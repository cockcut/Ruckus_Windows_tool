# -*- coding: utf-8 -*-
"""
HSITX Ruckus Technical Tool - GUI
PHP 포털과 유사한 메뉴 + CSV 샘플 다운로드/업로드 + AP 일괄 작업
"""

import os
import sys
import re
import csv
import shutil
import threading
import queue
import webbrowser
from datetime import datetime
from pathlib import Path
from tkinter import (
    Tk, Frame, Label, Button, Entry, Text, Scrollbar, Canvas, StringVar, BooleanVar,
    Toplevel, Checkbutton, LabelFrame, filedialog, messagebox, ttk, END, BOTH, X, Y, LEFT, RIGHT,
    TOP, BOTTOM, DISABLED, NORMAL, WORD, HORIZONTAL, VERTICAL,
)

# 로컬 모듈 (exe면 코드는 임시폴더, 결과/펌웨어는 exe 옆)
if getattr(sys, "frozen", False):
    ROOT = Path(sys.executable).resolve().parent
    sys.path.insert(0, str(Path(getattr(sys, "_MEIPASS", ROOT))))
else:
    ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))
from modules.ssh_helper import (
    process_ap,
    RuckusSSH,
    DEFAULT_USER,
    DEFAULT_PASSWORD,
    STANDARD_PASSWORD,
)
from modules.sz_api import (
    CONTROLLER_API_MAP,
    SmartZoneAPI,
    AP_ROW_FIELDS,
    ap_to_row,
    switch_to_row,
    save_csv,
)
from modules.unleashed_api import UnleashedAPI, save_csv as ul_save_csv
from modules.fw_builder import build_firmware_package, parse_bl7_name
from modules import updater as gh_updater
from modules.tftp_server import SimpleTftpServer
from modules.snmp_icx import query_icx

SAMPLES_DIR = ROOT / "samples"
UPLOAD_DIR = ROOT / "upload"
RESULTS_DIR = ROOT / "results"
RESULTS_AP_BATCH = RESULTS_DIR / "6_ap_batch"
RESULTS_SZ = RESULTS_DIR / "1_sz_api"
RESULTS_UNLEASHED = RESULTS_DIR / "2_unleashed_api"
RESULTS_OUI = RESULTS_DIR / "oui"
RESULTS_PSK = RESULTS_DIR / "3_ssid_psk"
RESULTS_FW = RESULTS_DIR / "7_fw_auto_upgrade"
RESULTS_SZFW = RESULTS_DIR / "9_fw_sz"
RESULTS_DPSK = RESULTS_DIR / "dpsk"
RESULTS_UDPSK = RESULTS_DIR / "dpsk_ul"
RESULTS_SNMP = RESULTS_DIR / "snmp"
FW_DIR = ROOT / "firmware"
LOG_DIR = ROOT / "log"
LOG_AP_BATCH = LOG_DIR / "6_ap_batch"
LOG_SZ = LOG_DIR / "1_sz_api"
LOG_UNLEASHED = LOG_DIR / "2_unleashed_api"
LOG_OUI = LOG_DIR / "oui"
LOG_PSK = LOG_DIR / "3_ssid_psk"
LOG_FW = LOG_DIR / "7_fw_auto_upgrade"
LOG_SZFW = LOG_DIR / "9_fw_sz"
LOG_DPSK = LOG_DIR / "dpsk"
LOG_UDPSK = LOG_DIR / "dpsk_ul"
LOG_SNMP = LOG_DIR / "snmp"
LOG_FWCLI = LOG_DIR / "8_fw_cli"
LOG_SINGLE = LOG_DIR / "12_single"
SAMPLES_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)
for _d in (
    RESULTS_DIR, RESULTS_AP_BATCH, RESULTS_SZ, RESULTS_UNLEASHED, RESULTS_OUI,
    RESULTS_PSK, RESULTS_FW, RESULTS_SZFW, RESULTS_DPSK, RESULTS_UDPSK, RESULTS_SNMP, FW_DIR,
    LOG_DIR, LOG_AP_BATCH, LOG_SZ, LOG_UNLEASHED, LOG_OUI, LOG_PSK, LOG_FW, LOG_SZFW,
    LOG_DPSK, LOG_UDPSK, LOG_SNMP, LOG_FWCLI, LOG_SINGLE,
):
    _d.mkdir(exist_ok=True)

RESULT_HINT = "결과파일은 최대 5개, 오래된순으로 삭제됨"
RESULT_KEEP = 5

# Ruckus AP 비밀번호 규칙
PW_SPECIAL = set(r"""~!@#$%^&*()-_=+[]{}|;:'",.<>/?""")
PW_RULE_KO = (
    "2차 기본계정 비밀번호 규칙\n"
    "• 8자 이상\n"
    "• 영문 대문자, 소문자, 숫자, 특수문자를 각각 1개 이상\n"
    "• 특수문자: ~ ! @ # $ % ^ & * ( ) - _ = + [ ] { } \\ | ; : ' \" , . < > / ?\n"
    "• ` 또는 $( 는 사용할 수 없음\n"
    "• 첫 글자는 ~ 일 수 없음"
)
PW_RULE_WARN = (
    "2차 비밀번호가 규칙에 맞지 않습니다.\n\n"
    + PW_RULE_KO
)
PW_COMPLEX_TIP = (
    "구형 펌웨어는 암호 복잡성을 지원하지 않을 수 있습니다.\n"
    "끄면 단순 비밀번호로도 2차 계정을 사용할 수 있습니다."
)


def validate_ruckus_password(pw: str):
    pw = pw or ""
    if len(pw) < 8:
        return False, "8자 이상이어야 합니다."
    if "`" in pw:
        return False, "` 문자는 사용할 수 없습니다."
    if "$(" in pw:
        return False, "$( 는 사용할 수 없습니다."
    if pw.startswith("~"):
        return False, "첫 글자는 ~ 일 수 없습니다."
    if not re.search(r"[a-z]", pw):
        return False, "영문 소문자가 1개 이상 필요합니다."
    if not re.search(r"[A-Z]", pw):
        return False, "영문 대문자가 1개 이상 필요합니다."
    if not re.search(r"[0-9]", pw):
        return False, "숫자가 1개 이상 필요합니다."
    if not any(c in PW_SPECIAL for c in pw):
        return False, "특수문자가 1개 이상 필요합니다."
    for c in pw:
        if not (c.isalnum() or c in PW_SPECIAL):
            return False, f"허용되지 않는 문자입니다: {c}"
    return True, ""


class BalloonTip:
    def __init__(self, widget, text: str, enabled: bool = True):
        self.widget = widget
        self.text = text
        self.enabled = enabled
        self.tip = None
        widget.bind("<Enter>", self._show)
        widget.bind("<Leave>", self._hide)

    def set_enabled(self, on: bool):
        self.enabled = bool(on)
        if not self.enabled:
            self._hide()

    def _show(self, _event=None):
        if self.tip or not self.enabled or not (self.text or "").strip():
            return
        x = self.widget.winfo_rootx() + 16
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 6
        self.tip = tw = Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        Label(
            tw, text=self.text, justify="left",
            bg="#fff8dc", fg="#333", relief="solid", borderwidth=1,
            font=("Segoe UI", 8), padx=8, pady=6,
        ).pack()

    def _hide(self, _event=None):
        if self.tip:
            self.tip.destroy()
            self.tip = None


def keep_latest_results(folder: Path, keep: int = RESULT_KEEP, suffixes=None):
    """Keep newest `keep` files. If suffixes given (e.g. ['_AP.csv']), keep that many of each type."""
    try:
        folder.mkdir(parents=True, exist_ok=True)
        for f in list(folder.glob("*.xml")):
            try:
                f.unlink()
            except Exception:
                pass
        groups = {}
        for f in folder.iterdir():
            if not f.is_file():
                continue
            key = "_all"
            if suffixes:
                hit = next((s for s in suffixes if f.name.endswith(s)), None)
                if not hit:
                    continue
                key = hit
            groups.setdefault(key, []).append(f)
        for files in groups.values():
            files.sort(key=lambda p: p.stat().st_mtime)
            extra = len(files) - keep
            if extra > 0:
                for f in files[:extra]:
                    try:
                        f.unlink()
                    except Exception:
                        pass
    except Exception:
        pass

# 버전 규칙:
#   - 소규모/버그픽스: 0.0.10p1, 0.0.10p2 ...
#   - 기능 추가·중규모: 0.0.11, 0.0.12 ...
#   - 대규모 구조 변경: 0.1.0, 0.2.0 ...
APP_VERSION = "0.0.1"
APP_TITLE = f"HSITX Ruckus Technical Tool v{APP_VERSION}"
BG = "#f4f4f9"
CARD = "#ffffff"
ACCENT = "#d9534f"
BTN_BG = "#f8f9fa"
BTN_ACTIVE = "#e2e6ea"
LINK = "#007bff"

OPERATIONS = [
    ("connect_sz", "SZ(SCG) IP 설정"),
    ("changeip", "AP IP 변경"),
    ("devicename", "Device Name 변경"),
    ("sz_devicename_changeip", "SZ + 호스트명 + IP 동시 변경"),
    ("reboot", "재부팅"),
    ("factory_reset", "공장 초기화 + 재부팅"),
]

# Excel이 구분자로 세미콜론을 쓰는 환경에서도 쉼표로 열리도록 sep=, 사용
SAMPLE_CSV_CONTENT = f"""sep=,
current_ip,username,password,new_ip,subnet,gateway,sz_ip,hostname
172.16.0.1,super,sp-admin,172.16.0.101,255.255.255.0,172.16.0.254,10.0.0.1,AP-Sample1
192.168.0.1,super,Ruckus!234,192.168.0.101,255.255.255.0,192.168.0.254,10.0.0.1,AP-Sample2
"""

SAMPLE_HEADERS = [
    "current_ip", "username", "password", "new_ip",
    "subnet", "gateway", "sz_ip", "hostname",
]
SAMPLE_ROWS = [
    ["172.16.0.1", "super", "sp-admin", "172.16.0.101",
     "255.255.255.0", "172.16.0.254", "10.0.0.1", "AP-Sample1"],
    ["192.168.0.1", "super", "Ruckus!234", "192.168.0.101",
     "255.255.255.0", "192.168.0.254", "10.0.0.1", "AP-Sample2"],
]


def _read_csv_text(path: str) -> str:
    """엑셀/메모장 저장 인코딩 자동 감지 (utf-8, cp949, euc-kr 등)"""
    raw = Path(path).read_bytes()
    # UTF-16 BOM
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        return raw.decode("utf-16")
    encodings = ("utf-8-sig", "utf-8", "cp949", "euc-kr", "utf-16", "latin-1")
    last_err = None
    for enc in encodings:
        try:
            return raw.decode(enc)
        except UnicodeDecodeError as e:
            last_err = e
            continue
    raise UnicodeDecodeError(
        "unknown", raw, 0, 1,
        f"지원 인코딩으로 읽을 수 없습니다: {last_err}",
    )


def _detect_delimiter(text: str) -> str:
    """쉼표/탭/세미콜론 자동 감지 (엑셀 저장 형식 대응)"""
    lines = []
    for line in text.splitlines():
        s = line.strip()
        if not s or s.startswith("#"):
            continue
        lines.append(line)
        if len(lines) >= 5:
            break
    sample = "\n".join(lines) if lines else text[:2048]
    if not sample.strip():
        return ","
    # 카운트 기반 (Sniffer가 실패하는 경우 대비)
    counts = {
        ",": sample.count(","),
        "\t": sample.count("\t"),
        ";": sample.count(";"),
    }
    # 탭이 많으면 TSV
    best = max(counts, key=counts.get)
    if counts[best] > 0:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
            return dialect.delimiter
        except Exception:
            return best
    return ","


def _split_row_fields(row: list) -> list:
    """한 칸에 전체가 들어온 경우(구분 실패) 재분리"""
    if len(row) == 1 and row[0]:
        cell = str(row[0])
        for sep in ("\t", ";", ","):
            if sep in cell:
                parts = [p.strip() for p in cell.split(sep)]
                if len(parts) >= 2:
                    return parts
    return row


def _row_to_dict(row: list, line_no: int):
    while len(row) < 8:
        row.append("")
    row = [str(c).strip().strip('"').strip("'") for c in row[:8]]
    if not row[0]:
        return None
    first = row[0].strip().lower()
    # sep=, 헤더, 주석 스킵
    if first.startswith("#") or first.startswith("sep="):
        return None
    if first in ("current_ip", "ip", "ap_ip"):
        return None
    ip = row[0].split()[0] if row[0] else ""
    if not ip or not any(ch.isdigit() for ch in ip):
        return None
    return {
        "ip": ip,
        "user": row[1] or DEFAULT_USER,
        "pass": row[2] or DEFAULT_PASSWORD,
        "new_ip": row[3],
        "subnet": row[4],
        "gw": row[5],
        "sz": row[6],
        "hostname": row[7],
        "line": line_no,
    }


def load_csv_rows(path: str) -> list:
    text = _read_csv_text(path)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # sep=, 가 있으면 강제 쉼표
    for line in text.splitlines()[:3]:
        if line.strip().lower().startswith("sep="):
            forced = line.split("=", 1)[-1].strip().strip('"')
            delim = forced if forced else ","
            break
    else:
        delim = _detect_delimiter(text)
    rows = []
    reader = csv.reader(text.splitlines(), delimiter=delim)
    for i, row in enumerate(reader, 1):
        if not row:
            continue
        row = _split_row_fields(row)
        item = _row_to_dict(row, i)
        if item:
            rows.append(item)
    return rows


def load_xlsx_rows(path: str) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for i, excel_row in enumerate(ws.iter_rows(values_only=True), 1):
        if not excel_row:
            continue
        row = ["" if c is None else str(c) for c in excel_row]
        item = _row_to_dict(row, i)
        if item:
            rows.append(item)
    wb.close()
    return rows


def load_table_file(path: str) -> list:
    """CSV 또는 Excel(.xlsx) 로드"""
    p = Path(path)
    suf = p.suffix.lower()
    if suf in (".xlsx", ".xlsm"):
        return load_xlsx_rows(path)
    return load_csv_rows(path)


class TextRedirector:
    """print 출력을 Text 위젯으로"""

    def __init__(self, q: queue.Queue):
        self.q = q

    def write(self, s):
        if s:
            self.q.put(s)

    def flush(self):
        pass


class App(Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1180x760")
        self.minsize(860, 640)
        self.configure(bg=BG)
        self._log_queue = queue.Queue()
        self._worker = None
        self._stop_flag = False
        self._csv_rows = []
        self._update_info = None
        self.alt_pw_complex = BooleanVar(value=True)

        self._build_main()
        self.after(100, self._drain_log)
        threading.Thread(target=self._check_github_update, daemon=True).start()

    def _clear_page(self):
        try:
            self.unbind_all("<MouseWheel>")
        except Exception:
            pass
        for w in self.winfo_children():
            w.destroy()
        try:
            self.geometry("1180x760")
        except Exception:
            pass

    def _scroll_page(self, padx=20, pady=16):
        """번호 메뉴 전체가 세로 스크롤되어 아래 로그가 가려지지 않게 함."""
        wrap = Frame(self, bg=BG)
        wrap.pack(fill=BOTH, expand=True)
        canvas = Canvas(wrap, bg=BG, highlightthickness=0)
        sb = Scrollbar(wrap, orient=VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        inner = Frame(canvas, bg=BG, padx=padx, pady=pady)
        win = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _cfg(_event=None):
            if not canvas.winfo_exists():
                return
            bbox = canvas.bbox("all")
            if bbox:
                canvas.configure(scrollregion=bbox)
            w = canvas.winfo_width()
            if w > 1:
                canvas.itemconfigure(win, width=w)

        inner.bind("<Configure>", _cfg)
        canvas.bind("<Configure>", _cfg)

        def _mw(event):
            if not canvas.winfo_exists():
                return
            canvas.yview_scroll(-1 if getattr(event, "delta", 0) > 0 else 1, "units")

        def _bind(_e=None):
            canvas.bind_all("<MouseWheel>", _mw)

        def _unbind(_e=None):
            try:
                canvas.unbind_all("<MouseWheel>")
            except Exception:
                pass

        inner.bind("<Enter>", _bind)
        inner.bind("<Leave>", _unbind)
        canvas.bind("<Enter>", _bind)
        return inner

    def _make_log(self, parent, title="실행 로그", height=8):
        """본문 아래 로그. 오른쪽 아래 손잡이를 세로로 드래그하면 높이만 조절."""
        Label(parent, text=title, font=("Segoe UI", 10, "bold"), bg=BG).pack(anchor="w", pady=(8, 2))
        box = Frame(parent, bg="#ced4da", highlightbackground="#adb5bd", highlightthickness=1)
        box.pack(fill=X)
        body = Frame(box, bg="#1e1e1e")
        body.pack(fill=BOTH, expand=True)
        self.log_text = Text(
            body, font=("Consolas", 9), height=height, wrap=WORD,
            bg="#1e1e1e", fg="#d4d4d4", insertbackground="white",
            relief="flat", padx=8, pady=8, borderwidth=0,
        )
        sb = Scrollbar(body, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=sb.set)
        self.log_text.pack(side=LEFT, fill=BOTH, expand=True)
        sb.pack(side=RIGHT, fill=Y)
        bar = Frame(box, bg="#f1f3f5", height=12)
        bar.pack(fill=X, side=BOTTOM)
        bar.pack_propagate(False)
        grip = Canvas(bar, width=22, height=12, bg="#f1f3f5", highlightthickness=0, cursor="sb_v_double_arrow")
        grip.pack(side=RIGHT, padx=4)
        grip.create_line(4, 9, 18, 9, fill="#868e96")
        grip.create_line(8, 6, 18, 6, fill="#868e96")
        grip.create_line(12, 3, 18, 3, fill="#868e96")
        st = {"y": 0, "h": height}

        def _press(e):
            st["y"] = e.y_root
            try:
                st["h"] = int(self.log_text.cget("height"))
            except Exception:
                st["h"] = height

        def _drag(e):
            dy = e.y_root - st["y"]
            lines = max(4, min(40, st["h"] + int(round(dy / 16.0))))
            self.log_text.configure(height=lines)

        for w in (grip, bar):
            w.bind("<ButtonPress-1>", _press)
            w.bind("<B1-Motion>", _drag)
            w.configure(cursor="sb_v_double_arrow")

    def _vscroll_pane(self, parent, width=None, bg=None):
        """왼쪽/오른쪽이 각각 세로 스크롤되는 패널."""
        bg = bg or BG
        wrap = Frame(parent, bg=bg)
        if width:
            wrap.configure(width=width)
            wrap.pack_propagate(False)
        canvas = Canvas(wrap, bg=bg, highlightthickness=0)
        if width:
            canvas.configure(width=max(40, width - 18))
        sb = Scrollbar(wrap, orient=VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side=RIGHT, fill=Y)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        inner = Frame(canvas, bg=bg)
        win = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _cfg(_event=None):
            if not canvas.winfo_exists():
                return
            bbox = canvas.bbox("all")
            if bbox:
                canvas.configure(scrollregion=bbox)
            w = canvas.winfo_width()
            if w > 1:
                canvas.itemconfigure(win, width=w)

        inner.bind("<Configure>", _cfg)
        canvas.bind("<Configure>", _cfg)

        def _mw(event):
            if not canvas.winfo_exists():
                return
            canvas.yview_scroll(-1 if getattr(event, "delta", 0) > 0 else 1, "units")
            return "break"

        def _on(_e=None):
            canvas.bind_all("<MouseWheel>", _mw)

        def _off(_e=None):
            try:
                canvas.unbind_all("<MouseWheel>")
            except Exception:
                pass

        canvas.bind("<Enter>", _on)
        canvas.bind("<Leave>", _off)
        inner.bind("<Enter>", _on)
        inner.bind("<Leave>", _off)
        return wrap, inner

    # ------------------------------------------------------------------
    # 메인 메뉴
    # ------------------------------------------------------------------
    def _build_main(self):
        self._clear_page()

        outer = Frame(self, bg=BG, padx=24, pady=20)
        outer.pack(fill=BOTH, expand=True)

        Label(
            outer, text=APP_TITLE, font=("Segoe UI", 18, "bold"),
            fg=ACCENT, bg=BG,
        ).pack(anchor="w")
        Label(
            outer,
            text=f"버전 {APP_VERSION}  |  비밀번호 정책: CSV Username, Password → 2차 기본계정",
            font=("Segoe UI", 9), fg="#666", bg=BG,
        ).pack(anchor="w", pady=(4, 8))
        upd_row = Frame(outer, bg=BG)
        upd_row.pack(anchor="w", pady=(0, 12))
        self._upd_check_btn = Button(
            upd_row, text="업데이트 확인", font=("Segoe UI", 9, "bold"),
            bg="#fff", fg=ACCENT, relief="solid", borderwidth=1,
            highlightbackground=ACCENT, padx=10, pady=2,
            command=self._manual_github_check, cursor="hand2",
        )
        self._upd_check_btn.pack(side=LEFT)
        self._upd_btn = Button(
            upd_row, text="업데이트", font=("Segoe UI", 9, "bold"),
            bg=ACCENT, fg="white", relief="flat", padx=12, pady=2,
            command=self._do_github_update, cursor="hand2",
        )
        self._upd_status = StringVar(value="")
        Label(upd_row, textvariable=self._upd_status, font=("Segoe UI", 9), fg="#666", bg=BG).pack(side=LEFT, padx=(10, 0))
        info = getattr(self, "_update_info", None) or {}
        if info.get("available"):
            if info.get("frozen"):
                self._upd_status.set("GitHub에 새 exe가 있습니다.")
            else:
                self._upd_status.set("GitHub에 새 버전이 있습니다.")
            self._upd_btn.pack(side=LEFT, padx=(10, 0))
        elif info.get("ok"):
            self._upd_status.set("최신 버전입니다.")
            self._upd_btn.pack_forget()
        elif info.get("message"):
            self._upd_status.set(str(info.get("message")))
            self._upd_btn.pack_forget()
        else:
            self._upd_btn.pack_forget()

        wrap = Frame(outer, bg=BG)
        wrap.pack(fill=BOTH, expand=True)
        canvas = Canvas(wrap, bg=CARD, highlightthickness=0)
        sb = Scrollbar(wrap, orient=VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        sb.pack(side=RIGHT, fill=Y)
        card = Frame(canvas, bg=CARD, padx=20, pady=16)
        win = canvas.create_window((0, 0), window=card, anchor="nw")

        def _main_cfg(_e=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfigure(win, width=canvas.winfo_width())
        card.bind("<Configure>", _main_cfg)
        canvas.bind("<Configure>", _main_cfg)

        def _mw(event):
            if not canvas.winfo_exists():
                try:
                    canvas.unbind_all("<MouseWheel>")
                except Exception:
                    pass
                return
            canvas.yview_scroll(-1 if getattr(event, "delta", 0) > 0 else 1, "units")
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _mw))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        menus = [
            ("1", "SmartZone 정보 보기", True),
            ("2", "언리시드 정보 보기", True),
            ("3", "PSK/SAE 패스워드 일괄변경 (SmartZone)", True),
            ("4", "SmartZone DPSK 관리", True),
            ("5", "Unleashed DPSK 관리", True),
            ("6", "AP IP 일괄 변경 / SZ 연동 / 공장초기화", True),
            ("7", "AP 펌웨어 자동 업그레이드", True),
            ("8", "AP 펌웨어 CLI 명령어 생성", True),
            ("9", "AP → SZ 펌웨어 업그레이드 + 연동", True),
            ("10", "ICX ARP 조회 (SNMP)", True),
            ("11", "OUI 조회", True),
            ("12", "단일 AP SSH 테스트", True),
        ]

        for num, title, enabled in menus:
            self._menu_btn(card, f"[ {num}. {title} ]", enabled, num)

        Label(
            outer,
            text="★ 표시 기능만 현재 동작합니다. 나머지는 추후 확장 가능합니다.",
            font=("Segoe UI", 8), fg="#888", bg=BG,
        ).pack(anchor="w", pady=(12, 0))

    def _menu_btn(self, parent, text, enabled, num):
        star = "  ★" if enabled else ""
        b = Button(
            parent,
            text=text + star,
            font=("Segoe UI", 11),
            bg=BTN_BG if enabled else "#eee",
            fg=LINK if enabled else "#999",
            activebackground=BTN_ACTIVE,
            relief="solid",
            borderwidth=1,
            anchor="w",
            padx=14, pady=8,
            cursor="hand2" if enabled else "arrow",
            command=(lambda n=num: self._on_menu(n)) if enabled else None,
            state=NORMAL if enabled else DISABLED,
        )
        b.pack(fill=X, pady=4)

    def _manual_github_check(self):
        if hasattr(self, "_upd_status"):
            self._upd_status.set("GitHub 확인 중...")
        if hasattr(self, "_upd_check_btn"):
            self._upd_check_btn.config(state=DISABLED)
        threading.Thread(target=self._check_github_update, daemon=True).start()

    def _check_github_update(self):
        frozen = bool(getattr(sys, "frozen", False))
        info = gh_updater.check_update(ROOT, frozen=frozen, current_version=APP_VERSION)
        self._update_info = info
        self.after(0, lambda: self._show_update_ui(info))

    def _show_update_ui(self, info: dict):
        if hasattr(self, "_upd_check_btn"):
            try:
                self._upd_check_btn.config(state=NORMAL)
            except Exception:
                pass
        if not info or not hasattr(self, "_upd_status"):
            return
        if not info.get("ok"):
            self._upd_status.set(info.get("message") or "업데이트 확인 실패")
            if hasattr(self, "_upd_btn"):
                self._upd_btn.pack_forget()
            return
        if info.get("available"):
            if info.get("frozen"):
                self._upd_status.set("GitHub에 새 exe가 있습니다.")
            else:
                self._upd_status.set("GitHub에 새 버전이 있습니다.")
            if hasattr(self, "_upd_btn") and not self._upd_btn.winfo_ismapped():
                self._upd_btn.pack(side=LEFT, padx=(10, 0))
        else:
            self._upd_status.set("최신 버전입니다.")
            if hasattr(self, "_upd_btn"):
                self._upd_btn.pack_forget()

    def _do_github_update(self):
        info = getattr(self, "_update_info", None) or {}
        frozen = bool(getattr(sys, "frozen", False))
        if frozen:
            msg = "GitHub에서 새 exe를 받아 지금 실행 파일을 교체할까요?"
        else:
            msg = "GitHub에서 최신 소스를 받아 덮어쓸까요?\nresults / upload / firmware .bl7 은 유지됩니다."
        if not messagebox.askyesno("업데이트", msg):
            return
        self._upd_status.set("업데이트 받는 중...")
        self._upd_btn.config(state=DISABLED)

        def work():
            result = gh_updater.apply_update(
                ROOT,
                expected_sha=info.get("remote") or "",
                frozen=frozen,
                exe_path=sys.executable if frozen else "",
                info=info,
            )
            self.after(0, lambda: self._after_github_update(result))

        threading.Thread(target=work, daemon=True).start()

    def _after_github_update(self, result: dict):
        if hasattr(self, "_upd_btn"):
            self._upd_btn.config(state=NORMAL)
        if result.get("ok"):
            bat = result.get("replace_bat")
            if bat:
                messagebox.showinfo("완료", result.get("message") or "exe 업데이트 준비됨")
                try:
                    import subprocess
                    env = os.environ.copy()
                    for k in list(env):
                        if k.startswith("_PYI") or k in ("PYTHONHOME", "PYTHONPATH"):
                            env.pop(k, None)
                    subprocess.Popen(
                        ["cmd", "/c", bat],
                        cwd=str(ROOT),
                        env=env,
                        close_fds=True,
                    )
                except Exception as e:
                    messagebox.showerror("업데이트", f"교체 스크립트 실행 실패:\n{e}")
                    return
            else:
                messagebox.showinfo("완료", result.get("message") or "업데이트 완료")
            self.destroy()
        else:
            if hasattr(self, "_upd_status"):
                self._upd_status.set("업데이트 실패")
            messagebox.showerror("업데이트 실패", result.get("message") or "실패")

    def _on_menu(self, num):
        if num == "1":
            self._build_sz()
        elif num == "2":
            self._build_unleashed()
        elif num == "3":
            self._build_psk()
        elif num == "4":
            self._build_dpsk()
        elif num == "5":
            self._build_uldpsk()
        elif num == "7":
            self._build_fwup()
        elif num == "8":
            self._build_fw()
        elif num == "9":
            self._build_fwsz()
        elif num == "10":
            self._build_icx()
        elif num == "6":
            self._build_ap_batch()
        elif num == "11":
            self._build_oui()
        elif num == "12":
            self._build_single_test()
        else:
            messagebox.showinfo("안내", "이 기능은 아직 GUI로 구현되지 않았습니다.")

    def _back_btn(self, parent):
        Button(
            parent, text="← 메인 메뉴", font=("Segoe UI", 10),
            bg=BTN_BG, relief="solid", borderwidth=1, padx=10, pady=4,
            command=self._build_main, cursor="hand2",
        ).pack(anchor="w", pady=(0, 10))

    # ------------------------------------------------------------------
    # 메뉴 6: AP 일괄 작업
    # ------------------------------------------------------------------
    def _build_ap_batch(self):
        self._clear_page()

        outer = self._scroll_page(padx=20, pady=16)

        self._back_btn(outer)
        Label(
            outer, text="6. AP 일괄 변경 / SZ 연동 / 공장초기화",
            font=("Segoe UI", 14, "bold"), fg=ACCENT, bg=BG,
        ).pack(anchor="w")

        # --- 파일 영역 ---
        file_fr = Frame(outer, bg=CARD, padx=12, pady=10,
                        highlightbackground="#dee2e6", highlightthickness=1)
        file_fr.pack(fill=X, pady=(10, 8))

        head = Frame(file_fr, bg=CARD)
        head.pack(fill=X)
        left_csv = Frame(head, bg=CARD)
        left_csv.pack(side=LEFT, anchor="n")
        Label(left_csv, text="1) CSV 파일", font=("Segoe UI", 10, "bold"), bg=CARD).pack(anchor="w")
        acc2 = LabelFrame(head, text="2차 기본 계정 (초기 비번 변경시 or CSV 실패시)", bg=CARD, fg="#444",
                          font=("Segoe UI", 8), padx=10, pady=6)
        acc2.pack(side=RIGHT, anchor="n")
        self.ap_user2 = StringVar(value="")
        self.ap_pass2 = StringVar(value="")
        Label(acc2, text="Username", width=10, anchor="w", bg=CARD, font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w", pady=2)
        Entry(acc2, textvariable=self.ap_user2, width=16, font=("Segoe UI", 10)).grid(row=0, column=1, sticky="w", pady=2)
        Label(acc2, text="Password", width=10, anchor="w", bg=CARD, font=("Segoe UI", 10)).grid(row=1, column=0, sticky="w", pady=2)
        self.ap_pass2_entry = Entry(acc2, textvariable=self.ap_pass2, width=16, font=("Segoe UI", 10), show="*")
        self.ap_pass2_entry.grid(row=1, column=1, sticky="w", pady=2)
        self.ap_pass2_tip = BalloonTip(self.ap_pass2_entry, PW_RULE_KO, enabled=self._alt_pw_complex_on())
        self.ap_pass2_show = BooleanVar(value=False)
        Checkbutton(acc2, text="표시", variable=self.ap_pass2_show, bg=CARD,
                    command=lambda: self.ap_pass2_entry.config(show="" if self.ap_pass2_show.get() else "*")).grid(row=1, column=2, sticky="w", padx=(6, 0))
        self._alt_pw_hint_row(acc2)

        row1 = Frame(left_csv, bg=CARD)
        row1.pack(fill=X, pady=6)
        Button(
            row1, text="샘플 다운로드 (CSV)", font=("Segoe UI", 10),
            bg="#28a745", fg="white", relief="flat", padx=12, pady=4,
            command=self._download_sample, cursor="hand2",
        ).pack(side=LEFT, padx=(0, 8))
        Button(
            row1, text="파일 업로드 (CSV)", font=("Segoe UI", 10),
            bg=LINK, fg="white", relief="flat", padx=12, pady=4,
            command=self._browse_csv, cursor="hand2",
        ).pack(side=LEFT)

        self.csv_path_var = StringVar(value="")
        Label(
            file_fr, textvariable=self.csv_path_var, font=("Segoe UI", 9),
            fg="#333", bg=CARD, wraplength=800, justify="left",
        ).pack(anchor="w", pady=(4, 0))

        self.csv_info_var = StringVar(value="업로드된 파일 없음")
        Label(
            file_fr, textvariable=self.csv_info_var, font=("Segoe UI", 9),
            fg="#666", bg=CARD,
        ).pack(anchor="w")

        # --- CSV 미리보기 ---
        Label(file_fr, text="CSV 미리보기", font=("Segoe UI", 9, "bold"), bg=CARD).pack(anchor="w", pady=(8, 2))
        prev_fr = Frame(file_fr, bg=CARD)
        prev_fr.pack(fill=X)
        cols = ("ip", "user", "pass", "new_ip", "subnet", "gw", "sz", "hostname")
        self.csv_tree = ttk.Treeview(
            prev_fr, columns=cols, show="headings", height=5, selectmode="browse",
        )
        headings = {
            "ip": "current_ip",
            "user": "username",
            "pass": "password",
            "new_ip": "new_ip",
            "subnet": "subnet",
            "gw": "gateway",
            "sz": "sz_ip",
            "hostname": "hostname",
        }
        widths = {"ip": 110, "user": 70, "pass": 90, "new_ip": 110, "subnet": 110, "gw": 100, "sz": 100, "hostname": 100}
        for c in cols:
            self.csv_tree.heading(c, text=headings[c])
            self.csv_tree.column(c, width=widths[c], anchor="w")
        prev_ys = Scrollbar(prev_fr, orient=VERTICAL, command=self.csv_tree.yview)
        prev_xs = Scrollbar(prev_fr, orient=HORIZONTAL, command=self.csv_tree.xview)
        self.csv_tree.configure(yscrollcommand=prev_ys.set, xscrollcommand=prev_xs.set)
        self.csv_tree.grid(row=0, column=0, sticky="nsew")
        prev_ys.grid(row=0, column=1, sticky="ns")
        prev_xs.grid(row=1, column=0, sticky="ew")
        prev_fr.grid_rowconfigure(0, weight=1)
        prev_fr.grid_columnconfigure(0, weight=1)

        # --- 작업 선택 ---
        op_fr = Frame(outer, bg=CARD, padx=12, pady=10,
                      highlightbackground="#dee2e6", highlightthickness=1)
        op_fr.pack(fill=X, pady=8)

        Label(op_fr, text="2) 작업 선택", font=("Segoe UI", 10, "bold"), bg=CARD).pack(anchor="w")
        self.op_var = StringVar(value=OPERATIONS[0][1])
        op_row = Frame(op_fr, bg=CARD)
        op_row.pack(fill=X, pady=6)
        self.op_combo = ttk.Combobox(
            op_row,
            textvariable=self.op_var,
            values=[v for _k, v in OPERATIONS],
            state="readonly",
            width=50,
            font=("Segoe UI", 10),
        )
        self.op_combo.current(0)
        self.op_combo.pack(side=LEFT)

        # --- 실행 ---
        run_fr = Frame(outer, bg=BG)
        run_fr.pack(fill=X, pady=8)
        self.run_btn = Button(
            run_fr, text="▶ 실행", font=("Segoe UI", 11, "bold"),
            bg=ACCENT, fg="white", relief="flat", padx=24, pady=6,
            command=self._start_batch, cursor="hand2",
        )
        self.run_btn.pack(side=LEFT, padx=(0, 8))
        self.stop_btn = Button(
            run_fr, text="중지", font=("Segoe UI", 10),
            bg="#6c757d", fg="white", relief="flat", padx=16, pady=6,
            command=self._stop_batch, state=DISABLED, cursor="hand2",
        )
        self.stop_btn.pack(side=LEFT, padx=(0, 8))
        Button(
            run_fr, text="결과 폴더 열기", font=("Segoe UI", 10),
            bg=BTN_BG, relief="solid", borderwidth=1, padx=12, pady=5,
            command=lambda: self._open_path(RESULTS_AP_BATCH), cursor="hand2",
        ).pack(side=LEFT)
        self._latest_result_btn(run_fr, RESULTS_AP_BATCH)
        self._log_action_btns(run_fr, LOG_AP_BATCH, bg=BG)
        Label(run_fr, text=RESULT_HINT, font=("Segoe UI", 8), fg="#888", bg=BG).pack(side=LEFT, padx=(8, 0))

        self.progress = ttk.Progressbar(outer, mode="determinate")
        self.progress.pack(fill=X, pady=(4, 4))
        self.status_var = StringVar(value="대기 중")
        Label(outer, textvariable=self.status_var, font=("Segoe UI", 9), bg=BG, fg="#555").pack(anchor="w")
        self._make_log(outer, "실행 로그", 8)
        self._csv_rows = []

    def _download_sample(self):
        """쉼표(,) 구분 CSV. 첫 줄 sep=, 은 엑셀이 세미콜론으로 열지 않게 함."""
        SAMPLES_DIR.mkdir(exist_ok=True)
        path = filedialog.asksaveasfilename(
            title="샘플 CSV 저장",
            defaultextension=".csv",
            initialfile="ap_sample.csv",
            initialdir=str(Path.home() / "Documents"),
            filetypes=[
                ("CSV files", "*.csv"),
                ("All files", "*.*"),
            ],
        )
        if not path:
            return
        try:
            p = Path(path)
            if p.suffix.lower() != ".csv":
                p = p.with_suffix(".csv")
            with open(p, "w", encoding="utf-8-sig", newline="") as f:
                f.write(SAMPLE_CSV_CONTENT)
            with open(SAMPLES_DIR / "ap_sample.csv", "w", encoding="utf-8-sig", newline="") as f:
                f.write(SAMPLE_CSV_CONTENT)
            messagebox.showinfo(
                "완료",
                f"샘플 CSV 저장:\n{p}\n\n"
                "열 구분은 쉼표(,) 입니다.\n"
                "첫 줄 sep=, 은 엑셀용 힌트이므로 삭제하지 마세요.\n"
                "엑셀에서 다른 이름 저장할 때도 CSV UTF-8 로 저장하세요.",
            )
        except Exception as e:
            messagebox.showerror("오류", str(e))

    def _browse_csv(self):
        path = filedialog.askopenfilename(
            title="AP 리스트 CSV 선택",
            filetypes=[
                ("CSV files", "*.csv"),
                ("All files", "*.*"),
            ],
            initialdir=str(SAMPLES_DIR) if SAMPLES_DIR.is_dir() else str(Path.home()),
        )
        if not path:
            return
        if Path(path).suffix.lower() != ".csv":
            messagebox.showwarning("안내", "CSV 파일만 업로드할 수 있습니다.")
            return
        try:
            rows = load_csv_rows(path)
            if not rows:
                messagebox.showwarning(
                    "경고",
                    "유효한 AP 행이 없습니다.\n"
                    "헤더 행을 제외하고 current_ip 열이 IP로 되어 있는지 확인하세요.",
                )
                return

            UPLOAD_DIR.mkdir(exist_ok=True)
            src = Path(path)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            dest_name = f"{src.stem}_{ts}{src.suffix}"
            dest = UPLOAD_DIR / dest_name
            shutil.copy2(src, dest)

            self._csv_rows = rows
            self._csv_upload_path = str(dest)
            self.csv_path_var.set(str(dest))
            self.csv_info_var.set(f"업로드됨: {len(rows)} 대 → upload/{dest_name}")
            self._fill_csv_preview(rows)
            self._log(f"파일 업로드: {src} → {dest} ({len(rows)} rows)\n")
        except Exception as e:
            messagebox.showerror("오류", f"파일 업로드 실패:\n{e}")

    def _fill_csv_preview(self, rows: list):
        if not hasattr(self, "csv_tree"):
            return
        self.csv_tree.delete(*self.csv_tree.get_children())
        for r in rows:
            self.csv_tree.insert(
                "",
                END,
                values=(
                    r.get("ip", ""),
                    r.get("user", ""),
                    r.get("pass", ""),
                    r.get("new_ip", ""),
                    r.get("subnet", ""),
                    r.get("gw", ""),
                    r.get("sz", ""),
                    r.get("hostname", ""),
                ),
            )

    def _selected_operation(self) -> str:
        val = (self.op_combo.get() or "").strip()
        for k, v in OPERATIONS:
            if val == v or val == k or val.startswith(f"{k} "):
                return k
        idx = self.op_combo.current()
        if 0 <= idx < len(OPERATIONS):
            return OPERATIONS[idx][0]
        return OPERATIONS[0][0]

    def _require_second_account(self, user_var, pass_var) -> bool:
        u = (user_var.get() if user_var is not None else "").strip()
        p = (pass_var.get() if pass_var is not None else "").strip()
        if not u or not p:
            msg = "2차 기본계정을 입력안하면 초기화된 AP에 접속이 불가합니다. 입력해주세요."
            if self._alt_pw_complex_on():
                msg += "\n\n" + PW_RULE_KO
            messagebox.showwarning("안내", msg)
            return False
        if self._alt_pw_complex_on():
            ok, reason = validate_ruckus_password(p)
            if not ok:
                messagebox.showwarning(
                    "2차 기본계정 비밀번호 규칙",
                    f"{reason}\n\n{PW_RULE_KO}",
                )
                return False
        return True

    def _sync_alt_pw_tips(self):
        on = self._alt_pw_complex_on()
        for name in ("ap_pass2_tip", "fwup_pass2_tip", "fwsz_pass2_tip"):
            tip = getattr(self, name, None)
            if tip:
                tip.set_enabled(on)

    def _alt_pw_complex_on(self) -> bool:
        var = getattr(self, "alt_pw_complex", None)
        try:
            return bool(var.get()) if var is not None else True
        except Exception:
            return True

    def _alt_pw_hint_row(self, parent):
        """힌트(작은 글씨) + 표시 아래 복잡성 체크."""
        Label(
            parent,
            text="강제변경 비밀번호 / CSV 실패 시 계정",
            font=("Segoe UI", 8), bg=CARD, fg="#888",
        ).grid(row=2, column=0, columnspan=2, sticky="w")
        cb = Checkbutton(
            parent, text="복잡성", variable=self.alt_pw_complex, bg=CARD,
            font=("Segoe UI", 8), command=self._sync_alt_pw_tips,
        )
        cb.grid(row=2, column=2, sticky="w", padx=(6, 0))
        BalloonTip(cb, PW_COMPLEX_TIP)

    def _start_batch(self):
        if self._worker and self._worker.is_alive():
            messagebox.showwarning("안내", "이미 실행 중입니다.")
            return
        if not self._csv_rows:
            messagebox.showwarning("안내", "먼저 CSV를 업로드하세요.")
            return
        if not self._require_second_account(
            getattr(self, "ap_user2", None), getattr(self, "ap_pass2", None)
        ):
            return
        op = self._selected_operation()
        self._stop_flag = False
        self.run_btn.config(state=DISABLED)
        self.stop_btn.config(state=NORMAL)
        self.progress["value"] = 0
        self.progress["maximum"] = len(self._csv_rows)
        self.status_var.set(f"실행 중: {op} (0/{len(self._csv_rows)})")
        self._log(f"\n===== 작업 시작: {op} / {len(self._csv_rows)} 대 =====\n")

        rows = list(self._csv_rows)
        self._worker = threading.Thread(
            target=self._batch_worker, args=(rows, op), daemon=True,
        )
        self._worker.start()

    def _stop_batch(self):
        self._stop_flag = True
        self._log("\n[중지 요청] 현재 AP 작업 후 중단합니다...\n")

    def _batch_worker(self, rows, operation):
        results = []
        old_stdout = sys.stdout
        sys.stdout = TextRedirector(self._log_queue)
        try:
            for i, row in enumerate(rows, 1):
                if self._stop_flag:
                    self._log_queue.put(f"\n사용자 중지로 종료 ({i - 1}/{len(rows)} 완료)\n")
                    break
                self._log_queue.put(f"\n[{i}/{len(rows)}] {row['ip']}\n")
                r = process_ap(
                    ip=row["ip"],
                    user=row["user"],
                    password=row["pass"],
                    operation=operation,
                    new_ip=row["new_ip"],
                    subnet=row["subnet"],
                    gw=row["gw"],
                    sz=row["sz"],
                    hostname=row["hostname"],
                    new_password=(self.ap_pass2.get() if hasattr(self, "ap_pass2") else ""),
                    standard_password=(self.ap_pass2.get() if hasattr(self, "ap_pass2") else ""),
                    fallback_user=(self.ap_user2.get() if hasattr(self, "ap_user2") else ""),
                    debug=True,
                )
                results.append(r)
                self.after(0, lambda v=i: self._update_progress(v, len(rows), operation))
        except Exception as e:
            self._log_queue.put(f"\n오류: {e}\n")
        finally:
            sys.stdout = old_stdout

        # 결과 저장
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out = RESULTS_AP_BATCH / f"result_{operation}_{ts}.csv"
            with open(out, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                w.writerow(["ip", "status", "message", "serial", "mac", "user", "password"])
                for r in results:
                    w.writerow([
                        r.get("ip", ""), r.get("status", ""), r.get("message", ""),
                        r.get("serial", ""), r.get("mac", ""),
                        r.get("user", ""), r.get("password", ""),
                    ])
            keep_latest_results(RESULTS_AP_BATCH)
            self._save_session_log(LOG_AP_BATCH, f"batch_{operation}")
            ok_n = sum(1 for r in results if r.get("status") == "OK")
            self._log_queue.put(f"\n===== 완료: 성공 {ok_n}/{len(results)} =====\n결과 파일: {out}\n")
        except Exception as e:
            self._log_queue.put(f"\n결과 저장 실패: {e}\n")

        self.after(0, self._batch_done)

    def _update_progress(self, current, total, op):
        self.progress["value"] = current
        self.status_var.set(f"실행 중: {op} ({current}/{total})")

    def _batch_done(self):
        self.run_btn.config(state=NORMAL)
        self.stop_btn.config(state=DISABLED)
        self.status_var.set("완료")
        messagebox.showinfo("완료", "작업이 끝났습니다.\n로그와 결과 CSV를 확인하세요.")

    def _open_results(self):
        path = str(RESULTS_DIR.resolve())
        try:
            os.startfile(path)  # Windows
        except Exception:
            messagebox.showinfo("결과 폴더", path)

    # ------------------------------------------------------------------
    # 메뉴 3: PSK/SAE 패스워드 일괄변경 (원본 psk/index.php 흐름)
    # ------------------------------------------------------------------
    def _build_psk(self):
        self._clear_page()

        outer = self._scroll_page(padx=20, pady=16)
        self._back_btn(outer)

        Label(
            outer, text="3. PSK/SAE 패스워드 일괄변경 (SmartZone)",
            font=("Segoe UI", 14, "bold"), fg=ACCENT, bg=BG,
        ).pack(anchor="w")
        Label(
            outer,
            text="Zone 조회 → WLAN 상세 조회 → Passphrase 변경",
            font=("Segoe UI", 9), fg="#666", bg=BG,
        ).pack(anchor="w", pady=(2, 8))

        form = Frame(outer, bg=CARD, padx=12, pady=8,
                     highlightbackground="#dee2e6", highlightthickness=1)
        form.pack(fill=X)
        form_left = Frame(form, bg=CARD)
        form_left.pack(side=LEFT, anchor="n")
        form_right = Frame(form, bg=CARD)
        form_right.pack(side=LEFT, anchor="n", padx=(16, 0))
        self.psk_hint_zone = StringVar(value="")
        self.psk_hint_wlan = StringVar(value="")
        self.psk_hint_pw = StringVar(value="")

        self.psk_ip = StringVar()
        self.psk_user = StringVar(value="admin")
        self.psk_pass = StringVar()
        self.psk_ctrl = StringVar(value="수동선택")
        self.psk_api = StringVar()
        self.psk_zone = StringVar(value="모든 Zone")
        self.psk_method = StringVar(value="WPA2")
        self.psk_new = StringVar()

        def row(parent, lbl, var, show=None):
            fr = Frame(parent, bg=CARD)
            fr.pack(fill=X, pady=1)
            Label(fr, text=lbl, width=14, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
            Entry(fr, textvariable=var, font=("Segoe UI", 10), width=20, show=show or "").pack(side=LEFT)

        row(form_left, "SZ IP/Domain", self.psk_ip)
        row(form_left, "Username", self.psk_user)
        row(form_left, "Password", self.psk_pass, show="*")

        fr_c = Frame(form_left, bg=CARD)
        fr_c.pack(fill=X, pady=1)
        Label(fr_c, text="컨트롤러 버전", width=14, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        self.psk_ctrl_combo = ttk.Combobox(
            fr_c, textvariable=self.psk_ctrl,
            values=list(CONTROLLER_API_MAP.keys()),
            state="readonly", width=18, font=("Segoe UI", 10),
        )
        self.psk_ctrl_combo.pack(side=LEFT)
        self.psk_ctrl_combo.bind("<<ComboboxSelected>>", self._psk_update_api_versions)

        fr_a = Frame(form_left, bg=CARD)
        fr_a.pack(fill=X, pady=1)
        Label(fr_a, text="API 버전", width=14, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        self.psk_api_combo = ttk.Combobox(
            fr_a, textvariable=self.psk_api, state="readonly", width=18, font=("Segoe UI", 10),
        )
        self.psk_api_combo.pack(side=LEFT)
        self._psk_update_api_versions()

        fr_z = Frame(form_left, bg=CARD)
        fr_z.pack(fill=X, pady=1)
        Label(fr_z, text="Zone", width=14, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        self.psk_zone_combo = ttk.Combobox(
            fr_z, textvariable=self.psk_zone, values=["모든 Zone"],
            state="readonly", width=18, font=("Segoe UI", 10),
        )
        self.psk_zone_combo.pack(side=LEFT)

        def same_btn(parent, text, cmd, bg, fg="white", bold=False, relief="flat", padx=0):
            wrap = Frame(parent, bg=CARD, width=220, height=34)
            wrap.pack(side=LEFT, pady=0, padx=padx)
            wrap.pack_propagate(False)
            font = ("Segoe UI", 10, "bold") if bold else ("Segoe UI", 10)
            b = Button(
                wrap, text=text, font=font, bg=bg, fg=fg, relief=relief,
                borderwidth=0 if relief == "flat" else 1,
                command=cmd, cursor="hand2",
            )
            b.pack(fill=BOTH, expand=True)
            return b

        zrow = Frame(form_right, bg=CARD)
        zrow.pack(anchor="w", pady=2)
        self.psk_zone_btn = same_btn(zrow, "Zone 목록 가져오기", self._psk_load_zones, LINK)
        Label(zrow, textvariable=self.psk_hint_zone, font=("Segoe UI", 9), bg=CARD, fg="#d9534f").pack(side=LEFT, padx=(8, 0))

        wrow = Frame(form_right, bg=CARD)
        wrow.pack(anchor="w", pady=2)
        self.psk_list_btn = same_btn(wrow, "WLAN 상세 가져오기", self._psk_list, LINK, bold=True)
        Label(wrow, textvariable=self.psk_hint_wlan, font=("Segoe UI", 9), bg=CARD, fg="#d9534f").pack(side=LEFT, padx=(8, 0))

        rrow = Frame(form_right, bg=CARD)
        rrow.pack(anchor="w", pady=2)
        same_btn(rrow, "결과 폴더 열기", lambda: self._open_path(RESULTS_PSK), BTN_BG, fg="#333", relief="solid")
        same_btn(rrow, "최근 결과 다운로드", lambda: self._download_latest_results(RESULTS_PSK), "#28a745")
        self._log_action_btns(rrow, LOG_PSK, bg=CARD)
        Label(rrow, text=RESULT_HINT, font=("Segoe UI", 8), fg="#888", bg=CARD).pack(side=LEFT, padx=(8, 0))

        chg = Frame(outer, bg=CARD, padx=12, pady=6,
                    highlightbackground="#dee2e6", highlightthickness=1)
        chg.pack(fill=X, pady=(6, 4))
        chg_left = Frame(chg, bg=CARD)
        chg_left.pack(side=LEFT, anchor="n")
        chg_right = Frame(chg, bg=CARD)
        chg_right.pack(side=LEFT, anchor="n", padx=(16, 0))
        self.psk_hint2 = StringVar(value="")
        hint2 = Frame(chg, bg=CARD)
        hint2.pack(side=LEFT, anchor="n", padx=(10, 0))
        Label(hint2, textvariable=self.psk_hint2, font=("Segoe UI", 9), bg=CARD, fg="#d9534f", wraplength=220, justify="left").pack(anchor="w")
        Label(chg_left, text="암호 변경 (PSK/SAE만)", font=("Segoe UI", 10, "bold"), bg=CARD).pack(anchor="w")
        fr_m = Frame(chg_left, bg=CARD)
        fr_m.pack(fill=X, pady=2)
        Label(fr_m, text="암호 방식", width=14, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        ttk.Combobox(
            fr_m, textvariable=self.psk_method, values=["WPA2", "WPA3", "WPA23_Mixed"],
            state="readonly", width=18, font=("Segoe UI", 10),
        ).pack(side=LEFT)
        fr_p = Frame(chg_left, bg=CARD)
        fr_p.pack(fill=X, pady=2)
        Label(fr_p, text="새 비밀번호", width=14, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        Entry(fr_p, textvariable=self.psk_new, font=("Segoe UI", 10), width=20, show="*").pack(side=LEFT)
        Label(
            chg_left, textvariable=self.psk_hint_pw, font=("Segoe UI", 9),
            bg=CARD, fg="#d9534f",
        ).pack(anchor="w", pady=(8, 0))
        brow1 = Frame(chg_right, bg=CARD)
        brow1.pack(anchor="w", pady=2)
        self.psk_run_btn = same_btn(brow1, "개별 WLAN 암호변경", lambda: self._psk_apply(False), ACCENT, bold=True)
        self.psk_all_btn = same_btn(brow1, "일괄 WLAN 암호변경", lambda: self._psk_apply(True), "#28a745", bold=True, padx=(10, 0))
        brow3 = Frame(chg_right, bg=CARD)
        brow3.pack(anchor="w", pady=2)
        same_btn(brow3, "QR 코드 보기", self._psk_show_qr, "#6f42c1")

        self.psk_summary = StringVar(value="1) Zone 목록 → 2) WLAN 상세 → 3) 위 목록만 변경")
        Label(outer, textvariable=self.psk_summary, font=("Segoe UI", 9), bg=BG, fg="#555").pack(anchor="w", pady=(6, 2))

        cols = ("zoneName", "ssid", "type", "dpsk", "method", "algorithm", "mfp", "passphrase", "saePassphrase", "wlanId")
        headings = {
            "zoneName": "Zone",
            "ssid": "SSID",
            "type": "유형",
            "dpsk": "DPSK",
            "method": "보안규격",
            "algorithm": "암호화방식",
            "mfp": "MFP",
            "passphrase": "Passphrase",
            "saePassphrase": "SAE Passphrase",
            "wlanId": "BSSID(WLAN)",
        }
        widths = {
            "zoneName": 110, "ssid": 110, "type": 100, "dpsk": 50, "method": 102,
            "algorithm": 72, "mfp": 72, "passphrase": 100, "saePassphrase": 100, "wlanId": 70,
        }

        style = ttk.Style()
        style.configure("PskOff.Treeview", foreground="#8a8a8a", fieldbackground="#ececec", background="#ececec")
        style.map("PskOff.Treeview",
                  foreground=[("selected", "#8a8a8a")],
                  background=[("selected", "#d8d8d8")])

        def make_tree(parent, height, selectmode, ttk_style="Treeview"):
            fr = Frame(parent, bg=CARD)
            tree = ttk.Treeview(
                fr, columns=cols, show="headings", height=height,
                selectmode=selectmode, style=ttk_style,
            )
            for c in cols:
                tree.heading(c, text=headings[c])
                tree.column(c, width=widths[c], minwidth=40, anchor="w", stretch=False)
            sb_y = Scrollbar(fr, orient=VERTICAL, command=tree.yview)
            sb_x = Scrollbar(fr, orient=HORIZONTAL, command=tree.xview)
            tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
            tree.grid(row=0, column=0, sticky="nsew")
            sb_y.grid(row=0, column=1, sticky="ns")
            sb_x.grid(row=1, column=0, sticky="ew")
            fr.grid_rowconfigure(0, weight=1)
            fr.grid_columnconfigure(0, weight=1)
            return fr, tree

        lists = Frame(outer, bg=BG)
        lists.pack(fill=BOTH, expand=True, pady=(0, 2))
        lists.grid_columnconfigure(0, weight=1)
        lists.grid_rowconfigure(1, weight=1)
        lists.grid_rowconfigure(3, weight=1)

        Label(
            lists, text="PSK / SAE WLAN (암호 변경 가능)",
            font=("Segoe UI", 9, "bold"), bg=BG, fg="#198754",
        ).grid(row=0, column=0, sticky="w")
        ok_fr, self.psk_tree = make_tree(lists, 6, "extended")
        ok_fr.grid(row=1, column=0, sticky="nsew", pady=(0, 4))

        Label(
            lists, text="PSK / SAE 아님 (암호 변경 불가)",
            font=("Segoe UI", 9, "bold"), bg=BG, fg="#888",
        ).grid(row=2, column=0, sticky="w")
        off_fr, self.psk_tree_off = make_tree(lists, 6, "none", "PskOff.Treeview")
        off_fr.grid(row=3, column=0, sticky="nsew")
        self._make_log(outer, "로그", 4)
        self._psk_rows = []
        self._psk_zones = []

    def _psk_update_api_versions(self, event=None):
        ctrl = self.psk_ctrl.get() or "수동선택"
        versions = CONTROLLER_API_MAP.get(ctrl, CONTROLLER_API_MAP["수동선택"])
        self.psk_api_combo["values"] = versions
        if versions:
            self.psk_api.set(versions[0])

    def _psk_client(self):
        host = self.psk_ip.get().strip()
        user = self.psk_user.get().strip()
        pw = self.psk_pass.get()
        api_ver = self.psk_api.get().strip()
        if not host or not user or not pw or not api_ver:
            messagebox.showwarning("안내", "SZ IP / Username / Password / API 버전을 입력하세요.")
            return None
        return SmartZoneAPI(host, user, pw, api_ver)

    def _psk_busy(self):
        if self._worker and self._worker.is_alive():
            messagebox.showwarning("안내", "이미 실행 중입니다.")
            return True
        return False

    def _psk_load_zones(self):
        client = self._psk_client()
        if not client or self._psk_busy():
            return
        self.psk_zone_btn.config(state=DISABLED)
        self._log(f"\n===== Zone 목록: {client.host} / {client.api_version} =====\n")

        def work():
            try:
                def log(msg):
                    self._log_queue.put(str(msg) + "\n")
                ok, msg = client.login()
                log(msg)
                if not ok:
                    self.after(0, lambda: self.psk_summary.set("실패: " + msg))
                    return
                zones = client.fetch_zones()
                self._psk_zones = zones
                names = ["모든 Zone"]
                for z in zones:
                    nm = z.get("name") or z.get("zoneName") or z.get("id")
                    names.append(str(nm))
                    log(f"  Zone {nm}  id={z.get('id')}")
                self.after(0, lambda: self._psk_set_zones(names, f"Zone {len(zones)} 개. WLAN 상세 가져오기를 클릭하세요."))
            except Exception as e:
                self._log_queue.put(f"예외: {e}\n")
                self.after(0, lambda: self.psk_summary.set(f"오류: {e}"))
            finally:
                self.after(0, lambda: self.psk_zone_btn.config(state=NORMAL))

        self._worker = threading.Thread(target=work, daemon=True)
        self._worker.start()

    def _psk_set_zones(self, names, summary):
        self.psk_zone_combo["values"] = names
        self.psk_zone.set(names[0] if names else "모든 Zone")
        self.psk_summary.set(summary)
        self.psk_hint_zone.set("WLAN 상세 가져오기를 클릭하세요.")
        self.psk_hint_wlan.set("")

    def _psk_selected_zone_id(self):
        label = self.psk_zone.get()
        if not label or label == "모든 Zone":
            return "ALL"
        for z in self._psk_zones:
            if (z.get("name") or z.get("zoneName")) == label:
                return z.get("id") or z.get("zoneId")
        return "ALL"

    def _psk_list(self):
        client = self._psk_client()
        if not client or self._psk_busy():
            return
        zone_id = self._psk_selected_zone_id()
        self.psk_list_btn.config(state=DISABLED)
        self.psk_summary.set("WLAN 상세 조회 중...")
        self._log(f"\n===== WLAN 상세 조회 zone={zone_id} =====\n")

        def work():
            try:
                def log(msg):
                    self._log_queue.put(str(msg) + "\n")
                ok, msg = client.login()
                log(msg)
                if not ok:
                    self.after(0, lambda: self.psk_summary.set("실패: " + msg))
                    return
                zid = None if zone_id == "ALL" else zone_id
                rows = client.collect_wlan_details(zid, log=log)
                self._psk_rows = rows
                self.after(0, lambda: self._psk_fill_tree(rows))
            except Exception as e:
                self._log_queue.put(f"예외: {e}\n")
                self.after(0, lambda: self.psk_summary.set(f"오류: {e}"))
            finally:
                self.after(0, lambda: self.psk_list_btn.config(state=NORMAL))

        self._worker = threading.Thread(target=work, daemon=True)
        self._worker.start()

    def _psk_row_values(self, r):
        return (
            r.get("zoneName", ""),
            r.get("ssid", ""),
            r.get("type", ""),
            r.get("dpsk", "N"),
            r.get("method", ""),
            r.get("algorithm", ""),
            r.get("mfp", "-"),
            r.get("passphrase", "-"),
            r.get("saePassphrase", "-"),
            r.get("wlanId", ""),
        )

    def _psk_fill_tree(self, rows):
        self.psk_tree.delete(*self.psk_tree.get_children())
        self.psk_tree_off.delete(*self.psk_tree_off.get_children())
        ok_n = off_n = 0
        for r in rows:
            vals = self._psk_row_values(r)
            if r.get("changeable"):
                self.psk_tree.insert("", END, values=vals)
                ok_n += 1
            else:
                self.psk_tree_off.insert("", END, values=vals)
                off_n += 1
        self.psk_summary.set(
            f"변경가능 {ok_n} 개  /  변경불가 {off_n} 개  (아래 목록은 선택·변경 안 됨)"
        )
        self.psk_hint_wlan.set("Zone을 선택할 수 있습니다.")
        self.psk_hint_zone.set("")
        self.psk_hint_pw.set("WLAN을 선택한 후 암호변경을 하세요.")

    def _wifi_qr_text(self, ssid: str, password: str, method: str) -> str:
        # PHP generate_qr.php 와 동일한 WIFI QR
        def esc(s: str) -> str:
            return (
                str(s)
                .replace("\\", "\\\\")
                .replace(";", "\\;")
                .replace(",", "\\,")
                .replace(":", "\\:")
                .replace('"', '\\"')
            )
        t = "SAE" if (method or "").upper() == "WPA3" else "WPA"
        return f"WIFI:T:{t};S:{esc(ssid)};P:{esc(password)};;"

    def _psk_show_qr(self):
        sel = self.psk_tree.selection()
        if not sel:
            messagebox.showwarning("안내", "QR을 볼 WLAN을 목록에서 선택하세요.")
            return
        vals = self.psk_tree.item(sel[0], "values")
        ssid = vals[1] if len(vals) > 1 else ""
        method = vals[4] if len(vals) > 4 else "WPA2"
        psk = vals[7] if len(vals) > 7 else "-"
        sae = vals[8] if len(vals) > 8 else "-"
        password = self.psk_new.get().strip()
        if not password:
            if method.upper() == "WPA3" and sae and sae != "-":
                password = sae
            elif psk and psk != "-":
                password = psk
            elif sae and sae != "-":
                password = sae
        if not password:
            messagebox.showwarning("안내", "표시할 비밀번호가 없습니다. 상세 조회로 Passphrase가 나오거나 새 비밀번호를 입력하세요.")
            return
        payload = self._wifi_qr_text(ssid, password, method)
        try:
            import io
            import qrcode
            from PIL import Image, ImageTk
        except Exception:
            messagebox.showerror(
                "모듈 없음",
                "qrcode / pillow 가 필요합니다.\nrun_script.bat 을 다시 실행해 모듈을 설치하세요.",
            )
            return
        img = qrcode.make(payload, box_size=8, border=2)
        if not hasattr(img, "save"):
            img = img.get_image()
        img = img.convert("RGB").resize((280, 280))
        bio = io.BytesIO()
        img.save(bio, format="PNG")
        bio.seek(0)
        from PIL import Image as PILImage
        tk_img = ImageTk.PhotoImage(PILImage.open(bio))
        win = Toplevel(self)
        win.title(f"QR - {ssid}")
        win.configure(bg=CARD)
        Label(win, text=ssid, font=("Segoe UI", 12, "bold"), bg=CARD).pack(pady=(12, 4))
        Label(win, text=f"{method}  |  {payload}", font=("Segoe UI", 8), bg=CARD, fg="#666", wraplength=360).pack()
        lbl = Label(win, image=tk_img, bg=CARD)
        lbl.image = tk_img
        lbl.pack(padx=16, pady=12)
        Button(win, text="닫기", command=win.destroy, padx=16, pady=4).pack(pady=(0, 12))
        self._log(f"QR: {ssid} {payload}\n")

    def _psk_apply(self, all_changeable):
        new_pw = self.psk_new.get()
        if not new_pw or len(new_pw) < 8:
            messagebox.showwarning("안내", "새 비밀번호는 8자 이상이어야 합니다.")
            return
        method = self.psk_method.get() or "WPA2"
        targets = []
        if all_changeable:
            for r in self._psk_rows:
                if r.get("changeable"):
                    targets.append(r)
            if not targets:
                messagebox.showwarning("안내", "변경 가능한 WLAN이 없습니다. (Standard_Open + WPA2/WPA3)")
                return
        else:
            sel = self.psk_tree.selection()
            if not sel:
                messagebox.showwarning("안내", "개별 변경할 WLAN을 목록에서 선택하세요.")
                return
            by_id = {r.get("wlanId"): r for r in self._psk_rows}
            for iid in sel:
                vals = self.psk_tree.item(iid, "values")
                wlan_id = vals[8] if len(vals) > 8 else ""
                r = by_id.get(wlan_id)
                if not r:
                    continue
                if not r.get("changeable"):
                    messagebox.showwarning(
                        "안내",
                        f"{r.get('ssid')} 은(는) 변경 불가입니다.\n유형={r.get('type')} / 보안={r.get('method')}",
                    )
                    return
                targets.append(r)
        names = ", ".join((t.get("ssid") or "") for t in targets[:8])
        extra = "" if len(targets) <= 8 else f" 외 {len(targets)-8}개"
        if not messagebox.askyesno(
            "확인",
            f"{len(targets)} 개 WLAN 을 {method} 로 변경합니다.\n\n{names}{extra}\n\n계속할까요?",
        ):
            return
        client = self._psk_client()
        if not client or self._psk_busy():
            return
        self.psk_run_btn.config(state=DISABLED)
        self.psk_all_btn.config(state=DISABLED)
        self.psk_summary.set("비밀번호 변경 중...")
        self._log(f"\n===== PSK 변경 {len(targets)} 개 method={method} =====\n")

        def work():
            try:
                def log(msg):
                    self._log_queue.put(str(msg) + "\n")
                ok, msg = client.login()
                log(msg)
                if not ok:
                    self.after(0, lambda: self.psk_summary.set("실패: " + msg))
                    return
                results = []
                ok_n = 0
                for i, t in enumerate(targets, 1):
                    log(f"[{i}/{len(targets)}] {t.get('zoneName')} / {t.get('ssid')}")
                    good, detail = client.update_wlan_password(
                        t["zoneId"], t["wlanId"], new_pw, method=method, log=log,
                    )
                    log("  " + detail)
                    results.append({
                        "zone": t.get("zoneName"),
                        "ssid": t.get("ssid"),
                        "type": t.get("type"),
                        "method": method,
                        "wlanId": t.get("wlanId"),
                        "status": "OK" if good else "FAIL",
                        "message": detail,
                    })
                    if good:
                        ok_n += 1
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                out = RESULTS_PSK / f"result_psk_{ts}.csv"
                RESULTS_PSK.mkdir(exist_ok=True)
                save_csv(
                    out,
                    ["zone", "ssid", "type", "method", "wlanId", "status", "message"],
                    results,
                )
                keep_latest_results(RESULTS_PSK)
                self._save_session_log(LOG_PSK, "psk")
                log(f"결과 파일: {out}")
                log("WLAN 목록 자동 새로고침...")
                zone_id = self._psk_selected_zone_id()
                zid = None if zone_id == "ALL" else zone_id
                rows = client.collect_wlan_details(zid, log=log)
                self._psk_rows = rows
                summary = f"완료: 성공 {ok_n}/{len(targets)}  (목록 갱신됨)"
                self.after(0, lambda r=rows: self._psk_fill_tree(r))
                self.after(0, lambda: self.psk_summary.set(summary))
                self.after(0, lambda: messagebox.showinfo("완료", summary + f"\n\n{out}"))
            except Exception as e:
                self._log_queue.put(f"예외: {e}\n")
                self.after(0, lambda: self.psk_summary.set(f"오류: {e}"))
            finally:
                self.after(0, lambda: self.psk_run_btn.config(state=NORMAL))
                self.after(0, lambda: self.psk_all_btn.config(state=NORMAL))

        self._worker = threading.Thread(target=work, daemon=True)
        self._worker.start()

    # ------------------------------------------------------------------
    # 메뉴 1: SmartZone 정보 보기
    # ------------------------------------------------------------------

    def _build_dpsk(self):
        self._clear_page()
        outer = Frame(self, bg=BG, padx=12, pady=10)
        outer.pack(fill=BOTH, expand=True)
        top = Frame(outer, bg=BG)
        top.pack(fill=X)
        self._back_btn(top)
        Label(top, text="4. SmartZone DPSK 관리", font=("Segoe UI", 14, "bold"),
              fg=ACCENT, bg=BG).pack(side=LEFT, padx=12)

        self.dpsk_ip = StringVar()
        self.dpsk_user = StringVar(value="admin")
        self.dpsk_pass = StringVar()
        self.dpsk_ctrl = StringVar(value="7.2.0" if "7.2.0" in CONTROLLER_API_MAP else "수동선택")
        self.dpsk_api = StringVar()
        self.dpsk_zone_filter = StringVar(value="전체 Zone")
        self.dpsk_search = StringVar()
        self.dpsk_create_zone = StringVar(value="선택")
        self.dpsk_create_wlan = StringVar()
        self.dpsk_create_user = StringVar()
        self.dpsk_create_count = StringVar(value="1")
        self.dpsk_create_psk = StringVar()
        self.dpsk_create_role = StringVar(value="선택 안 함")
        self.dpsk_create_vlan = StringVar()
        self.dpsk_create_group = StringVar(value="False (개인)")
        self._dpsk_api_cli = None
        self._dpsk_zones = []
        self._dpsk_all = []
        self._dpsk_wlans = []
        self._dpsk_roles = []

        body = Frame(outer, bg=BG)
        body.pack(fill=BOTH, expand=True, pady=(8, 0))

        side_wrap = Frame(body, bg=BG, width=268)
        side_wrap.pack(side=LEFT, fill=Y, padx=(0, 10))
        side_wrap.pack_propagate(False)
        side_canvas = Canvas(side_wrap, bg=BG, highlightthickness=0, width=248)
        side_sb = Scrollbar(side_wrap, orient=VERTICAL, command=side_canvas.yview)
        side_canvas.configure(yscrollcommand=side_sb.set)
        side_canvas.pack(side=LEFT, fill=BOTH, expand=True)
        side_sb.pack(side=RIGHT, fill=Y)
        side = Frame(side_canvas, bg=BG)
        side_win = side_canvas.create_window((0, 0), window=side, anchor="nw")

        def _dpsk_side_scroll(_event=None):
            side_canvas.configure(scrollregion=side_canvas.bbox("all"))
            side_canvas.itemconfigure(side_win, width=side_canvas.winfo_width())

        side.bind("<Configure>", _dpsk_side_scroll)
        side_canvas.bind("<Configure>", _dpsk_side_scroll)

        def _dpsk_mousewheel(event):
            delta = -1 if getattr(event, "delta", 0) > 0 else 1
            if getattr(event, "num", None) == 4:
                delta = -1
            elif getattr(event, "num", None) == 5:
                delta = 1
            side_canvas.yview_scroll(delta, "units")

        side_canvas.bind("<Enter>", lambda e: side_canvas.bind_all("<MouseWheel>", _dpsk_mousewheel))
        side_canvas.bind("<Leave>", lambda e: side_canvas.unbind_all("<MouseWheel>"))
        side.bind("<Enter>", lambda e: side.bind_all("<MouseWheel>", _dpsk_mousewheel))
        side.bind("<Leave>", lambda e: side.unbind_all("<MouseWheel>"))

        def stack_label(parent, text):
            Label(parent, text=text, bg=CARD, font=("Segoe UI", 8), fg="#555").pack(anchor="w", pady=(6, 0))

        box1 = LabelFrame(side, text="1. 접속 설정", bg=CARD, fg="#333", font=("Segoe UI", 9, "bold"),
                          padx=8, pady=6)
        box1.pack(fill=X)
        stack_label(box1, "SZ IP/Domain")
        Entry(box1, textvariable=self.dpsk_ip, font=("Segoe UI", 10)).pack(fill=X)
        stack_label(box1, "Username")
        Entry(box1, textvariable=self.dpsk_user, font=("Segoe UI", 10)).pack(fill=X)
        stack_label(box1, "Password")
        Entry(box1, textvariable=self.dpsk_pass, show="*", font=("Segoe UI", 10)).pack(fill=X)
        stack_label(box1, "컨트롤러 버전")
        cb = ttk.Combobox(box1, textvariable=self.dpsk_ctrl, values=list(CONTROLLER_API_MAP.keys()),
                          state="readonly", width=26)
        cb.pack(fill=X)
        cb.bind("<<ComboboxSelected>>", lambda e: self._dpsk_update_api())
        stack_label(box1, "API 버전")
        self.dpsk_api_combo = ttk.Combobox(box1, textvariable=self.dpsk_api, state="readonly", width=26)
        self.dpsk_api_combo.pack(fill=X)
        Button(box1, text="로그인 & DPSK 조회", font=("Segoe UI", 9, "bold"), bg=LINK, fg="white",
               relief="flat", pady=5, command=self._dpsk_refresh, cursor="hand2").pack(fill=X, pady=(10, 4))
        self._dpsk_update_api()

        box2 = LabelFrame(side, text="2. DPSK 생성", bg=CARD, fg="#333", font=("Segoe UI", 9, "bold"),
                          padx=8, pady=6)
        box2.pack(fill=X, pady=(10, 0))
        stack_label(box2, "Zone")
        self.dpsk_cz_combo = ttk.Combobox(box2, textvariable=self.dpsk_create_zone, state="readonly")
        self.dpsk_cz_combo.pack(fill=X)
        self.dpsk_cz_combo.bind("<<ComboboxSelected>>", lambda e: self._dpsk_load_create_wlans())
        stack_label(box2, "WLAN 선택 (DPSK Enabled)")
        self.dpsk_cw_combo = ttk.Combobox(box2, textvariable=self.dpsk_create_wlan, state="readonly")
        self.dpsk_cw_combo.pack(fill=X)
        stack_label(box2, "Number of DPSKs")
        Entry(box2, textvariable=self.dpsk_create_count, font=("Segoe UI", 10)).pack(fill=X)
        stack_label(box2, "User Name")
        Entry(box2, textvariable=self.dpsk_create_user, font=("Segoe UI", 10)).pack(fill=X)
        stack_label(box2, "Passphrase (비우면 자동생성)")
        Entry(box2, textvariable=self.dpsk_create_psk, font=("Segoe UI", 10)).pack(fill=X)
        stack_label(box2, "User Role")
        self.dpsk_role_combo = ttk.Combobox(box2, textvariable=self.dpsk_create_role, state="readonly")
        self.dpsk_role_combo.pack(fill=X)
        stack_label(box2, "VLAN ID (1 – 4094, 선택 사항)")
        Entry(box2, textvariable=self.dpsk_create_vlan, font=("Segoe UI", 10)).pack(fill=X)
        stack_label(box2, "Group DPSK")
        ttk.Combobox(box2, textvariable=self.dpsk_create_group,
                     values=["False (개인)", "True (그룹)"], state="readonly").pack(fill=X)
        Button(box2, text="DPSK 생성하기", font=("Segoe UI", 9, "bold"), bg="#28a745", fg="white",
               relief="flat", pady=5, command=self._dpsk_create, cursor="hand2").pack(fill=X, pady=(10, 4))

        main = Frame(body, bg=CARD, padx=10, pady=8, highlightbackground="#dee2e6", highlightthickness=1)
        main.pack(side=LEFT, fill=BOTH, expand=True)
        filt = Frame(main, bg=CARD)
        filt.pack(fill=X)
        Label(filt, text="Zone 필터:", bg=CARD).pack(side=LEFT)
        self.dpsk_zone_combo = ttk.Combobox(filt, textvariable=self.dpsk_zone_filter, state="readonly", width=28)
        self.dpsk_zone_combo.pack(side=LEFT, padx=4)
        self.dpsk_zone_combo.bind("<<ComboboxSelected>>", lambda e: self._dpsk_fill_tree())
        Label(filt, text="검색:", bg=CARD).pack(side=LEFT, padx=(12, 0))
        Entry(filt, textvariable=self.dpsk_search, width=28).pack(side=LEFT, padx=4)
        Button(filt, text="검색", bg=LINK, fg="white", relief="flat", padx=10,
               command=self._dpsk_fill_tree).pack(side=LEFT)

        head = Frame(main, bg=CARD)
        head.pack(fill=X, pady=(8, 2))
        Label(head, text="DPSK 목록", font=("Segoe UI", 10, "bold"), bg=CARD, fg=LINK).pack(side=LEFT)
        self.dpsk_count = StringVar(value="총 0개")
        Label(head, textvariable=self.dpsk_count, bg=CARD, fg="#555").pack(side=RIGHT)
        Button(head, text="CSV 다운로드", bg="#28a745", fg="white", relief="flat", padx=8,
               command=self._dpsk_csv).pack(side=RIGHT, padx=4)
        Button(head, text="선택 항목 삭제", bg="#dc3545", fg="white", relief="flat", padx=8,
               command=self._dpsk_delete).pack(side=RIGHT, padx=4)
        Button(head, text="결과 폴더 열기", bg=BTN_BG, relief="solid", borderwidth=1, padx=8,
               command=lambda: self._open_path(RESULTS_DPSK)).pack(side=RIGHT, padx=4)
        Button(head, text="최근 결과 다운로드", bg="#28a745", fg="white", relief="flat", padx=8,
               command=lambda: self._download_latest_results(RESULTS_DPSK)).pack(side=RIGHT, padx=4)

        cols = ("zone", "wlan", "user", "psk", "mac", "role", "vlan", "group", "created", "exp", "status")
        tree_fr = Frame(main, bg=CARD)
        tree_fr.pack(fill=BOTH, expand=True)
        self.dpsk_tree = ttk.Treeview(tree_fr, columns=cols, show="headings", height=16, selectmode="extended")
        headers = {
            "zone": "Zone", "wlan": "WLAN 이름", "user": "User Name", "psk": "Passphrase",
            "mac": "MAC", "role": "User Role", "vlan": "VLAN", "group": "Group DPSK",
            "created": "생성일시", "exp": "만료일시", "status": "상태",
        }
        widths = {"zone": 90, "wlan": 110, "user": 90, "psk": 110, "mac": 110, "role": 80,
                  "vlan": 50, "group": 80, "created": 120, "exp": 120, "status": 70}
        for c in cols:
            self.dpsk_tree.heading(c, text=headers[c])
            self.dpsk_tree.column(c, width=widths[c], anchor="w")
        ysb = Scrollbar(tree_fr, command=self.dpsk_tree.yview)
        xsb = Scrollbar(tree_fr, orient=HORIZONTAL, command=self.dpsk_tree.xview)
        self.dpsk_tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.dpsk_tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        tree_fr.grid_rowconfigure(0, weight=1)
        tree_fr.grid_columnconfigure(0, weight=1)

        self.dpsk_status = StringVar(value="왼쪽에서 로그인 후 DPSK를 조회하세요.")
        Label(outer, textvariable=self.dpsk_status, font=("Segoe UI", 9), bg=BG, fg="#555").pack(anchor="w", pady=(6, 0))

    def _dpsk_update_api(self):
        vers = CONTROLLER_API_MAP.get(self.dpsk_ctrl.get(), CONTROLLER_API_MAP.get("수동선택", []))
        if hasattr(self, "dpsk_api_combo"):
            self.dpsk_api_combo["values"] = vers
        if vers and self.dpsk_api.get() not in vers:
            self.dpsk_api.set(vers[0])

    def _dpsk_fmt_time(self, raw):
        raw = raw or "-"
        if raw in ("-", ""):
            return "-", "-"
        if str(raw) == "Unlimited":
            return "무제한", "Active"
        if "from first use" in str(raw).lower():
            return "미사용", "Active"
        try:
            s = str(raw).replace("/", "-")
            dt = datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
        except Exception:
            try:
                dt = datetime.fromisoformat(str(raw).replace("Z", "+00:00")).replace(tzinfo=None)
            except Exception:
                return str(raw), "-"
        # SZ 값은 UTC로 오는 경우가 많아 KST +9
        try:
            from datetime import timedelta
            local = dt + timedelta(hours=9)
        except Exception:
            local = dt
        status = "Expired" if local < datetime.now() else "Active"
        return local.strftime("%Y/%m/%d %H:%M:%S"), status

    def _dpsk_refresh(self):
        host = self.dpsk_ip.get().strip()
        user = self.dpsk_user.get().strip()
        pw = self.dpsk_pass.get()
        api = self.dpsk_api.get().strip()
        if not (host and user and pw and api):
            messagebox.showwarning("안내", "SZ IP / Username / Password / API 버전을 입력하세요.")
            return
        try:
            cli = SmartZoneAPI(host, user, pw, api)
            ok, msg = cli.login()
            if not ok:
                messagebox.showerror("로그인 실패", msg)
                return
            self._dpsk_api_cli = cli
            zones = cli.fetch_zones()
            self._dpsk_zones = zones
            labels = ["전체 Zone"]
            rows = []
            wlan_cache = {}
            for z in zones:
                zid = z.get("id") or ""
                zname = z.get("name") or zid
                labels.append(f"{zname}")
                dpsks = cli.fetch_dpsk_list(zid)
                for d in dpsks:
                    wid = d.get("wlanId") or ""
                    key = f"{zid}:{wid}"
                    if key not in wlan_cache and wid:
                        code, wd = cli.fetch_wlan(zid, wid)
                        wlan_cache[key] = (wd.get("name") if isinstance(wd, dict) else "") or ""
                    created, _ = self._dpsk_fmt_time(d.get("creationDateTime"))
                    exp, st = self._dpsk_fmt_time(d.get("expirationDateTime"))
                    if str(d.get("expirationDateTime")) == "Unlimited":
                        exp, st = "무제한", "Active"
                    rows.append({
                        "id": d.get("id") or "",
                        "zone_id": zid,
                        "zone": zname,
                        "wlan_id": wid,
                        "wlan": wlan_cache.get(key, ""),
                        "user": d.get("userName") or "",
                        "psk": d.get("passphrase") or "",
                        "mac": d.get("macAddress") or "Unbound",
                        "role": d.get("userRoleId") or "-",
                        "vlan": d.get("vlanId") if d.get("vlanId") not in (None, "") else "-",
                        "group": "True" if d.get("groupDpsk") else "False",
                        "created": created,
                        "exp": exp,
                        "status": st,
                    })
            self._dpsk_all = rows
            self.dpsk_zone_combo["values"] = labels
            if self.dpsk_zone_filter.get() not in labels:
                self.dpsk_zone_filter.set("전체 Zone")
            names = [z.get("name") or z.get("id") for z in zones]
            self.dpsk_cz_combo["values"] = names
            if names and self.dpsk_create_zone.get() in ("", "선택"):
                self.dpsk_create_zone.set(names[0])
                self._dpsk_load_create_wlans()
            try:
                roles = cli.fetch_user_roles()
            except Exception:
                roles = []
            self._dpsk_roles = roles
            role_labels = ["선택 안 함"] + [f"{r.get('name')}  [{(r.get('id') or '')[:8]}]" for r in roles]
            if hasattr(self, "dpsk_role_combo"):
                self.dpsk_role_combo["values"] = role_labels
            if self.dpsk_create_role.get() not in role_labels:
                self.dpsk_create_role.set("선택 안 함")
            self._dpsk_fill_tree()
            self.dpsk_status.set(f"{msg}  /  DPSK {len(rows)}건")
        except Exception as e:
            messagebox.showerror("오류", str(e))

    def _dpsk_fill_tree(self):
        tree = getattr(self, "dpsk_tree", None)
        if tree is None:
            return
        tree.delete(*tree.get_children())
        zf = self.dpsk_zone_filter.get()
        q = (self.dpsk_search.get() or "").strip().lower()
        n = 0
        for r in self._dpsk_all:
            if zf and zf not in ("ALL", "전체 Zone") and r.get("zone") != zf:
                continue
            blob = " ".join(str(r.get(k, "")) for k in ("zone", "wlan", "user", "psk", "mac", "role")).lower()
            if q and q not in blob:
                continue
            tree.insert("", END, iid=r["id"] or f"row{n}", values=(
                r["zone"], r["wlan"], r["user"], r["psk"], r["mac"], r.get("role", "-"),
                r["vlan"], r["group"], r["created"], r["exp"], r["status"],
            ))
            n += 1
        self.dpsk_count.set(f"총 {n}개")

    def _dpsk_zone_id(self, name):
        for z in self._dpsk_zones:
            if (z.get("name") or z.get("id")) == name or z.get("id") == name:
                return z.get("id")
        return ""

    def _dpsk_load_create_wlans(self):
        cli = self._dpsk_api_cli
        zid = self._dpsk_zone_id(self.dpsk_create_zone.get())
        if not cli or not zid:
            return
        try:
            wlans = cli.fetch_dpsk_wlans(zid)
            self._dpsk_wlans = wlans
            labels = [f"{w.get('name')}  ({w.get('ssid')})" for w in wlans]
            self.dpsk_cw_combo["values"] = labels
            if labels:
                self.dpsk_create_wlan.set(labels[0])
            else:
                self.dpsk_create_wlan.set("")
        except Exception as e:
            messagebox.showerror("WLAN 조회 실패", str(e))

    def _dpsk_create(self):
        cli = self._dpsk_api_cli
        zid = self._dpsk_zone_id(self.dpsk_create_zone.get())
        label = self.dpsk_create_wlan.get()
        username = self.dpsk_create_user.get().strip()
        if not cli or not zid:
            messagebox.showwarning("안내", "먼저 목록 조회를 하세요.")
            return
        wlan = None
        for w in self._dpsk_wlans:
            if f"{w.get('name')}  ({w.get('ssid')})" == label or w.get("id") == label:
                wlan = w
                break
        if not wlan:
            messagebox.showwarning("안내", "DPSK Enabled WLAN을 선택하세요.")
            return
        try:
            amount = int((self.dpsk_create_count.get() or "1").strip())
        except ValueError:
            messagebox.showwarning("안내", "Number of DPSKs 는 숫자여야 합니다.")
            return
        if amount < 1 or amount > 500:
            messagebox.showwarning("안내", "Number of DPSKs 범위는 1–500 입니다.")
            return
        if amount == 1 and username and any(r.get("zone_id") == zid and r.get("user") == username for r in self._dpsk_all):
            if not messagebox.askyesno("중복", f"사용자 '{username}' 이(가) 이미 있습니다. 강제 생성할까요?"):
                return
        vlan = None
        vs = self.dpsk_create_vlan.get().strip()
        if vs:
            try:
                vlan = int(vs)
            except ValueError:
                messagebox.showwarning("안내", "VLAN은 숫자여야 합니다.")
                return
        group = str(self.dpsk_create_group.get()).lower().startswith("true")
        role_id = ""
        role_label = self.dpsk_create_role.get()
        for rr in getattr(self, "_dpsk_roles", []):
            if f"{rr.get('name')}  [{rr.get('id')[:8]}]" == role_label or rr.get("name") == role_label:
                role_id = rr.get("id") or ""
                break
        passphrase = self.dpsk_create_psk.get().strip()
        code, body = cli.create_dpsk(
            zid, wlan["id"], username, group, vlan,
            amount=amount, passphrase=passphrase, user_role_id=role_id,
        )
        ok = code in (200, 201, 204)
        info = []
        if isinstance(body, dict):
            info = body.get("dpskInfoList") or []
            if body.get("error"):
                ok = False
        if ok:
            lines = []
            for it in info[:20]:
                if isinstance(it, dict):
                    lines.append(f"{it.get('userName','')}  /  {it.get('passphrase','')}")
            extra = "" if len(info) <= 20 else f"\n... 외 {len(info)-20}건"
            messagebox.showinfo("완료", f"{len(info) or amount}건 생성\n" + "\n".join(lines) + extra)
            self._dpsk_refresh()
        else:
            messagebox.showerror("생성 실패", str(body)[:500])

    def _dpsk_delete(self):
        cli = self._dpsk_api_cli
        if not cli:
            return
        sel = self.dpsk_tree.selection()
        if not sel:
            messagebox.showwarning("안내", "삭제할 DPSK를 선택하세요.")
            return
        if not messagebox.askyesno("확인", f"{len(sel)}건을 삭제할까요?"):
            return
        by = {}
        lookup = {r["id"]: r for r in self._dpsk_all}
        for iid in sel:
            r = lookup.get(iid)
            if not r:
                continue
            key = (r["zone_id"], r["wlan_id"])
            by.setdefault(key, []).append(r["id"])
        ok_n = fail_n = 0
        for (zid, wid), ids in by.items():
            code, body = cli.delete_dpsk(zid, wid, ids)
            if code in (200, 201, 204):
                ok_n += len(ids)
            else:
                fail_n += len(ids)
        messagebox.showinfo("삭제", f"성공 {ok_n} / 실패 {fail_n}")
        self._dpsk_refresh()

    def _dpsk_csv(self):
        rows = []
        zf = self.dpsk_zone_filter.get()
        q = (self.dpsk_search.get() or "").strip().lower()
        for r in self._dpsk_all:
            if zf and zf not in ("ALL", "전체 Zone") and r.get("zone") != zf:
                continue
            blob = " ".join(str(r.get(k, "")) for k in ("zone", "wlan", "user", "psk", "mac", "role")).lower()
            if q and q not in blob:
                continue
            rows.append(r)
        if not rows:
            messagebox.showwarning("안내", "저장할 목록이 없습니다.")
            return
        dest = filedialog.asksaveasfilename(
            title="DPSK CSV 저장",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=f"dpsk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        if not dest:
            return
        out = Path(dest)
        with open(out, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Zone", "WLAN 이름", "User Name", "Passphrase", "MAC", "User Role", "VLAN", "Group DPSK", "생성일시", "만료일시", "상태"])
            for r in rows:
                w.writerow([r["zone"], r["wlan"], r["user"], r["psk"], r["mac"], r.get("role", "-"), r["vlan"], r["group"],
                            r["created"], r["exp"], r["status"]])
        try:
            if RESULTS_DPSK.exists():
                for old in RESULTS_DPSK.glob("dpsk_*.csv"):
                    try:
                        old.unlink()
                    except Exception:
                        pass
        except Exception:
            pass
        try:
            os.startfile(str(out.resolve()))
        except Exception:
            pass
        messagebox.showinfo("저장", f"{len(rows)}건 저장\n{out}")


    def _build_icx(self):
        self._clear_page()
        outer = self._scroll_page(padx=20, pady=16)
        self._back_btn(outer)
        Label(outer, text="10. ICX ARP 조회 (SNMP)", font=("Segoe UI", 14, "bold"),
              fg=ACCENT, bg=BG).pack(anchor="w")
        Label(outer, text="SNMPv2c  ·  sysDescr / sysName / ifDescr / ipNetToPhysicalPhysAddress",
              font=("Segoe UI", 9), fg="#666", bg=BG).pack(anchor="w", pady=(2, 8))

        self.icx_ip = StringVar()
        self.icx_comm = StringVar(value="public")
        self._icx_rows = []
        form = Frame(outer, bg=CARD, padx=12, pady=10, highlightbackground="#dee2e6", highlightthickness=1)
        form.pack(fill=X)
        def fld(lbl, var, w=20, show=""):
            fr = Frame(form, bg=CARD)
            fr.pack(side=LEFT, padx=(0, 12))
            Label(fr, text=lbl, bg=CARD, font=("Segoe UI", 8), fg="#666").pack(anchor="w")
            Entry(fr, textvariable=var, width=w, show=show, font=("Segoe UI", 10)).pack()
        fld("ICX Switch IP", self.icx_ip, 18)
        fld("SNMP Community", self.icx_comm, 16)
        Button(form, text="조회", font=("Segoe UI", 10, "bold"), bg=LINK, fg="white",
               relief="flat", padx=16, pady=6, command=self._icx_query, cursor="hand2").pack(side=LEFT, pady=(12, 0))
        Button(form, text="CSV 다운로드", bg="#28a745", fg="white", relief="flat", padx=12, pady=6,
               command=self._icx_csv).pack(side=LEFT, padx=8, pady=(12, 0))
        Button(form, text="결과 폴더 열기", bg=BTN_BG, relief="solid", borderwidth=1, padx=10, pady=6,
               command=lambda: self._open_path(RESULTS_SNMP)).pack(side=LEFT, padx=(0, 6), pady=(12, 0))
        Button(form, text="최근 결과 다운로드", bg="#28a745", fg="white", relief="flat", padx=10, pady=6,
               command=lambda: self._download_latest_results(RESULTS_SNMP)).pack(side=LEFT, pady=(12, 0))

        self.icx_info = StringVar(value="Community / Switch IP 입력 후 조회하세요.")
        Label(outer, textvariable=self.icx_info, font=("Segoe UI", 9), bg=BG, fg="#555").pack(anchor="w", pady=(8, 4))

        cols = ("ip", "mac", "mac_lower", "iface")
        tree_fr = Frame(outer, bg=CARD)
        tree_fr.pack(fill=BOTH, expand=True)
        self.icx_tree = ttk.Treeview(tree_fr, columns=cols, show="headings", height=18)
        headers = {"ip": "IP", "mac": "MAC", "mac_lower": "MAC (lowercase)", "iface": "VLAN / Interface"}
        widths = {"ip": 140, "mac": 150, "mac_lower": 150, "iface": 280}
        for c in cols:
            self.icx_tree.heading(c, text=headers[c])
            self.icx_tree.column(c, width=widths[c], anchor="w")
        ysb = Scrollbar(tree_fr, command=self.icx_tree.yview)
        xsb = Scrollbar(tree_fr, orient=HORIZONTAL, command=self.icx_tree.xview)
        self.icx_tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.icx_tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        tree_fr.grid_rowconfigure(0, weight=1)
        tree_fr.grid_columnconfigure(0, weight=1)

    def _icx_query(self):
        host = self.icx_ip.get().strip()
        comm = self.icx_comm.get().strip()
        if not host or not comm:
            messagebox.showwarning("안내", "Switch IP 와 Community 를 입력하세요.")
            return
        self.icx_info.set("조회 중...")
        self.icx_tree.delete(*self.icx_tree.get_children())

        def work():
            try:
                res = query_icx(host, comm)
            except Exception as e:
                res = {"ok": False, "error": str(e)}
            self.after(0, lambda: self._icx_done(res))

        threading.Thread(target=work, daemon=True).start()

    def _icx_done(self, res):
        if not res.get("ok"):
            self.icx_info.set(res.get("error") or "실패")
            messagebox.showerror("SNMP 실패", res.get("error") or "실패")
            return
        rows = res.get("rows") or []
        self._icx_rows = rows
        for r in rows:
            self.icx_tree.insert("", END, values=(r.get("ip"), r.get("mac"), r.get("mac_lower"), r.get("iface")))
        self.icx_info.set(
            f"Hostname: {res.get('hostname')}   Model: {res.get('model')}   Version: {res.get('version')}   ARP {len(rows)}건"
        )

    def _icx_csv(self):
        rows = getattr(self, "_icx_rows", [])
        if not rows:
            messagebox.showwarning("안내", "먼저 조회하세요.")
            return
        dest = filedialog.asksaveasfilename(
            title="ICX ARP CSV 저장",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=f"snmp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        if not dest:
            return
        out = Path(dest)
        with open(out, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["IP", "MAC", "VLAN", "MAC (lowercase)"])
            for r in rows:
                w.writerow([r.get("ip"), r.get("mac"), r.get("iface"), r.get("mac_lower")])
        try:
            if RESULTS_SNMP.exists():
                for old in RESULTS_SNMP.glob("snmp_*.csv"):
                    old.unlink()
        except Exception:
            pass
        try:
            os.startfile(str(out.resolve()))
        except Exception:
            pass
        messagebox.showinfo("저장", f"{len(rows)}건 저장\n{out}")

    def _build_sz(self):


        self._clear_page()

        outer = self._scroll_page(padx=20, pady=16)
        self._back_btn(outer)

        Label(
            outer, text="1. SmartZone 정보 보기",
            font=("Segoe UI", 14, "bold"), fg=ACCENT, bg=BG,
        ).pack(anchor="w")
        Label(
            outer, text="SZ API로 AP / BSSID / Switch 리스트를 조회합니다.",
            font=("Segoe UI", 9), fg="#666", bg=BG,
        ).pack(anchor="w", pady=(2, 10))

        form = Frame(outer, bg=CARD, padx=14, pady=12,
                     highlightbackground="#dee2e6", highlightthickness=1)
        form.pack(fill=X)

        self.sz_ip = StringVar()
        self.sz_user = StringVar(value="admin")
        self.sz_pass = StringVar()
        self.sz_ctrl = StringVar(value="수동선택")
        self.sz_api = StringVar()

        fields = Frame(form, bg=CARD)
        fields.pack(fill=X)
        left = Frame(fields, bg=CARD)
        right = Frame(fields, bg=CARD)
        left.pack(side=LEFT, anchor="n")
        right.pack(side=LEFT, anchor="n", padx=(24, 0))

        def row(parent, lbl, var, show=None):
            fr = Frame(parent, bg=CARD)
            fr.pack(fill=X, pady=3)
            Label(fr, text=lbl, width=16, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
            Entry(fr, textvariable=var, font=("Segoe UI", 10), width=20, show=show or "").pack(side=LEFT)

        row(left, "SZ IP/Domain", self.sz_ip)
        row(left, "Username", self.sz_user)
        row(left, "Password", self.sz_pass, show="*")

        fr_c = Frame(right, bg=CARD)
        fr_c.pack(fill=X, pady=3)
        Label(fr_c, text="컨트롤러 버전", width=16, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        self.sz_ctrl_combo = ttk.Combobox(
            fr_c, textvariable=self.sz_ctrl,
            values=list(CONTROLLER_API_MAP.keys()),
            state="readonly", width=18, font=("Segoe UI", 10),
        )
        self.sz_ctrl_combo.pack(side=LEFT)
        self.sz_ctrl_combo.bind("<<ComboboxSelected>>", self._sz_update_api_versions)

        fr_a = Frame(right, bg=CARD)
        fr_a.pack(fill=X, pady=3)
        Label(fr_a, text="API 버전", width=16, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        self.sz_api_combo = ttk.Combobox(
            fr_a, textvariable=self.sz_api, state="readonly", width=18, font=("Segoe UI", 10),
        )
        self.sz_api_combo.pack(side=LEFT)
        self._sz_update_api_versions()

        btn_row = Frame(form, bg=CARD)
        btn_row.pack(fill=X, pady=(10, 0))
        self.sz_run_btn = Button(
            btn_row, text="조회 실행", font=("Segoe UI", 10, "bold"),
            bg=ACCENT, fg="white", relief="flat", padx=16, pady=5,
            command=self._sz_run, cursor="hand2",
        )
        self.sz_run_btn.pack(side=LEFT, padx=(0, 8))
        Button(
            btn_row, text="결과 폴더 열기", font=("Segoe UI", 10),
            bg=BTN_BG, relief="solid", borderwidth=1, padx=12, pady=5,
            command=lambda: self._open_path(RESULTS_SZ), cursor="hand2",
        ).pack(side=LEFT, padx=(0, 8))
        Button(
            btn_row, text="최근 결과 다운로드", font=("Segoe UI", 10),
            bg="#28a745", fg="white", relief="flat", padx=12, pady=5,
            command=lambda: self._download_latest_results(
                RESULTS_SZ, ["_AP.csv", "_BSSID.csv", "_Switch.csv"]
            ),
            cursor="hand2",
        ).pack(side=LEFT)
        self._log_action_btns(btn_row, LOG_SZ, bg=CARD)
        Label(btn_row, text=RESULT_HINT, font=("Segoe UI", 8), fg="#888", bg=CARD).pack(side=LEFT, padx=(8, 0))

        self.sz_summary = StringVar(value="대기 중")
        Label(outer, textvariable=self.sz_summary, font=("Segoe UI", 9), bg=BG, fg="#555").pack(anchor="w", pady=(8, 2))

        # Notebook: AP / BSSID / Switch
        nb = ttk.Notebook(outer)
        nb.pack(fill=BOTH, expand=True, pady=(4, 4))
        self.sz_trees = {}
        tabs = [
            ("AP", AP_ROW_FIELDS),
            ("BSSID(WLAN)", ("deviceName", "apMac", "wlanName", "bssid", "radioid", "ip", "eirp2G", "eirp5G", "eirp6G")),
            ("Switch", ("switchName", "ipAddress", "macAddress", "serialNumber", "model", "status", "firmwareVersion")),
        ]
        for title, cols in tabs:
            fr = Frame(nb, bg=CARD)
            nb.add(fr, text=title)
            tree = ttk.Treeview(fr, columns=cols, show="headings", height=12)
            ap_w = {
                "AP_Name": 130, "IP": 110, "AP_MAC": 130, "serial": 120, "Model": 90,
                "channel2G": 80, "channel5G": 80, "channel6G": 80,
                "status": 80, "config_status": 100, "firmwareVer": 110,
                "airtime2G": 80, "airtime5G": 80, "airtime6G": 80,
                "noise2G": 70, "noise5G": 70, "noise6G": 70,
                "eirp2G": 70, "eirp5G": 70, "eirp6G": 70,
                "Clients": 60, "poePort": 80, "ZoneDomain": 120,
            }
            for c in cols:
                tree.heading(c, text=c)
                tree.column(c, width=ap_w.get(c, 120), anchor="w", stretch=False)
            sb_y = Scrollbar(fr, orient=VERTICAL, command=tree.yview)
            sb_x = Scrollbar(fr, orient=HORIZONTAL, command=tree.xview)
            tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
            tree.grid(row=0, column=0, sticky="nsew")
            sb_y.grid(row=0, column=1, sticky="ns")
            sb_x.grid(row=1, column=0, sticky="ew")
            fr.grid_rowconfigure(0, weight=1)
            fr.grid_columnconfigure(0, weight=1)
            self.sz_trees[title] = (tree, cols)

        self._make_log(outer, "로그", 8)
        self._sz_result = None

    def _sz_update_api_versions(self, event=None):
        ctrl = self.sz_ctrl.get() or "수동선택"
        versions = CONTROLLER_API_MAP.get(ctrl, CONTROLLER_API_MAP["수동선택"])
        self.sz_api_combo["values"] = versions
        if versions:
            self.sz_api.set(versions[0])

    def _sz_run(self):
        host = self.sz_ip.get().strip()
        user = self.sz_user.get().strip()
        pw = self.sz_pass.get()
        api_ver = self.sz_api.get().strip()
        if not host or not user or not pw or not api_ver:
            messagebox.showwarning("안내", "SZ IP / Username / Password / API 버전을 입력하세요.")
            return
        if self._worker and self._worker.is_alive():
            messagebox.showwarning("안내", "이미 실행 중입니다.")
            return

        self.sz_run_btn.config(state=DISABLED)
        self.sz_summary.set("조회 중...")
        self._log(f"\n===== SmartZone 조회: {host} / {api_ver} =====\n")

        def work():
            try:
                client = SmartZoneAPI(host, user, pw, api_ver)

                def log(msg):
                    self._log_queue.put(str(msg) + "\n")

                result = client.collect_all(log=log)
                self._sz_result = result
                if not result.get("ok"):
                    self.after(0, lambda: self.sz_summary.set("실패: " + result.get("error", "")))
                    return

                # CSV 저장
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                prefix = RESULTS_SZ / f"sz_{host}_{ts}"
                ap_rows = [ap_to_row(a) for a in result["aps"]]
                sw_rows = [switch_to_row(s) for s in result["switches"]]
                bssid_rows = result["bssid_rows"]

                ap_fields = list(AP_ROW_FIELDS)
                b_fields = list(bssid_rows[0].keys()) if bssid_rows else [
                    "deviceName", "apMac", "wlanName", "bssid", "radioid", "ip", "eirp2G", "eirp5G", "eirp6G"
                ]
                s_fields = list(sw_rows[0].keys()) if sw_rows else [
                    "switchName", "ipAddress", "macAddress", "serialNumber", "model", "status", "firmwareVersion"
                ]

                p1 = save_csv(Path(str(prefix) + "_AP.csv"), ap_fields, ap_rows)
                p2 = save_csv(Path(str(prefix) + "_BSSID.csv"), b_fields, bssid_rows)
                p3 = save_csv(Path(str(prefix) + "_Switch.csv"), s_fields, sw_rows)
                keep_latest_results(RESULTS_SZ, suffixes=["_AP.csv", "_BSSID.csv", "_Switch.csv"])
                self._save_session_log(LOG_SZ, "sz")
                log(f"CSV 저장:\n  {p1}\n  {p2}\n  {p3}")

                summary = (
                    f"완료 | controller={result.get('controller_version') or 'N/A'} | "
                    f"cluster={result.get('cluster_name') or 'N/A'} | "
                    f"AP={len(ap_rows)} BSSID={len(bssid_rows)} Switch={len(sw_rows)}"
                )
                self.after(0, lambda: self._sz_fill_trees(result, summary))
            except Exception as e:
                self._log_queue.put(f"예외: {e}\n")
                self.after(0, lambda: self.sz_summary.set(f"오류: {e}"))
            finally:
                self.after(0, lambda: self.sz_run_btn.config(state=NORMAL))

        self._worker = threading.Thread(target=work, daemon=True)
        self._worker.start()

    def _sz_fill_trees(self, result, summary):
        self.sz_summary.set(summary)
        # AP
        tree, cols = self.sz_trees["AP"]
        tree.delete(*tree.get_children())
        for ap in result.get("aps") or []:
            r = ap_to_row(ap)
            tree.insert("", END, values=tuple(r.get(c, "") for c in cols))
        # BSSID
        tree, cols = self.sz_trees["BSSID(WLAN)"]
        tree.delete(*tree.get_children())
        for r in result.get("bssid_rows") or []:
            tree.insert("", END, values=tuple(r.get(c, "") for c in cols))
        # Switch
        tree, cols = self.sz_trees["Switch"]
        tree.delete(*tree.get_children())
        for sw in result.get("switches") or []:
            r = switch_to_row(sw)
            tree.insert("", END, values=tuple(r.get(c, "") for c in cols))
        messagebox.showinfo("완료", summary + "\n\nCSV는 results/1_sz_api 폴더에 저장되었습니다.")

    # ------------------------------------------------------------------
    # 메뉴 2: Unleashed 정보 보기
    # ------------------------------------------------------------------

    def _build_uldpsk(self):
        self._clear_page()
        outer = Frame(self, bg=BG, padx=12, pady=10)
        outer.pack(fill=BOTH, expand=True)
        top = Frame(outer, bg=BG)
        top.pack(fill=X)
        self._back_btn(top)
        Label(top, text="5. Unleashed DPSK 관리", font=("Segoe UI", 14, "bold"),
              fg=ACCENT, bg=BG).pack(side=LEFT, padx=12)

        self.udpsk_ip = StringVar()
        self.udpsk_user = StringVar(value="admin")
        self.udpsk_pass = StringVar()
        self.udpsk_search = StringVar()
        self.udpsk_wlan = StringVar()
        self.udpsk_user_name = StringVar()
        self.udpsk_count = StringVar(value="1")
        self.udpsk_vlan = StringVar()
        self._udpsk_cli = None
        self._udpsk_wlans = []
        self._udpsk_all = []

        body = Frame(outer, bg=BG)
        body.pack(fill=BOTH, expand=True, pady=(8, 0))
        side_wrap = Frame(body, bg=BG, width=268)
        side_wrap.pack(side=LEFT, fill=Y, padx=(0, 10))
        side_wrap.pack_propagate(False)
        side_canvas = Canvas(side_wrap, bg=BG, highlightthickness=0, width=248)
        side_sb = Scrollbar(side_wrap, orient=VERTICAL, command=side_canvas.yview)
        side_canvas.configure(yscrollcommand=side_sb.set)
        side_canvas.pack(side=LEFT, fill=BOTH, expand=True)
        side_sb.pack(side=RIGHT, fill=Y)
        side = Frame(side_canvas, bg=BG)
        side_win = side_canvas.create_window((0, 0), window=side, anchor="nw")

        def _side_cfg(_e=None):
            side_canvas.configure(scrollregion=side_canvas.bbox("all"))
            side_canvas.itemconfigure(side_win, width=side_canvas.winfo_width())
        side.bind("<Configure>", _side_cfg)
        side_canvas.bind("<Configure>", _side_cfg)

        def _mw(event):
            delta = -1 if getattr(event, "delta", 0) > 0 else 1
            side_canvas.yview_scroll(delta, "units")
        side.bind("<Enter>", lambda e: side.bind_all("<MouseWheel>", _mw))
        side.bind("<Leave>", lambda e: side.unbind_all("<MouseWheel>"))

        def sl(parent, text):
            Label(parent, text=text, bg=CARD, font=("Segoe UI", 8), fg="#555").pack(anchor="w", pady=(6, 0))

        box1 = LabelFrame(side, text="1. 접속 설정", bg=CARD, fg="#333", font=("Segoe UI", 9, "bold"), padx=8, pady=6)
        box1.pack(fill=X)
        sl(box1, "Unleashed IP/Domain")
        Entry(box1, textvariable=self.udpsk_ip, font=("Segoe UI", 10)).pack(fill=X)
        sl(box1, "Username")
        Entry(box1, textvariable=self.udpsk_user, font=("Segoe UI", 10)).pack(fill=X)
        sl(box1, "Password")
        Entry(box1, textvariable=self.udpsk_pass, show="*", font=("Segoe UI", 10)).pack(fill=X)
        Button(box1, text="로그인 & DPSK 조회", font=("Segoe UI", 9, "bold"), bg=LINK, fg="white",
               relief="flat", pady=5, command=self._udpsk_refresh, cursor="hand2").pack(fill=X, pady=(10, 4))

        box2 = LabelFrame(side, text="2. DPSK 생성", bg=CARD, fg="#333", font=("Segoe UI", 9, "bold"), padx=8, pady=6)
        box2.pack(fill=X, pady=(10, 0))
        sl(box2, "WLAN 선택 (DPSK Enabled)")
        self.udpsk_wlan_combo = ttk.Combobox(box2, textvariable=self.udpsk_wlan, state="readonly")
        self.udpsk_wlan_combo.pack(fill=X)
        sl(box2, "Number of DPSKs")
        Entry(box2, textvariable=self.udpsk_count, font=("Segoe UI", 10)).pack(fill=X)
        sl(box2, "User Name")
        Entry(box2, textvariable=self.udpsk_user_name, font=("Segoe UI", 10)).pack(fill=X)
        sl(box2, "VLAN ID (선택)")
        Entry(box2, textvariable=self.udpsk_vlan, font=("Segoe UI", 10)).pack(fill=X)
        Button(box2, text="DPSK 생성하기", font=("Segoe UI", 9, "bold"), bg="#28a745", fg="white",
               relief="flat", pady=5, command=self._udpsk_create, cursor="hand2").pack(fill=X, pady=(10, 4))

        main = Frame(body, bg=CARD, padx=10, pady=8, highlightbackground="#dee2e6", highlightthickness=1)
        main.pack(side=LEFT, fill=BOTH, expand=True)
        filt = Frame(main, bg=CARD)
        filt.pack(fill=X)
        Label(filt, text="검색:", bg=CARD).pack(side=LEFT)
        Entry(filt, textvariable=self.udpsk_search, width=28).pack(side=LEFT, padx=4)
        Button(filt, text="검색", bg=LINK, fg="white", relief="flat", padx=10,
               command=self._udpsk_fill).pack(side=LEFT)
        head = Frame(main, bg=CARD)
        head.pack(fill=X, pady=(8, 2))
        Label(head, text="DPSK 목록", font=("Segoe UI", 10, "bold"), bg=CARD, fg=LINK).pack(side=LEFT)
        self.udpsk_count_lbl = StringVar(value="총 0개")
        Label(head, textvariable=self.udpsk_count_lbl, bg=CARD, fg="#555").pack(side=RIGHT)
        Button(head, text="CSV 다운로드", bg="#28a745", fg="white", relief="flat", padx=8,
               command=self._udpsk_csv).pack(side=RIGHT, padx=4)
        Button(head, text="선택 항목 삭제", bg="#dc3545", fg="white", relief="flat", padx=8,
               command=self._udpsk_delete).pack(side=RIGHT, padx=4)
        Button(head, text="결과 폴더 열기", bg=BTN_BG, relief="solid", borderwidth=1, padx=8,
               command=lambda: self._open_path(RESULTS_UDPSK)).pack(side=RIGHT, padx=4)
        Button(head, text="최근 결과 다운로드", bg="#28a745", fg="white", relief="flat", padx=8,
               command=lambda: self._download_latest_results(RESULTS_UDPSK)).pack(side=RIGHT, padx=4)

        cols = ("wlan", "dpsk_len", "shared_dpsk", "shared_num", "user", "psk", "vlan",
                "clients", "usage", "mac", "period", "status", "start_point",
                "limit_dpsk", "limit_num", "created", "expires")
        tree_fr = Frame(main, bg=CARD)
        tree_fr.pack(fill=BOTH, expand=True)
        self.udpsk_tree = ttk.Treeview(tree_fr, columns=cols, show="headings", height=16, selectmode="extended")
        headers = {
            "wlan": "WLAN 이름", "dpsk_len": "DPSK 길이", "shared_dpsk": "공유 DPSK",
            "shared_num": "공유 수", "user": "User Name", "psk": "Passphrase", "vlan": "VLAN",
            "clients": "사용 단말수", "usage": "Usage", "mac": "MAC 주소",
            "period": "사용가능기간", "status": "상태", "start_point": "시작 방식",
            "limit_dpsk": "DPSK 제한", "limit_num": "제한 수", "created": "생성일시", "expires": "만료일시",
        }
        widths = {
            "wlan": 150, "dpsk_len": 70, "shared_dpsk": 70, "shared_num": 50,
            "user": 120, "psk": 100, "vlan": 50, "clients": 70, "usage": 50,
            "mac": 120, "period": 80, "status": 90, "start_point": 70,
            "limit_dpsk": 70, "limit_num": 50, "created": 130, "expires": 130,
        }
        for c in cols:
            self.udpsk_tree.heading(c, text=headers[c])
            self.udpsk_tree.column(c, width=widths[c], anchor="w")
        self.udpsk_tree.tag_configure("active", foreground="#198754")
        self.udpsk_tree.tag_configure("expired", foreground="#dc3545")
        ysb = Scrollbar(tree_fr, command=self.udpsk_tree.yview)
        xsb = Scrollbar(tree_fr, orient=HORIZONTAL, command=self.udpsk_tree.xview)
        self.udpsk_tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.udpsk_tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        tree_fr.grid_rowconfigure(0, weight=1)
        tree_fr.grid_columnconfigure(0, weight=1)
        self.udpsk_status = StringVar(value="Unleashed IP / 계정 입력 후 조회하세요.")
        Label(outer, textvariable=self.udpsk_status, font=("Segoe UI", 9), bg=BG, fg="#555").pack(anchor="w", pady=(6, 0))

    def _udpsk_refresh(self):
        host = self.udpsk_ip.get().strip()
        user = self.udpsk_user.get().strip()
        pw = self.udpsk_pass.get()
        if not (host and user and pw):
            messagebox.showwarning("안내", "Unleashed IP / Username / Password 를 입력하세요.")
            return
        try:
            cli = UnleashedAPI(host, user, pw)
            ok, msg = cli.login()
            if not ok:
                messagebox.showerror("로그인 실패", msg)
                return
            self._udpsk_cli = cli
            wlans = cli.fetch_dpsk_wlans()
            self._udpsk_wlans = wlans
            wmap = {w.get("id"): w for w in wlans}
            labels = [f"{w.get('name')}  ({w.get('ssid')})" for w in wlans]
            self.udpsk_wlan_combo["values"] = labels
            if labels and not self.udpsk_wlan.get():
                self.udpsk_wlan.set(labels[0])
            self._udpsk_all = cli.fetch_dpsk_list(wmap)
            self._udpsk_fill()
            self.udpsk_status.set(f"{msg}  /  DPSK {len(self._udpsk_all)}건 / WLAN {len(wlans)}개")
        except Exception as e:
            messagebox.showerror("오류", str(e))

    def _udpsk_fill(self):
        tree = getattr(self, "udpsk_tree", None)
        if tree is None:
            return
        tree.delete(*tree.get_children())
        q = (self.udpsk_search.get() or "").strip().lower()
        n = 0
        for r in self._udpsk_all:
            blob = " ".join(str(r.get(k, "")) for k in ("wlan", "user", "psk", "mac")).lower()
            if q and q not in blob:
                continue
            iid = r.get("id") or f"row{n}"
            tag = "expired" if "만료" == r.get("status") else "active"
            tree.insert("", END, iid=str(iid), tags=(tag,), values=(
                r.get("wlan"), r.get("dpsk_len"), r.get("shared_dpsk"), r.get("shared_num"),
                r.get("user"), r.get("psk"), r.get("vlan"), r.get("clients"), r.get("usage"),
                r.get("mac"), r.get("period"), r.get("status"), r.get("start_point"),
                r.get("limit_dpsk"), r.get("limit_num"), r.get("created"), r.get("expires"),
            ))
            n += 1
        self.udpsk_count_lbl.set(f"총 {n}개")

    def _udpsk_wlan_id(self):
        label = self.udpsk_wlan.get()
        for w in self._udpsk_wlans:
            if f"{w.get('name')}  ({w.get('ssid')})" == label or w.get("id") == label:
                return w.get("id")
        return ""

    def _udpsk_create(self):
        cli = self._udpsk_cli
        wid = self._udpsk_wlan_id()
        if not cli or not wid:
            messagebox.showwarning("안내", "먼저 조회 후 DPSK WLAN을 선택하세요.")
            return
        try:
            amount = int((self.udpsk_count.get() or "1").strip())
        except ValueError:
            messagebox.showwarning("안내", "개수는 숫자여야 합니다.")
            return
        username = self.udpsk_user_name.get().strip()
        if amount == 1 and username:
            if any(r.get("wlan_id") == wid and (r.get("user") or "").lower() == username.lower() for r in self._udpsk_all):
                if not messagebox.askyesno("중복", f"사용자 '{username}' 이 이미 있습니다. 강제 생성할까요?"):
                    return
        ok, text = cli.create_dpsk(wid, username, self.udpsk_vlan.get().strip(), amount)
        if ok:
            messagebox.showinfo("완료", f"{amount}건 생성 요청 완료")
            self._udpsk_refresh()
        else:
            messagebox.showerror("생성 실패", text[:400])

    def _udpsk_delete(self):
        cli = self._udpsk_cli
        if not cli:
            return
        sel = self.udpsk_tree.selection()
        if not sel:
            messagebox.showwarning("안내", "삭제할 항목을 선택하세요.")
            return
        if not messagebox.askyesno("확인", f"{len(sel)}건을 삭제할까요?"):
            return
        ok, text = cli.delete_dpsk(list(sel))
        if ok:
            messagebox.showinfo("삭제", f"{len(sel)}건 삭제 요청 완료")
            self._udpsk_refresh()
        else:
            messagebox.showerror("삭제 실패", text[:400])

    def _udpsk_csv(self):
        q = (self.udpsk_search.get() or "").strip().lower()
        rows = []
        for r in self._udpsk_all:
            blob = " ".join(str(r.get(k, "")) for k in ("wlan", "user", "psk", "mac")).lower()
            if q and q not in blob:
                continue
            rows.append(r)
        if not rows:
            messagebox.showwarning("안내", "저장할 목록이 없습니다.")
            return
        dest = filedialog.asksaveasfilename(
            title="Unleashed DPSK CSV 저장",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile=f"udpsk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
        )
        if not dest:
            return
        out = Path(dest)
        with open(out, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["WLAN 이름", "DPSK 길이", "공유 DPSK", "공유 수", "User Name", "Passphrase",
                        "VLAN", "사용 단말수", "Usage", "MAC 주소", "사용가능기간", "상태",
                        "시작 방식", "DPSK 제한", "제한 수", "생성일시", "만료일시"])
            for r in rows:
                w.writerow([
                    r.get("wlan"), r.get("dpsk_len"), r.get("shared_dpsk"), r.get("shared_num"),
                    r.get("user"), r.get("psk"), r.get("vlan"), r.get("clients"), r.get("usage"),
                    r.get("mac"), r.get("period"), r.get("status"), r.get("start_point"),
                    r.get("limit_dpsk"), r.get("limit_num"), r.get("created"), r.get("expires"),
                ])
        try:
            if RESULTS_UDPSK.exists():
                for old in RESULTS_UDPSK.glob("udpsk_*.csv"):
                    old.unlink()
        except Exception:
            pass
        try:
            os.startfile(str(out.resolve()))
        except Exception:
            pass
        messagebox.showinfo("저장", f"{len(rows)}건 저장\n{out}")

    def _build_unleashed(self):

        self._clear_page()

        outer = self._scroll_page(padx=20, pady=16)
        self._back_btn(outer)

        Label(
            outer, text="2. 언리시드 정보 보기",
            font=("Segoe UI", 14, "bold"), fg=ACCENT, bg=BG,
        ).pack(anchor="w")
        Label(
            outer,
            text="Unleashed 마스터 AP에 로그인하여 AP / WLAN(VAP) 통계를 조회합니다.",
            font=("Segoe UI", 9), fg="#666", bg=BG,
        ).pack(anchor="w", pady=(2, 10))

        form = Frame(outer, bg=CARD, padx=14, pady=12,
                     highlightbackground="#dee2e6", highlightthickness=1)
        form.pack(fill=X)

        self.ul_ip = StringVar()
        self.ul_user = StringVar(value="admin")
        self.ul_pass = StringVar()

        def row(lbl, var, show=None):
            fr = Frame(form, bg=CARD)
            fr.pack(fill=X, pady=3)
            Label(fr, text=lbl, width=22, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
            Entry(fr, textvariable=var, font=("Segoe UI", 10), width=20, show=show or "").pack(side=LEFT)

        row("Unleashed IP/Domain", self.ul_ip)
        row("Username", self.ul_user)
        row("Password", self.ul_pass, show="*")

        btn_row = Frame(form, bg=CARD)
        btn_row.pack(fill=X, pady=(10, 0))
        self.ul_run_btn = Button(
            btn_row, text="조회 실행", font=("Segoe UI", 10, "bold"),
            bg=ACCENT, fg="white", relief="flat", padx=16, pady=5,
            command=self._ul_run, cursor="hand2",
        )
        self.ul_run_btn.pack(side=LEFT, padx=(0, 8))
        Button(
            btn_row, text="결과 폴더 열기", font=("Segoe UI", 10),
            bg=BTN_BG, relief="solid", borderwidth=1, padx=12, pady=5,
            command=lambda: self._open_path(RESULTS_UNLEASHED), cursor="hand2",
        ).pack(side=LEFT, padx=(0, 8))
        Button(
            btn_row, text="최근 결과 다운로드", font=("Segoe UI", 10),
            bg="#28a745", fg="white", relief="flat", padx=12, pady=5,
            command=lambda: self._download_latest_results(
                RESULTS_UNLEASHED, ["_AP.csv", "_WLAN.csv"]
            ),
            cursor="hand2",
        ).pack(side=LEFT)
        self._log_action_btns(btn_row, LOG_UNLEASHED, bg=CARD)
        Label(btn_row, text=RESULT_HINT, font=("Segoe UI", 8), fg="#888", bg=CARD).pack(side=LEFT, padx=(8, 0))

        self.ul_summary = StringVar(value="대기 중")
        Label(outer, textvariable=self.ul_summary, font=("Segoe UI", 9), bg=BG, fg="#555").pack(
            anchor="w", pady=(8, 2)
        )

        nb = ttk.Notebook(outer)
        nb.pack(fill=BOTH, expand=True, pady=(4, 4))
        self.ul_trees = {}
        tabs = [
            ("AP", (
                "mac", "ap-name", "model", "ip", "netmask", "gateway",
                "serial-number", "firmware-version", "num-sta",
                "eth0_Physical", "2G_ch", "5G_ch", "6G_ch",
            )),
            ("BSSID(WLAN)", (
                "BSSID", "SSID", "Radio_Band", "AP_mac",
                "Radio_Type", "VAP_Up_Status", "Channel",
            )),
        ]
        for title, cols in tabs:
            fr = Frame(nb, bg=CARD)
            nb.add(fr, text=title)
            tree = ttk.Treeview(fr, columns=cols, show="headings", height=12)
            for c in cols:
                tree.heading(c, text="BSSID(WLAN)" if c == "BSSID" else c)
                tree.column(c, width=140, anchor="w")
            sb_y = Scrollbar(fr, orient=VERTICAL, command=tree.yview)
            sb_x = Scrollbar(fr, orient=HORIZONTAL, command=tree.xview)
            tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
            tree.grid(row=0, column=0, sticky="nsew")
            sb_y.grid(row=0, column=1, sticky="ns")
            sb_x.grid(row=1, column=0, sticky="ew")
            fr.grid_rowconfigure(0, weight=1)
            fr.grid_columnconfigure(0, weight=1)
            self.ul_trees[title] = (tree, cols)

        self._make_log(outer, "로그", 8)

    def _ul_run(self):
        host = self.ul_ip.get().strip()
        user = self.ul_user.get().strip()
        pw = self.ul_pass.get()
        if not host or not user or not pw:
            messagebox.showwarning("안내", "Unleashed IP / Username / Password 를 입력하세요.")
            return
        if self._worker and self._worker.is_alive():
            messagebox.showwarning("안내", "이미 실행 중입니다.")
            return

        self.ul_run_btn.config(state=DISABLED)
        self.ul_summary.set("조회 중...")
        self._log(f"\n===== Unleashed 조회: {host} =====\n")

        def work():
            try:
                client = UnleashedAPI(host, user, pw)

                def log(msg):
                    self._log_queue.put(str(msg) + "\n")

                result = client.collect_all(log=log)
                if not result.get("ok"):
                    self.after(0, lambda: self.ul_summary.set("실패: " + result.get("error", "")))
                    return

                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                RESULTS_UNLEASHED.mkdir(parents=True, exist_ok=True)
                ap_rows = result.get("aps") or []
                wlan_rows = result.get("wlans") or []
                ap_fields = [
                    "mac", "ap-name", "model", "ip", "netmask", "gateway",
                    "serial-number", "firmware-version", "num-sta",
                    "eth0_Physical", "eth1_Physical", "2G_ch", "5G_ch", "6G_ch",
                ]
                wlan_fields = [
                    "BSSID", "SSID", "Radio_Band", "AP_mac",
                    "Radio_Type", "VAP_Up_Status", "Channel",
                ]
                p1 = ul_save_csv(
                    RESULTS_UNLEASHED / f"ul_{host}_{ts}_AP.csv", ap_fields, ap_rows
                )
                p2 = ul_save_csv(
                    RESULTS_UNLEASHED / f"ul_{host}_{ts}_WLAN.csv", wlan_fields, wlan_rows
                )
                keep_latest_results(RESULTS_UNLEASHED, suffixes=["_AP.csv", "_WLAN.csv"])
                self._save_session_log(LOG_UNLEASHED, "ul")
                log(f"CSV 저장:\n  {p1}\n  {p2}")

                summary = f"완료 | AP={len(ap_rows)} WLAN={len(wlan_rows)}"
                self.after(0, lambda: self._ul_fill_trees(result, summary))
            except Exception as e:
                self._log_queue.put(f"예외: {e}\n")
                self.after(0, lambda: self.ul_summary.set(f"오류: {e}"))
            finally:
                self.after(0, lambda: self.ul_run_btn.config(state=NORMAL))

        self._worker = threading.Thread(target=work, daemon=True)
        self._worker.start()

    def _ul_fill_trees(self, result, summary):
        self.ul_summary.set(summary)
        tree, cols = self.ul_trees["AP"]
        tree.delete(*tree.get_children())
        for r in result.get("aps") or []:
            tree.insert("", END, values=tuple(r.get(c, "") for c in cols))
        tree, cols = self.ul_trees["BSSID(WLAN)"]
        tree.delete(*tree.get_children())
        for r in result.get("wlans") or []:
            tree.insert("", END, values=tuple(r.get(c, "") for c in cols))
        messagebox.showinfo(
            "완료",
            summary + "\n\nCSV/XML 은 results/2_unleashed_api 폴더에 저장되었습니다.",
        )



    # ------------------------------------------------------------------
    # 메뉴 9: AP → SZ 펌웨어 업그레이드 + 연동
    # ------------------------------------------------------------------
    def _build_fwsz(self):
        self._clear_page()
        outer = self._scroll_page(padx=20, pady=16)
        self._back_btn(outer)
        Label(outer, text="9. AP → SZ 펌웨어 업그레이드 + 연동", font=("Segoe UI", 14, "bold"), fg=ACCENT, bg=BG).pack(anchor="w")
        Label(outer, text="SZ HTTPS로 model_version.rcks 다운로드 → 완료 후 set scg / hostname / IP",
              font=("Segoe UI", 9), fg="#666", bg=BG).pack(anchor="w", pady=(2, 8))

        form = Frame(outer, bg=CARD, padx=14, pady=12, highlightbackground="#dee2e6", highlightthickness=1)
        form.pack(fill=X)

        self.fwsz_api_ip = StringVar()
        self.fwsz_user = StringVar(value="admin")
        self.fwsz_pass = StringVar()
        self.fwsz_ctrl = StringVar(value="수동선택")
        self.fwsz_api = StringVar()
        self.fwsz_zone = StringVar()
        self.fwsz_ver = StringVar()
        self.fwsz_user2 = StringVar(value="")
        self.fwsz_pass2 = StringVar(value="")
        self._fwsz_zones = []

        fields = Frame(form, bg=CARD)
        fields.pack(fill=X)
        left = Frame(fields, bg=CARD)
        mid = Frame(fields, bg=CARD)
        acc2 = LabelFrame(fields, text="2차 기본 계정 (초기 비번 변경시 or CSV 실패시)", bg=CARD, fg="#444",
                          font=("Segoe UI", 8), padx=10, pady=6)
        left.pack(side=LEFT, anchor="n")
        mid.pack(side=LEFT, anchor="n", padx=(24, 0))
        acc2.pack(side=RIGHT, anchor="n")
        Label(acc2, text="Username", width=10, anchor="w", bg=CARD, font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w", pady=2)
        Entry(acc2, textvariable=self.fwsz_user2, width=16, font=("Segoe UI", 10)).grid(row=0, column=1, sticky="w", pady=2)
        Label(acc2, text="Password", width=10, anchor="w", bg=CARD, font=("Segoe UI", 10)).grid(row=1, column=0, sticky="w", pady=2)
        self.fwsz_pass2_entry = Entry(acc2, textvariable=self.fwsz_pass2, width=16, font=("Segoe UI", 10), show="*")
        self.fwsz_pass2_entry.grid(row=1, column=1, sticky="w", pady=2)
        self.fwsz_pass2_tip = BalloonTip(self.fwsz_pass2_entry, PW_RULE_KO, enabled=self._alt_pw_complex_on())
        self.fwsz_pass2_show = BooleanVar(value=False)
        Checkbutton(acc2, text="표시", variable=self.fwsz_pass2_show, bg=CARD,
                    command=lambda: self.fwsz_pass2_entry.config(show="" if self.fwsz_pass2_show.get() else "*")).grid(row=1, column=2, sticky="w", padx=(6, 0))
        self._alt_pw_hint_row(acc2)

        def row(parent, lbl, var, show=None):
            fr = Frame(parent, bg=CARD)
            fr.pack(fill=X, pady=3)
            Label(fr, text=lbl, width=16, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
            Entry(fr, textvariable=var, font=("Segoe UI", 10), width=20, show=show or "").pack(side=LEFT)

        row(left, "SZ IP/Domain", self.fwsz_api_ip)
        row(left, "Username", self.fwsz_user)
        row(left, "Password", self.fwsz_pass, show="*")
        fr_z = Frame(left, bg=CARD)
        fr_z.pack(fill=X, pady=3)
        Label(fr_z, text="Zone", width=16, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        self.fwsz_zone_combo = ttk.Combobox(fr_z, textvariable=self.fwsz_zone, state="readonly", width=20, font=("Segoe UI", 10))
        self.fwsz_zone_combo.pack(side=LEFT)
        self.fwsz_zone_combo.bind("<<ComboboxSelected>>", lambda e: self._fwsz_load_fw())

        fr_c = Frame(mid, bg=CARD)
        fr_c.pack(fill=X, pady=3)
        Label(fr_c, text="컨트롤러 버전", width=16, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        self.fwsz_ctrl_combo = ttk.Combobox(
            fr_c, textvariable=self.fwsz_ctrl,
            values=list(CONTROLLER_API_MAP.keys()),
            state="readonly", width=20, font=("Segoe UI", 10),
        )
        self.fwsz_ctrl_combo.pack(side=LEFT)
        self.fwsz_ctrl_combo.bind("<<ComboboxSelected>>", lambda e: self._fwsz_update_api())

        fr_a = Frame(mid, bg=CARD)
        fr_a.pack(fill=X, pady=3)
        Label(fr_a, text="API 버전", width=16, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        self.fwsz_api_combo = ttk.Combobox(fr_a, textvariable=self.fwsz_api, state="readonly", width=20, font=("Segoe UI", 10))
        self.fwsz_api_combo.pack(side=LEFT)
        self._fwsz_update_api()

        fr_v = Frame(mid, bg=CARD)
        fr_v.pack(fill=X, pady=3)
        Label(fr_v, text="Zone 펌웨어 버전", width=16, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        self.fwsz_ver_combo = ttk.Combobox(fr_v, textvariable=self.fwsz_ver, width=20, font=("Segoe UI", 10))
        self.fwsz_ver_combo.pack(side=LEFT)

        brow = Frame(form, bg=CARD)
        brow.pack(fill=X, pady=(8, 0))
        Button(brow, text="Zone / 펌웨어 목록", font=("Segoe UI", 10, "bold"), bg=LINK, fg="white",
               relief="flat", padx=10, pady=4, command=self._fwsz_fetch_meta, cursor="hand2").pack(side=LEFT, padx=(0, 6))
        Button(brow, text="결과 폴더", font=("Segoe UI", 10),
               bg=BTN_BG, relief="solid", borderwidth=1, padx=10, pady=5,
               command=lambda: self._open_path(RESULTS_SZFW), cursor="hand2").pack(side=LEFT)
        self._latest_result_btn(brow, RESULTS_SZFW)
        self._log_action_btns(brow, LOG_SZFW, bg=CARD)
        Label(brow, text=RESULT_HINT, font=("Segoe UI", 8), fg="#888", bg=CARD).pack(side=LEFT, padx=(8, 0))

        rule = Frame(outer, bg=CARD, padx=12, pady=8, highlightbackground="#dee2e6", highlightthickness=1)
        rule.pack(fill=X, pady=(8, 4))
        Label(rule, text="AP Registration Rule (Default Zone 방지)", font=("Segoe UI", 10, "bold"), bg=CARD).pack(anchor="w")
        Label(
            rule,
            text="SZ에 Rule을 만든 뒤, Provision Tag면 AP에 set provisioning-tag 를 넣고 연동합니다.",
            font=("Segoe UI", 8), fg="#666", bg=CARD,
        ).pack(anchor="w")
        rbtn = Frame(rule, bg=CARD)
        rbtn.pack(fill=X, pady=(4, 4))
        Button(rbtn, text="Rule 조회", font=("Segoe UI", 10), bg=LINK, fg="white", relief="flat",
               padx=10, pady=4, command=self._fwsz_rule_list, cursor="hand2").pack(side=LEFT, padx=(0, 6))
        Button(rbtn, text="Rule 생성", font=("Segoe UI", 10), bg="#28a745", fg="white", relief="flat",
               padx=10, pady=4, command=self._fwsz_rule_create, cursor="hand2").pack(side=LEFT, padx=(0, 6))
        Button(rbtn, text="Rule 삭제", font=("Segoe UI", 10), bg=ACCENT, fg="white", relief="flat",
               padx=10, pady=4, command=self._fwsz_rule_delete, cursor="hand2").pack(side=LEFT)
        rlist = Frame(rule, bg=CARD)
        rlist.pack(fill=X)
        cols_r = ("priority", "desc", "type", "zone", "value")
        self.fwsz_rule_tree = ttk.Treeview(rlist, columns=cols_r, show="headings", height=4)
        for c, w, h in (
            ("priority", 50, "우선"),
            ("desc", 160, "설명"),
            ("type", 120, "유형"),
            ("zone", 140, "Zone"),
            ("value", 220, "값"),
        ):
            self.fwsz_rule_tree.heading(c, text=h)
            self.fwsz_rule_tree.column(c, width=w, anchor="w")
        ysr = Scrollbar(rlist, orient=VERTICAL, command=self.fwsz_rule_tree.yview)
        self.fwsz_rule_tree.configure(yscrollcommand=ysr.set)
        self.fwsz_rule_tree.pack(side=LEFT, fill=X, expand=True)
        ysr.pack(side=LEFT, fill=Y)
        self.fwsz_rule_tree.bind("<<TreeviewSelect>>", lambda e: self._fwsz_rule_on_select())
        self._fwsz_rules = []

        rform = Frame(rule, bg=CARD)
        rform.pack(fill=X, pady=(6, 0))
        self.fwsz_rule_desc = StringVar()
        self.fwsz_rule_type = StringVar(value="ProvisionTag")
        self.fwsz_rule_from = StringVar()
        self.fwsz_rule_to = StringVar()
        self.fwsz_rule_net = StringVar()
        self.fwsz_rule_mask = StringVar(value="255.255.255.0")
        self.fwsz_rule_tag = StringVar()
        self.fwsz_apply_tag = BooleanVar(value=True)
        Label(rform, text="설명", bg=CARD, font=("Segoe UI", 9)).pack(side=LEFT)
        Entry(rform, textvariable=self.fwsz_rule_desc, width=16, font=("Segoe UI", 9)).pack(side=LEFT, padx=(4, 8))
        Label(rform, text="유형", bg=CARD, font=("Segoe UI", 9)).pack(side=LEFT)
        self.fwsz_rule_type_combo = ttk.Combobox(
            rform, textvariable=self.fwsz_rule_type,
            values=["ProvisionTag", "IPAddressRange", "Subnet"],
            state="readonly", width=16, font=("Segoe UI", 9),
        )
        self.fwsz_rule_type_combo.pack(side=LEFT, padx=(4, 8))
        self.fwsz_rule_type_combo.bind("<<ComboboxSelected>>", lambda e: self._fwsz_rule_toggle_fields())

        self.fwsz_rule_fields = Frame(rform, bg=CARD)
        self.fwsz_rule_fields.pack(side=LEFT, fill=X, expand=True)

        self.fwsz_rule_fr_tag = Frame(self.fwsz_rule_fields, bg=CARD)
        Label(self.fwsz_rule_fr_tag, text="Tag", bg=CARD, font=("Segoe UI", 9)).pack(side=LEFT)
        Entry(self.fwsz_rule_fr_tag, textvariable=self.fwsz_rule_tag, width=18, font=("Segoe UI", 9)).pack(side=LEFT, padx=(4, 0))

        self.fwsz_rule_fr_ip = Frame(self.fwsz_rule_fields, bg=CARD)
        Label(self.fwsz_rule_fr_ip, text="From IP", bg=CARD, font=("Segoe UI", 9)).pack(side=LEFT)
        Entry(self.fwsz_rule_fr_ip, textvariable=self.fwsz_rule_from, width=14, font=("Segoe UI", 9)).pack(side=LEFT, padx=(4, 8))
        Label(self.fwsz_rule_fr_ip, text="To IP", bg=CARD, font=("Segoe UI", 9)).pack(side=LEFT)
        Entry(self.fwsz_rule_fr_ip, textvariable=self.fwsz_rule_to, width=14, font=("Segoe UI", 9)).pack(side=LEFT, padx=(4, 0))

        self.fwsz_rule_fr_net = Frame(self.fwsz_rule_fields, bg=CARD)
        Label(self.fwsz_rule_fr_net, text="Network Address", bg=CARD, font=("Segoe UI", 9)).pack(side=LEFT)
        Entry(self.fwsz_rule_fr_net, textvariable=self.fwsz_rule_net, width=14, font=("Segoe UI", 9)).pack(side=LEFT, padx=(4, 8))
        Label(self.fwsz_rule_fr_net, text="Subnet Mask", bg=CARD, font=("Segoe UI", 9)).pack(side=LEFT)
        Entry(self.fwsz_rule_fr_net, textvariable=self.fwsz_rule_mask, width=14, font=("Segoe UI", 9)).pack(side=LEFT, padx=(4, 0))

        self.fwsz_apply_tag_chk = Checkbutton(
            rule, text="업그레이드+연동 시 AP에 provisioning-tag 적용",
            variable=self.fwsz_apply_tag, bg=CARD, font=("Segoe UI", 9),
        )
        self.fwsz_apply_tag_chk.pack(anchor="w", pady=(4, 0))
        self._fwsz_rule_toggle_fields()

        box = Frame(outer, bg=CARD, padx=12, pady=8, highlightbackground="#dee2e6", highlightthickness=1)
        box.pack(fill=X, pady=(8, 4))
        Label(box, text="AP 리스트",
              font=("Segoe UI", 10, "bold"), bg=CARD).pack(anchor="w")
        row1 = Frame(box, bg=CARD)
        row1.pack(fill=X, pady=4)
        Button(row1, text="샘플 다운로드 (CSV)", font=("Segoe UI", 10), bg="#28a745", fg="white", relief="flat",
               padx=10, pady=3, command=self._download_sample).pack(side=LEFT, padx=(0, 6))
        Button(row1, text="파일 업로드 (CSV)", font=("Segoe UI", 10), bg=LINK, fg="white", relief="flat",
               padx=10, pady=3, command=self._browse_csv).pack(side=LEFT)
        if not hasattr(self, "csv_path_var"):
            self.csv_path_var = StringVar()
        if not hasattr(self, "csv_info_var"):
            self.csv_info_var = StringVar(value="업로드된 파일 없음")
        Label(box, textvariable=self.csv_path_var, font=("Segoe UI", 8), bg=CARD).pack(anchor="w")
        Label(box, textvariable=self.csv_info_var, font=("Segoe UI", 8), bg=CARD, fg="#666").pack(anchor="w")
        prev = Frame(box, bg=CARD)
        prev.pack(fill=X)
        cols = ("ip", "user", "pass", "new_ip", "subnet", "gw", "sz", "hostname")
        self.csv_tree = ttk.Treeview(prev, columns=cols, show="headings", height=4)
        for c in cols:
            self.csv_tree.heading(c, text=c)
            self.csv_tree.column(c, width=90, anchor="w")
        ysb = Scrollbar(prev, orient=VERTICAL, command=self.csv_tree.yview)
        xsb = Scrollbar(prev, orient=HORIZONTAL, command=self.csv_tree.xview)
        self.csv_tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.csv_tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        prev.grid_rowconfigure(0, weight=1)
        prev.grid_columnconfigure(0, weight=1)
        if getattr(self, "_csv_rows", None):
            self._fill_csv_preview(self._csv_rows)

        run = Frame(outer, bg=BG)
        run.pack(fill=X, pady=6)
        self.fwsz_run_btn = Button(
            run, text="▶ 업그레이드 + SZ 연동실행(자동IP → 고정IP)",
            font=("Segoe UI", 10, "bold"),
            bg=ACCENT, fg="white", relief="flat", padx=12, pady=5,
            command=lambda: self._fwsz_start(change_ip=True), cursor="hand2",
        )
        self.fwsz_run_btn.pack(side=LEFT, padx=(0, 8))
        self.fwsz_run_keep_btn = Button(
            run, text="▶ 업그레이드 + SZ 연동실행(IP변경안함)",
            font=("Segoe UI", 10, "bold"),
            bg="#17a2b8", fg="white", relief="flat", padx=12, pady=5,
            command=lambda: self._fwsz_start(change_ip=False), cursor="hand2",
        )
        self.fwsz_run_keep_btn.pack(side=LEFT, padx=(0, 8))
        self.fwsz_stop_btn = Button(run, text="중지", bg="#6c757d", fg="white", relief="flat", padx=12, pady=5,
                                    command=self._stop_batch, state=DISABLED)
        self.fwsz_stop_btn.pack(side=LEFT)
        self.fwsz_status = StringVar(value="대기 — SZ 펌웨어 버전과 AP CSV를 준비한 뒤 실행")
        Label(outer, textvariable=self.fwsz_status, font=("Segoe UI", 9), bg=BG, fg="#555").pack(anchor="w")
        self.progress = ttk.Progressbar(outer, mode="determinate")
        self.progress.pack(fill=X, pady=4)
        self._make_log(outer, "실행 로그", 8)

    def _fwsz_update_api(self):
        vers = CONTROLLER_API_MAP.get(self.fwsz_ctrl.get(), CONTROLLER_API_MAP.get("수동선택", []))
        if hasattr(self, "fwsz_api_combo"):
            self.fwsz_api_combo["values"] = vers
        if vers and self.fwsz_api.get() not in vers:
            self.fwsz_api.set(vers[0])

    def _fwsz_fetch_meta(self):
        host = self.fwsz_api_ip.get().strip()
        user = self.fwsz_user.get().strip()
        pw = self.fwsz_pass.get()
        api = self.fwsz_api.get().strip()
        if not (host and user and pw and api):
            messagebox.showwarning("안내", "SZ API IP / Username / Password / API 버전을 입력하세요.")
            return
        try:
            api_cli = SmartZoneAPI(host, user, pw, api)
            ok, msg = api_cli.login()
            if not ok:
                messagebox.showerror("SZ 로그인 실패", msg)
                return
            zones = api_cli.fetch_zones()
            self._fwsz_zones = zones
            names = []
            for z in zones:
                nm = (z.get("name") or z.get("id") or "").strip()
                if nm:
                    names.append(nm)
            self.fwsz_zone_combo["values"] = names
            if names:
                self.fwsz_zone.set(names[0])
                self._fwsz_api_cli = api_cli
                self._fwsz_load_fw()
            self._log(f"SZ Zone {len(names)}개 조회\n")
        except Exception as e:
            messagebox.showerror("오류", str(e))

    def _fwsz_load_fw(self):
        api_cli = getattr(self, "_fwsz_api_cli", None)
        zones = getattr(self, "_fwsz_zones", [])
        if not api_cli or not zones:
            return
        sel = (self.fwsz_zone.get() or "").strip()
        zone = None
        for z in zones:
            nm = (z.get("name") or "").strip()
            zid = (z.get("id") or "").strip()
            if sel in (nm, zid) or sel.startswith(nm + " "):
                zone = z
                break
        if not zone:
            zone = zones[0]
        try:
            zid = zone.get("id") or ""
            vers = api_cli.fetch_ap_firmware(zid)
            current = api_cli.zone_current_fw(zone)
            if not current and zid:
                current = api_cli.zone_current_fw(api_cli.fetch_zone(zid))
            labels = []
            for v in vers:
                labels.append(f"{v} *" if current and v == current else v)
            if current and current not in vers:
                labels.insert(0, f"{current} *")
            self.fwsz_ver_combo["values"] = labels
            pick = next((lb for lb in labels if lb.endswith(" *")), labels[0] if labels else "")
            if pick:
                self.fwsz_ver.set(pick)
            extra = f"  현재 * {current}" if current else ""
            self._log(f"펌웨어 {len(vers)}개{extra}: {', '.join(labels[:8])}\n")
        except Exception as e:
            self._log(f"펌웨어 목록 실패: {e}\n")

    def _fwsz_cli(self):
        cli = getattr(self, "_fwsz_api_cli", None)
        if cli and getattr(cli, "service_ticket", None):
            return cli
        host = self.fwsz_api_ip.get().strip()
        user = self.fwsz_user.get().strip()
        pw = self.fwsz_pass.get()
        api = self.fwsz_api.get().strip()
        if not (host and user and pw and api):
            messagebox.showwarning("안내", "SZ API IP / Username / Password / API 버전을 입력하세요.")
            return None
        cli = SmartZoneAPI(host, user, pw, api)
        ok, msg = cli.login()
        if not ok:
            messagebox.showerror("SZ 로그인 실패", msg)
            return None
        self._fwsz_api_cli = cli
        return cli

    def _fwsz_zone_id(self) -> str:
        sel = (self.fwsz_zone.get() or "").strip()
        for z in getattr(self, "_fwsz_zones", []) or []:
            if sel == (z.get("name") or "").strip() or sel == (z.get("id") or "").strip():
                return z.get("id") or ""
        return ""

    def _fwsz_rule_toggle_fields(self):
        typ = (self.fwsz_rule_type.get() if hasattr(self, "fwsz_rule_type") else "ProvisionTag") or "ProvisionTag"
        for fr in (
            getattr(self, "fwsz_rule_fr_tag", None),
            getattr(self, "fwsz_rule_fr_ip", None),
            getattr(self, "fwsz_rule_fr_net", None),
        ):
            if fr is not None:
                fr.pack_forget()
        if typ == "IPAddressRange":
            if hasattr(self, "fwsz_rule_fr_ip"):
                self.fwsz_rule_fr_ip.pack(side=LEFT)
        elif typ == "Subnet":
            if hasattr(self, "fwsz_rule_fr_net"):
                self.fwsz_rule_fr_net.pack(side=LEFT)
        else:
            if hasattr(self, "fwsz_rule_fr_tag"):
                self.fwsz_rule_fr_tag.pack(side=LEFT)
        chk = getattr(self, "fwsz_apply_tag_chk", None)
        if chk is not None:
            if typ == "ProvisionTag":
                chk.pack(anchor="w", pady=(4, 0))
            else:
                chk.pack_forget()
                if hasattr(self, "fwsz_apply_tag"):
                    self.fwsz_apply_tag.set(False)

    def _fwsz_rule_value(self, item: dict) -> str:
        t = item.get("type") or ""
        if t == "IPAddressRange" or item.get("ipAddressRange"):
            rng = item.get("ipAddressRange") or {}
            return f"{rng.get('fromIp', '')} ~ {rng.get('toIp', '')}"
        if t == "Subnet" or item.get("subnet"):
            sub = item.get("subnet") or {}
            return f"{sub.get('network') or sub.get('ipAddress') or sub.get('networkAddress') or ''} / {sub.get('subnetMask') or ''}"
        if t == "ProvisionTag" or item.get("provisionTag"):
            return str(item.get("provisionTag") or "")
        gps = item.get("gpsCoordinates") or {}
        if gps:
            return f"{gps.get('latitude')},{gps.get('longitude')}"
        return ""

    def _fwsz_rule_list(self):
        cli = self._fwsz_cli()
        if not cli:
            return
        try:
            rules = cli.fetch_ap_rules()
            detailed = []
            for r in rules:
                rid = r.get("id") or ""
                full = cli.fetch_ap_rule(rid) if rid else r
                if not isinstance(full, dict) or not full.get("id"):
                    full = r
                detailed.append(full)
            self._fwsz_rules = detailed
            tree = getattr(self, "fwsz_rule_tree", None)
            if tree:
                tree.delete(*tree.get_children())
                for it in detailed:
                    zone = ""
                    mz = it.get("mobilityZone") or {}
                    if isinstance(mz, dict):
                        zone = mz.get("name") or mz.get("id") or ""
                    tree.insert("", END, iid=it.get("id") or "", values=(
                        it.get("priority") or "",
                        it.get("description") or "",
                        it.get("type") or "",
                        zone,
                        self._fwsz_rule_value(it),
                    ))
            self._log(f"AP Registration Rule {len(detailed)}개 조회\n")
        except Exception as e:
            messagebox.showerror("오류", str(e))

    def _fwsz_rule_on_select(self):
        tree = getattr(self, "fwsz_rule_tree", None)
        if not tree:
            return
        sel = tree.selection()
        if not sel:
            return
        rid = sel[0]
        item = next((x for x in getattr(self, "_fwsz_rules", []) if x.get("id") == rid), None)
        if not item:
            return
        self.fwsz_rule_desc.set(item.get("description") or "")
        self.fwsz_rule_type.set(item.get("type") or "ProvisionTag")
        tag = item.get("provisionTag") or ""
        if tag:
            self.fwsz_rule_tag.set(tag)
        rng = item.get("ipAddressRange") or {}
        if rng:
            self.fwsz_rule_from.set(rng.get("fromIp") or "")
            self.fwsz_rule_to.set(rng.get("toIp") or "")
        sub = item.get("subnet") or {}
        if sub:
            self.fwsz_rule_net.set(sub.get("network") or sub.get("ipAddress") or "")
            self.fwsz_rule_mask.set(sub.get("subnetMask") or self.fwsz_rule_mask.get())
        self._fwsz_rule_toggle_fields()

    def _fwsz_rule_create(self):
        cli = self._fwsz_cli()
        if not cli:
            return
        zid = self._fwsz_zone_id()
        if not zid:
            messagebox.showwarning("안내", "먼저 Zone / 펌웨어 목록으로 Zone을 선택하세요.")
            return
        typ = (self.fwsz_rule_type.get() or "ProvisionTag").strip()
        desc = (self.fwsz_rule_desc.get() or "").strip() or f"rule-{typ}"
        body = {"description": desc, "type": typ, "mobilityZone": {"id": zid}}
        if typ == "ProvisionTag":
            tag = (self.fwsz_rule_tag.get() or "").strip()
            if not tag:
                messagebox.showwarning("안내", "Provision Tag 값을 입력하세요.")
                return
            body["provisionTag"] = tag
        elif typ == "IPAddressRange":
            a = (self.fwsz_rule_from.get() or "").strip()
            b = (self.fwsz_rule_to.get() or "").strip()
            if not a or not b:
                messagebox.showwarning("안내", "From IP / To IP를 입력하세요.")
                return
            body["ipAddressRange"] = {"fromIp": a, "toIp": b}
        elif typ == "Subnet":
            net = (self.fwsz_rule_net.get() or "").strip()
            mask = (self.fwsz_rule_mask.get() or "").strip()
            if not net or not mask:
                messagebox.showwarning("안내", "Subnet 주소와 마스크를 입력하세요.")
                return
            body["subnet"] = {"network": net, "subnetMask": mask}
        else:
            messagebox.showwarning("안내", "지원하지 않는 유형입니다.")
            return
        ok, resp = cli.create_ap_rule(body)
        if ok:
            self._log(f"Rule 생성 성공: {desc} / {typ}\n")
            self._fwsz_rule_list()
        else:
            messagebox.showerror("Rule 생성 실패", str(resp))

    def _fwsz_rule_delete(self):
        cli = self._fwsz_cli()
        if not cli:
            return
        tree = getattr(self, "fwsz_rule_tree", None)
        sel = tree.selection() if tree else ()
        if not sel:
            messagebox.showwarning("안내", "삭제할 Rule을 리스트에서 선택하세요.")
            return
        rid = sel[0]
        if not messagebox.askyesno("확인", "선택한 AP Registration Rule을 삭제할까요?"):
            return
        ok, resp = cli.delete_ap_rule(rid)
        if ok:
            self._log(f"Rule 삭제: {rid}\n")
            self._fwsz_rule_list()
        else:
            messagebox.showerror("삭제 실패", str(resp))

    def _fwsz_start(self, change_ip: bool = True):
        if self._worker and self._worker.is_alive():
            messagebox.showwarning("안내", "이미 실행 중입니다.")
            return
        if not getattr(self, "_csv_rows", None):
            messagebox.showwarning("안내", "먼저 AP CSV를 업로드하세요.")
            return
        host = self.fwsz_api_ip.get().strip()
        ver = self.fwsz_ver.get().replace("*", "").strip()
        if not host or not ver:
            messagebox.showwarning("안내", "SZ API IP와 펌웨어 버전을 입력하세요.")
            return
        if not self._require_second_account(
            getattr(self, "fwsz_user2", None), getattr(self, "fwsz_pass2", None)
        ):
            return
        mode = "자동IP → 고정IP" if change_ip else "IP변경안함"
        if not messagebox.askyesno(
            "확인",
            f"{len(self._csv_rows)} 대\nSZ {host}\n버전 {ver}\n업그레이드 + SZ 연동 ({mode}) 시작할까요?",
        ):
            return
        user2 = (self.fwsz_user2.get() or "").strip()
        pass2 = (self.fwsz_pass2.get() or "").strip()
        ptag = ""
        if getattr(self, "fwsz_apply_tag", None) and self.fwsz_apply_tag.get():
            ptag = (self.fwsz_rule_tag.get() or "").strip()
        rows = list(self._csv_rows)
        self._stop_flag = False
        self.fwsz_run_btn.config(state=DISABLED)
        if hasattr(self, "fwsz_run_keep_btn"):
            self.fwsz_run_keep_btn.config(state=DISABLED)
        self.fwsz_stop_btn.config(state=NORMAL)
        self.progress["value"] = 0
        self.progress["maximum"] = len(rows)
        self._log(
            f"\n===== SZ 펌웨어 업그레이드 {len(rows)} 대 / {ver} / {host} / {mode} =====\n"
        )

        def work():
            results = []
            old = sys.stdout
            sys.stdout = TextRedirector(self._log_queue)
            try:
                for i, row in enumerate(rows, 1):
                    if self._stop_flag:
                        self._log_queue.put("사용자 중지\n")
                        break
                    self._log_queue.put(f"\n[{i}/{len(rows)}] {row['ip']}\n")
                    r = process_ap(
                        ip=row["ip"],
                        user=row.get("user") or "",
                        password=row.get("pass") or "",
                        operation="fw_upgrade_sz",
                        new_ip=row.get("new_ip") or "",
                        subnet=row.get("subnet") or "",
                        gw=row.get("gw") or "",
                        sz=row.get("sz") or "",
                        hostname=row.get("hostname") or "",
                        new_password=pass2,
                        debug=True,
                        fw_host=host,
                        fw_file=ver,
                        fw_change_ip=change_ip,
                        provision_tag=ptag,
                        standard_password=pass2,
                        fallback_user=user2,
                        try_factory=False,
                    )
                    results.append(r)
                    self.after(0, lambda v=i: self.progress.config(value=v))
                    self.after(0, lambda v=i: self.fwsz_status.set(f"진행 {v}/{len(rows)}"))
            except Exception as e:
                self._log_queue.put(f"예외: {e}\n")
            finally:
                sys.stdout = old
            try:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                out = RESULTS_SZFW / f"result_fw_sz_{ts}.csv"
                RESULTS_SZFW.mkdir(exist_ok=True)
                with open(out, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f)
                    w.writerow(["ip", "status", "message", "model", "serial", "mac"])
                    for r in results:
                        w.writerow([r.get("ip"), r.get("status"), r.get("message"),
                                    r.get("model"), r.get("serial"), r.get("mac")])
                keep_latest_results(RESULTS_SZFW)
                self._save_session_log(LOG_SZFW, "fw_sz")
                ok_n = sum(1 for r in results if r.get("status") == "OK")
                self._log_queue.put(f"\n===== 완료: 성공 {ok_n}/{len(results)} =====\n결과: {out}\n")
                self.after(0, lambda: self.fwsz_status.set(f"완료 성공 {ok_n}/{len(results)}"))
                self.after(0, lambda: messagebox.showinfo("완료", f"성공 {ok_n}/{len(results)}\n\n{out}"))
            except Exception as e:
                self._log_queue.put(f"결과 저장 실패: {e}\n")
            self.after(0, lambda: self.fwsz_run_btn.config(state=NORMAL))
            self.after(0, lambda: getattr(self, "fwsz_run_keep_btn", self.fwsz_run_btn).config(state=NORMAL))
            self.after(0, lambda: self.fwsz_stop_btn.config(state=DISABLED))

        self._worker = threading.Thread(target=work, daemon=True)
        self._worker.start()

    # ------------------------------------------------------------------
    # 메뉴 7: AP 펌웨어 자동 업그레이드
    # ------------------------------------------------------------------
    def _build_fwup(self):
        self._clear_page()
        outer = self._scroll_page(padx=20, pady=16)
        self._back_btn(outer)
        Label(outer, text="7. AP 펌웨어 자동 업그레이드", font=("Segoe UI", 14, "bold"), fg=ACCENT, bg=BG).pack(anchor="w")
        Label(outer, text="CSV AP SSH 접속 → 선택한 .bl7 → TFTP fw update → reboot",
              font=("Segoe UI", 9), fg="#666", bg=BG).pack(anchor="w", pady=(2, 8))

        form = Frame(outer, bg=CARD, padx=12, pady=8, highlightbackground="#dee2e6", highlightthickness=1)
        form.pack(fill=X)
        form.grid_columnconfigure(0, weight=1)
        left = Frame(form, bg=CARD)
        right = LabelFrame(
            form, text="2차 기본 계정 (초기 비번 변경시 or CSV 실패시)",
            bg=CARD, fg="#444", font=("Segoe UI", 8),
            padx=10, pady=8, labelanchor="nw",
        )
        left.grid(row=0, column=0, sticky="nw")
        right.grid(row=0, column=1, sticky="ne", padx=(16, 0))
        self.fw_server_ip = StringVar(value="")
        self.fw_server_port = StringVar(value="69")
        self.fwup_proto = StringVar(value="tftp")
        self.fwup_file = StringVar(value="펌웨어 파일 선택")
        self.fwup_factory = BooleanVar(value=True)
        self.fwup_user2 = StringVar(value="")
        self.fwup_pass2 = StringVar(value="")

        def row(parent, lbl, var, show="", width=20):
            fr = Frame(parent, bg=CARD)
            fr.pack(fill=X, pady=2)
            Label(fr, text=lbl, width=16, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
            Entry(fr, textvariable=var, width=width, font=("Segoe UI", 10), show=show).pack(side=LEFT)
            return fr

        ip_row = row(left, "TFTP 서버 IP", self.fw_server_ip, width=20)
        Button(ip_row, text="IP 감지", font=("Segoe UI", 9),
               bg=BTN_BG, relief="solid", borderwidth=1, padx=8, pady=1,
               command=self._fw_refill_ip, cursor="hand2").pack(side=LEFT, padx=(8, 0))
        row(left, "TFTP 포트", self.fw_server_port, width=20)
        fr = Frame(left, bg=CARD)
        fr.pack(fill=X, pady=2)
        Label(fr, text="프로토콜", width=16, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        Label(fr, text="tftp", width=18, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        fr = Frame(left, bg=CARD)
        fr.pack(fill=X, pady=2)
        Label(fr, text="펌웨어 파일", width=16, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        self.fwup_combo = ttk.Combobox(fr, textvariable=self.fwup_file, state="readonly", width=36)
        self.fwup_combo.pack(side=LEFT)
        Button(
            fr, text="새로고침", font=("Segoe UI", 9, "bold"),
            bg="#fff", fg=ACCENT, relief="solid", borderwidth=1,
            highlightbackground=ACCENT, padx=10, pady=1,
            command=self._fwup_refresh_clicked, cursor="hand2",
        ).pack(side=LEFT, padx=(8, 0))
        self._fwup_refresh_combo()

        Label(right, text="Username", width=10, anchor="w", bg=CARD, font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w", pady=2)
        Entry(right, textvariable=self.fwup_user2, width=16, font=("Segoe UI", 10)).grid(row=0, column=1, sticky="w", pady=2)
        Label(right, text="Password", width=10, anchor="w", bg=CARD, font=("Segoe UI", 10)).grid(row=1, column=0, sticky="w", pady=2)
        self.fwup_pass2_entry = Entry(right, textvariable=self.fwup_pass2, width=16, font=("Segoe UI", 10), show="*")
        self.fwup_pass2_entry.grid(row=1, column=1, sticky="w", pady=2)
        self.fwup_pass2_tip = BalloonTip(self.fwup_pass2_entry, PW_RULE_KO, enabled=self._alt_pw_complex_on())
        self.fwup_pass2_show = BooleanVar(value=False)
        Checkbutton(right, text="표시", variable=self.fwup_pass2_show, bg=CARD,
                    command=lambda: self.fwup_pass2_entry.config(show="" if self.fwup_pass2_show.get() else "*")).grid(row=1, column=2, sticky="w", padx=(6, 0))
        self._alt_pw_hint_row(right)

        Checkbutton(form, text="완료 시 set factory + reboot (동일 버전은 생략)", variable=self.fwup_factory, bg=CARD).grid(row=1, column=0, columnspan=2, sticky="w", pady=(6, 0))
        brow = Frame(form, bg=CARD)
        brow.grid(row=2, column=0, columnspan=2, sticky="w", pady=(6, 0))
        Button(brow, text="펌웨어 업로드 (최대 10개씩)", font=("Segoe UI", 10, "bold"), bg=LINK, fg="white",
               relief="flat", padx=10, pady=4, command=self._fw_upload, cursor="hand2").pack(side=LEFT, padx=(0, 6))
        Button(brow, text="내장 TFTP 시작", font=("Segoe UI", 10, "bold"), bg="#17a2b8", fg="white", relief="flat",
               padx=10, pady=4, command=lambda: self._fw_start_builtin_tftp("upgrade"), cursor="hand2").pack(side=LEFT, padx=(0, 6))
        Button(brow, text="TFTP 중지", font=("Segoe UI", 10),
               bg=BTN_BG, relief="solid", borderwidth=1, padx=10, pady=5,
               command=self._fw_stop_builtin_tftp, cursor="hand2").pack(side=LEFT, padx=(0, 6))
        Button(brow, text="펌웨어 폴더", font=("Segoe UI", 10),
               bg=BTN_BG, relief="solid", borderwidth=1, padx=10, pady=5,
               command=lambda: self._open_path(FW_DIR), cursor="hand2").pack(side=LEFT, padx=(0, 6))
        Button(brow, text="결과 폴더", font=("Segoe UI", 10),
               bg=BTN_BG, relief="solid", borderwidth=1, padx=10, pady=5,
               command=lambda: self._open_path(RESULTS_FW), cursor="hand2").pack(side=LEFT)
        self._latest_result_btn(brow, RESULTS_FW)
        self._log_action_btns(brow, LOG_FW, bg=CARD)
        Label(brow, text=RESULT_HINT, font=("Segoe UI", 8), fg="#888", bg=CARD).pack(side=LEFT, padx=(8, 0))

        # 메뉴 6과 동일한 CSV 업로드 UI
        box = Frame(outer, bg=CARD, padx=12, pady=8, highlightbackground="#dee2e6", highlightthickness=1)
        box.pack(fill=X, pady=(8, 4))
        Label(box, text="AP 리스트", font=("Segoe UI", 10, "bold"), bg=CARD).pack(anchor="w")
        row1 = Frame(box, bg=CARD)
        row1.pack(fill=X, pady=4)
        Button(row1, text="샘플 다운로드 (CSV)", font=("Segoe UI", 10), bg="#28a745", fg="white", relief="flat",
               padx=10, pady=3, command=self._download_sample).pack(side=LEFT, padx=(0, 6))
        Button(row1, text="파일 업로드 (CSV)", font=("Segoe UI", 10), bg=LINK, fg="white", relief="flat",
               padx=10, pady=3, command=self._browse_csv).pack(side=LEFT)
        if not hasattr(self, "csv_path_var"):
            self.csv_path_var = StringVar()
        if not hasattr(self, "csv_info_var"):
            self.csv_info_var = StringVar(value="업로드된 파일 없음")
        Label(box, textvariable=self.csv_path_var, font=("Segoe UI", 8), bg=CARD).pack(anchor="w")
        Label(box, textvariable=self.csv_info_var, font=("Segoe UI", 8), bg=CARD, fg="#666").pack(anchor="w")
        prev = Frame(box, bg=CARD)
        prev.pack(fill=X)
        cols = ("ip", "user", "pass", "new_ip", "subnet", "gw", "sz", "hostname")
        self.csv_tree = ttk.Treeview(prev, columns=cols, show="headings", height=4)
        for c in cols:
            self.csv_tree.heading(c, text=c)
            self.csv_tree.column(c, width=90, anchor="w")
        ysb = Scrollbar(prev, orient=VERTICAL, command=self.csv_tree.yview)
        xsb = Scrollbar(prev, orient=HORIZONTAL, command=self.csv_tree.xview)
        self.csv_tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.csv_tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        prev.grid_rowconfigure(0, weight=1)
        prev.grid_columnconfigure(0, weight=1)
        if getattr(self, "_csv_rows", None):
            self._fill_csv_preview(self._csv_rows)

        run = Frame(outer, bg=BG)
        run.pack(fill=X, pady=6)
        self.fwup_run_btn = Button(run, text="▶ 업그레이드 실행", font=("Segoe UI", 11, "bold"), bg=ACCENT, fg="white",
                                   relief="flat", padx=16, pady=5, command=self._fwup_start, cursor="hand2")
        self.fwup_run_btn.pack(side=LEFT, padx=(0, 8))
        self.fwup_stop_btn = Button(run, text="중지", bg="#6c757d", fg="white", relief="flat", padx=12, pady=5,
                                    command=self._stop_batch, state=DISABLED)
        self.fwup_stop_btn.pack(side=LEFT)
        self.fwup_status = StringVar(value="대기 — TFTP 시작 후 CSV 업로드하고 실행")
        Label(outer, textvariable=self.fwup_status, font=("Segoe UI", 9), bg=BG, fg="#555").pack(anchor="w")
        self.progress = ttk.Progressbar(outer, mode="determinate")
        self.progress.pack(fill=X, pady=4)
        self._make_log(outer, "실행 로그", 8)

    def _fwup_start(self):
        if self._worker and self._worker.is_alive():
            messagebox.showwarning("안내", "이미 실행 중입니다.")
            return
        if not getattr(self, "_csv_rows", None):
            messagebox.showwarning("안내", "먼저 AP CSV를 업로드하세요.")
            return
        host = self.fw_server_ip.get().strip()
        if not host:
            messagebox.showwarning("안내", "TFTP 서버 IP를 입력하세요.")
            return
        if not list(FW_DIR.glob("*.bl7")):
            messagebox.showwarning("안내", "firmware 폴더에 .bl7 이 없습니다. 메뉴 8에서 업로드하세요.")
            return
        if not self._tftp_is_running():
            messagebox.showwarning(
                "안내",
                "내장 TFTP 데몬을 시작하지 않았습니다. 내장 TFTP 를 시작하세요.",
            )
            return
        if not self._require_second_account(
            getattr(self, "fwup_user2", None), getattr(self, "fwup_pass2", None)
        ):
            return
        if not messagebox.askyesno("확인", f"{len(self._csv_rows)} 대 AP 펌웨어 업그레이드를 시작할까요?"):
            return
        sel = (self.fwup_file.get() or "").strip()
        fw_file = sel.split()[0] if sel else ""
        if not fw_file.lower().endswith(".bl7"):
            messagebox.showwarning("안내", "펌웨어 파일(.bl7)을 선택하세요.")
            return
        proto = self.fwup_proto.get() or "tftp"
        port = self.fw_server_port.get().strip() or "69"
        factory = bool(self.fwup_factory.get())
        user2 = (self.fwup_user2.get() or "").strip()
        pass2 = (self.fwup_pass2.get() or "").strip()
        rows = list(self._csv_rows)
        self._stop_flag = False
        self.fwup_run_btn.config(state=DISABLED)
        self.fwup_stop_btn.config(state=NORMAL)
        self.progress["value"] = 0
        self.progress["maximum"] = len(rows)
        self._log(f"\n===== 펌웨어 업그레이드 {len(rows)} 대 / {fw_file} / {proto} {host}:{port} =====\n")

        def work():
            results = []
            old = sys.stdout
            sys.stdout = TextRedirector(self._log_queue)
            try:
                for i, row in enumerate(rows, 1):
                    if self._stop_flag:
                        self._log_queue.put("사용자 중지\n")
                        break
                    self._log_queue.put(f"\n[{i}/{len(rows)}] {row['ip']}\n")
                    r = process_ap(
                        ip=row["ip"],
                        user=row.get("user") or "",
                        password=row.get("pass") or "",
                        operation="fw_upgrade",
                        new_password=pass2,
                        debug=True,
                        fw_host=host,
                        fw_port=port,
                        fw_proto=proto,
                        fw_file=fw_file,
                        fw_factory=factory,
                        standard_password=pass2,
                        fallback_user=user2,
                        try_factory=False,
                    )
                    results.append(r)
                    self.after(0, lambda v=i: self.progress.config(value=v))
                    self.after(0, lambda v=i: self.fwup_status.set(f"진행 {v}/{len(rows)}"))
            except Exception as e:
                self._log_queue.put(f"예외: {e}\n")
            finally:
                sys.stdout = old
            try:
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                out = RESULTS_FW / f"result_fw_upgrade_{ts}.csv"
                RESULTS_FW.mkdir(exist_ok=True)
                with open(out, "w", encoding="utf-8-sig", newline="") as f:
                    w = csv.writer(f)
                    w.writerow(["ip", "status", "message", "model", "serial", "mac"])
                    for r in results:
                        w.writerow([r.get("ip"), r.get("status"), r.get("message"),
                                    r.get("model"), r.get("serial"), r.get("mac")])
                keep_latest_results(RESULTS_FW)
                self._save_session_log(LOG_FW, "fw_upgrade")
                ok_n = sum(1 for r in results if r.get("status") == "OK")
                self._log_queue.put(f"\n===== 완료: 성공 {ok_n}/{len(results)} =====\n결과: {out}\n")
                self.after(0, lambda: self.fwup_status.set(f"완료 성공 {ok_n}/{len(results)}"))
                self.after(0, lambda: messagebox.showinfo("완료", f"성공 {ok_n}/{len(results)}\n\n{out}"))
            except Exception as e:
                self._log_queue.put(f"결과 저장 실패: {e}\n")
            self.after(0, lambda: self.fwup_run_btn.config(state=NORMAL))
            self.after(0, lambda: self.fwup_stop_btn.config(state=DISABLED))

        self._worker = threading.Thread(target=work, daemon=True)
        self._worker.start()

    def _fwup_refresh_clicked(self):
        self._fwup_refresh_combo()
        n = len(list(FW_DIR.glob("*.bl7"))) if FW_DIR.is_dir() else 0
        self._log(f"목록 새로고침: .bl7 {n}개 ({FW_DIR})\n")

    def _fwup_refresh_combo(self):
        items = []
        if FW_DIR.is_dir():
            for f in sorted(FW_DIR.glob("*.bl7")):
                m, v = parse_bl7_name(f)
                items.append(f"{f.name}  [{m} {v}]")
        if hasattr(self, "fwup_combo"):
            self.fwup_combo["values"] = items
        cur = self.fwup_file.get() if hasattr(self, "fwup_file") else ""
        if items:
            if cur not in items:
                self.fwup_file.set(items[0])
        else:
            items = ["펌웨어 파일 선택"]
            if hasattr(self, "fwup_combo"):
                self.fwup_combo["values"] = items
            self.fwup_file.set("펌웨어 파일 선택")

    # ------------------------------------------------------------------
    # 메뉴 8: 펌웨어 CLI 명령어 생성 (bl7 업로드)
    # ------------------------------------------------------------------
    def _build_fw(self):
        self._clear_page()

        outer = self._scroll_page(padx=20, pady=16)
        self._back_btn(outer)

        Label(
            outer, text="8. AP 펌웨어 CLI 명령어 생성",
            font=("Segoe UI", 14, "bold"), fg=ACCENT, bg=BG,
        ).pack(anchor="w")
        Label(
            outer,
            text=".bl7 업로드 → .rcks / HTML 명령어 자동 생성 (원본 fw.sh 동일). 최대 10개씩 업로드, 추가 업로드 후 다시 생성 가능.",
            font=("Segoe UI", 9), fg="#666", bg=BG, wraplength=860, justify="left",
        ).pack(anchor="w", pady=(2, 10))

        form = Frame(outer, bg=CARD, padx=14, pady=12,
                     highlightbackground="#dee2e6", highlightthickness=1)
        form.pack(fill=X)

        self.fw_server_ip = StringVar(value="")
        self.fw_server_port = StringVar(value="69")
        self.fw_proto = StringVar(value="tftp")
        self.fw_location = StringVar(value="")  # TFTP 는 불필요

        def row(lbl, var, width=20):
            fr = Frame(form, bg=CARD)
            fr.pack(fill=X, pady=3)
            Label(fr, text=lbl, width=24, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
            Entry(fr, textvariable=var, font=("Segoe UI", 10), width=width).pack(side=LEFT)
            return fr

        ip_row = row("TFTP 서버 IP (해당 PC)", self.fw_server_ip)
        Button(
            ip_row, text="IP 감지", font=("Segoe UI", 9),
            bg=BTN_BG, relief="solid", borderwidth=1, padx=8, pady=2,
            command=self._fw_refill_ip, cursor="hand2",
        ).pack(side=LEFT, padx=(8, 0))
        row("TFTP 포트 [69]", self.fw_server_port)

        fr_p = Frame(form, bg=CARD)
        fr_p.pack(fill=X, pady=3)
        Label(fr_p, text="프로토콜", width=24, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
        Label(fr_p, text="tftp", width=18, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)

        Label(
            form,
            text="※ 권장: 「내장 TFTP 시작」— 설정 없이 firmware 폴더를 자동 공유합니다. (포트 69, 관리자 권한 필요할 수 있음)",
            font=("Segoe UI", 8), fg="#888", bg=CARD, wraplength=800, justify="left",
        ).pack(anchor="w", pady=(4, 0))

        btn_row = Frame(form, bg=CARD)
        btn_row.pack(fill=X, pady=(12, 0))
        Button(
            btn_row, text="펌웨어 업로드 (최대 10개씩)", font=("Segoe UI", 10, "bold"),
            bg=LINK, fg="white", relief="flat", padx=14, pady=5,
            command=self._fw_upload, cursor="hand2",
        ).pack(side=LEFT, padx=(0, 8))
        Button(
            btn_row, text="▶ 목록 생성", font=("Segoe UI", 10, "bold"),
            bg=ACCENT, fg="white", relief="flat", padx=14, pady=5,
            command=self._fw_generate, cursor="hand2",
        ).pack(side=LEFT, padx=(0, 8))
        Button(
            btn_row, text="▶ 내장 TFTP 시작", font=("Segoe UI", 10, "bold"),
            bg="#17a2b8", fg="white", relief="flat", padx=14, pady=5,
            command=lambda: self._fw_start_builtin_tftp("list"), cursor="hand2",
        ).pack(side=LEFT, padx=(0, 8))
        Button(
            btn_row, text="TFTP 중지", font=("Segoe UI", 10),
            bg=BTN_BG, relief="solid", borderwidth=1, padx=10, pady=5,
            command=self._fw_stop_builtin_tftp, cursor="hand2",
        ).pack(side=LEFT, padx=(0, 8))
        Button(
            btn_row, text="펌웨어 폴더", font=("Segoe UI", 10),
            bg=BTN_BG, relief="solid", borderwidth=1, padx=12, pady=5,
            command=lambda: self._open_path(FW_DIR), cursor="hand2",
        ).pack(side=LEFT, padx=(0, 8))
        Button(
            btn_row, text="펌웨어 스크립트 보기", font=("Segoe UI", 10, "bold"),
            bg="#fd7e14", fg="white", relief="flat", padx=12, pady=5,
            command=self._fw_open_index, cursor="hand2",
        ).pack(side=LEFT)
        self._log_action_btns(btn_row, LOG_FWCLI, bg=CARD)

        self.fw_info = StringVar(value=self._fw_scan_info())
        Label(outer, textvariable=self.fw_info, font=("Segoe UI", 9), bg=BG, fg="#555").pack(
            anchor="w", pady=(8, 4)
        )

        hdr = Frame(outer, bg=BG)
        hdr.pack(fill=X, pady=(4, 0))
        Label(hdr, text="등록된 .bl7 파일", font=("Segoe UI", 10, "bold"), bg=BG).pack(side=LEFT)
        Button(
            hdr, text="새로고침", font=("Segoe UI", 9, "bold"),
            bg="#fff", fg=ACCENT, relief="solid", borderwidth=1,
            highlightbackground=ACCENT, padx=10, pady=1,
            command=self._fw_refresh_clicked, cursor="hand2",
        ).pack(side=LEFT, padx=(8, 0))
        list_fr = Frame(outer, bg=CARD)
        list_fr.pack(fill=BOTH, expand=True, pady=(2, 4))
        cols = ("name", "size", "model", "version")
        self.fw_tree = ttk.Treeview(list_fr, columns=cols, show="headings", height=10)
        for c, w, h in (
            ("name", 280, "파일명"),
            ("size", 90, "크기"),
            ("model", 80, "모델"),
            ("version", 140, "버전"),
        ):
            self.fw_tree.heading(c, text=h)
            self.fw_tree.column(c, width=w, anchor="w")
        ysb = Scrollbar(list_fr, orient=VERTICAL, command=self.fw_tree.yview)
        xsb = Scrollbar(list_fr, orient=HORIZONTAL, command=self.fw_tree.xview)
        self.fw_tree.configure(yscrollcommand=ysb.set, xscrollcommand=xsb.set)
        self.fw_tree.grid(row=0, column=0, sticky="nsew")
        ysb.grid(row=0, column=1, sticky="ns")
        xsb.grid(row=1, column=0, sticky="ew")
        list_fr.grid_rowconfigure(0, weight=1)
        list_fr.grid_columnconfigure(0, weight=1)
        self._fw_refresh_list()
        self._make_log(outer, "로그", 8)

    def _fw_scan_info(self) -> str:
        FW_DIR.mkdir(exist_ok=True)
        n = len(list(FW_DIR.glob("*.bl7")))
        return f"펌웨어 폴더: {FW_DIR}  |  .bl7 파일 {n}개"

    def _fw_refresh_clicked(self):
        self._fw_refresh_list()
        n = len(list(FW_DIR.glob("*.bl7")))
        self._log(f"목록 새로고침: .bl7 {n}개 ({FW_DIR})\n")

    def _fw_refresh_list(self):
        if not hasattr(self, "fw_tree"):
            return
        self.fw_tree.delete(*self.fw_tree.get_children())
        from modules.fw_builder import parse_bl7_name
        for f in sorted(FW_DIR.glob("*.bl7")):
            model, ver = parse_bl7_name(f)
            size = f.stat().st_size
            size_s = f"{size/1024/1024:.2f} MB" if size > 1024 * 1024 else f"{size/1024:.1f} KB"
            self.fw_tree.insert("", END, values=(f.name, size_s, model, ver))
        self.fw_info.set(self._fw_scan_info())

    def _fw_upload(self):
        paths = filedialog.askopenfilenames(
            title="펌웨어 .bl7 선택 (최대 10개)",
            filetypes=[("BL7 firmware", "*.bl7"), ("All files", "*.*")],
        )
        if not paths:
            return
        paths = list(paths)[:10]
        FW_DIR.mkdir(exist_ok=True)
        ok_n = 0
        for pth in paths:
            try:
                src = Path(pth)
                if src.suffix.lower() != ".bl7":
                    self._log(f"스킵 (bl7 아님): {src.name}\n")
                    continue
                dest = FW_DIR / src.name
                shutil.copy2(src, dest)
                ok_n += 1
                self._log(f"업로드: {src.name} → {dest}\n")
            except Exception as e:
                self._log(f"실패: {pth} → {e}\n")
        self._fw_refresh_list()
        if hasattr(self, "fwup_combo"):
            self._fwup_refresh_combo()
        messagebox.showinfo(
            "업로드",
            f"{ok_n}개 파일을 firmware 폴더에 저장했습니다.\n"
            f"메뉴 7 펌웨어 파일 목록에서 선택하세요.",
        )

    def _fw_generate(self):
        ip = self.fw_server_ip.get().strip()
        proto = (self.fw_proto.get() or "tftp").strip().lower()
        default_port = "69" if proto == "tftp" else "80"
        port_s = self.fw_server_port.get().strip() or default_port
        loc = self.fw_location.get().strip() if hasattr(self, "fw_location") else ""
        if not ip:
            messagebox.showwarning("안내", "TFTP/HTTP 서버 IP를 입력하세요. (보통 이 PC의 IP)")
            return
        try:
            port = int(port_s)
        except ValueError:
            messagebox.showwarning("안내", "포트는 숫자여야 합니다.")
            return
        files = list(FW_DIR.glob("*.bl7"))
        if not files:
            messagebox.showwarning("안내", "firmware 폴더에 .bl7 파일이 없습니다. 먼저 업로드하세요.")
            return
        self._log(f"\n===== 펌웨어 목록 생성: {len(files)} files, {proto}://{ip}:{port} =====\n")
        try:
            result = build_firmware_package(
                bl7_files=files,
                server_ip=ip,
                server_port=port,
                output_path=FW_DIR,
                location_path=loc,
                protocol=proto,
            )
            if not result.get("ok"):
                messagebox.showerror("오류", result.get("message", "실패"))
                self._log(result.get("message", "") + "\n")
                return
            self._log(result["message"] + "\n")
            self._log(f"index: {result.get('index_html')}\n")
            for r in result.get("rcks_files") or []:
                self._log(f"  rcks: {r}\n")
            self._fw_refresh_list()
            messagebox.showinfo(
                "완료",
                f"{result['message']}\n\n"
                f"서버: {result.get('server')}\n"
                f"경로: {result.get('location_path')}\n\n"
                f"「펌웨어 스크립트 보기」를 눌러 목록을 열고 CLI를 복사하세요.",
            )
            self._save_session_log(LOG_FWCLI, "fw_cli")
            if not self._tftp_is_running():
                messagebox.showwarning(
                    "안내",
                    "AP에서 펌웨어 다운로드를 하려면 내장 TFTP 시작을 눌러 내장 TFTP 데몬을 실행하세요.",
                )
        except Exception as e:
            self._log(f"예외: {e}\n")
            messagebox.showerror("오류", str(e))

    @staticmethod
    def _detect_local_ip() -> str:
        """이 PC의 LAN IP 자동 감지 (AP가 접근 가능한 주소)"""
        import socket
        # 1) UDP 소켓으로 기본 경로 인터페이스 IP
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.settimeout(0.5)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            if ip and not ip.startswith("127."):
                return ip
        except Exception:
            pass
        # 2) hostname 조회
        try:
            for info in socket.getaddrinfo(socket.gethostname(), None, socket.AF_INET):
                ip = info[4][0]
                if ip and not ip.startswith("127."):
                    return ip
        except Exception:
            pass
        return ""

    def _fw_refill_ip(self):
        ip = self._detect_local_ip()
        if ip:
            self.fw_server_ip.set(ip)
            messagebox.showinfo("IP 감지", f"이 PC IP: {ip}")
        else:
            messagebox.showwarning("IP 감지", "자동 감지 실패. 수동으로 입력하세요.")

    def _tftp_is_running(self) -> bool:
        srv = getattr(self, "_tftp_server", None)
        return bool(srv and getattr(srv, "running", False))

    def _fw_start_builtin_tftp(self, kind="upgrade"):
        """설정 불필요 — firmware 폴더를 바로 공유하는 내장 TFTP"""
        port_s = self.fw_server_port.get().strip() or "69"
        try:
            port = int(port_s)
        except ValueError:
            messagebox.showwarning("안내", "포트는 숫자여야 합니다.")
            return
        # 기존 서버 중지
        self._fw_stop_builtin_tftp(silent=True)
        FW_DIR.mkdir(exist_ok=True)
        try:
            srv = SimpleTftpServer(
                root=FW_DIR,
                host="0.0.0.0",
                port=port,
                log=lambda m: self._log_queue.put(str(m) + "\n"),
            )
            srv.start()
            self._tftp_server = srv
            ip = self.fw_server_ip.get().strip() or "(미입력)"
            self._log(f"내장 TFTP 서버 시작: 0.0.0.0:{port} → {FW_DIR}\n")
            if kind == "list":
                how = (
                    "「목록 생성」 후 AP에서 fw update 하면 됩니다.\n"
                    "포트 69 바인드 실패 시 관리자 권한으로 실행하거나\n"
                    "포트를 6969 등으로 바꾸세요."
                )
            else:
                how = (
                    "TFTP 서버 IP를 확인(또는 IP 감지)한 뒤\n"
                    "CSV를 올리고 「업그레이드 실행」하면\n"
                    "AP가 이 TFTP에서 .bl7을 받아 업그레이드합니다.\n"
                    "포트 69 바인드 실패 시 관리자 권한으로 실행하거나\n"
                    "포트를 6969 등으로 바꾸세요."
                )
            messagebox.showinfo(
                "내장 TFTP 서버",
                f"TFTP 서버가 시작되었습니다. (설정 불필요)\n\n"
                f"• 공유 폴더: {FW_DIR}\n"
                f"• 포트: {port}\n"
                f"• AP에 넣을 서버 IP: {ip}\n\n"
                f"{how}",
            )
        except OSError as e:
            messagebox.showerror("TFTP 시작 실패", str(e))
            self._log(f"TFTP 시작 실패: {e}\n")

    def _fw_stop_builtin_tftp(self, silent=False):
        srv = getattr(self, "_tftp_server", None)
        if srv:
            try:
                srv.stop()
            except Exception as e:
                self._log(f"TFTP 중지 오류: {e}\n")
            self._tftp_server = None
            if not silent:
                messagebox.showinfo("TFTP", "내장 TFTP 서버를 중지했습니다.")
        elif not silent:
            messagebox.showinfo("TFTP", "실행 중인 내장 TFTP 서버가 없습니다.")

    def _fw_open_index(self):
        idx = FW_DIR / "index.html"
        if not idx.is_file():
            messagebox.showwarning("안내", "index.html 이 없습니다. 먼저 「목록 생성」을 실행하세요.")
            return
        try:
            os.startfile(str(idx.resolve()))
        except Exception as e:
            messagebox.showinfo("경로", str(idx.resolve()) + f"\n{e}")

    # ------------------------------------------------------------------
    # 메뉴 12: 단일 AP 테스트
    # ------------------------------------------------------------------
    def _build_single_test(self):
        self._clear_page()

        outer = self._scroll_page(padx=20, pady=16)
        self._back_btn(outer)

        Label(
            outer, text="12. 단일 AP SSH 테스트",
            font=("Segoe UI", 14, "bold"), fg=ACCENT, bg=BG,
        ).pack(anchor="w")
        Label(
            outer,
            text=f"기본 정책: {DEFAULT_USER}/{DEFAULT_PASSWORD} → {STANDARD_PASSWORD}",
            font=("Segoe UI", 9), fg="#666", bg=BG,
        ).pack(anchor="w", pady=(2, 10))

        form = Frame(outer, bg=CARD, padx=14, pady=12,
                     highlightbackground="#dee2e6", highlightthickness=1)
        form.pack(fill=X)

        self.s_ip = StringVar()
        self.s_user = StringVar(value=DEFAULT_USER)
        self.s_pass = StringVar(value="")  # 비우면 자동 정책

        def row(label, var, show=None):
            fr = Frame(form, bg=CARD)
            fr.pack(fill=X, pady=3)
            Label(fr, text=label, width=18, anchor="w", bg=CARD, font=("Segoe UI", 10)).pack(side=LEFT)
            Entry(fr, textvariable=var, font=("Segoe UI", 10), width=36, show=show or "").pack(side=LEFT)

        row("AP IP", self.s_ip)
        row("Username", self.s_user)
        row(f"Password (빈칸=자동)", self.s_pass, show="*")

        btn_row = Frame(form, bg=CARD)
        btn_row.pack(fill=X, pady=(10, 0))
        self.s_connect_btn = Button(
            btn_row, text="연결 테스트", font=("Segoe UI", 10, "bold"),
            bg=LINK, fg="white", relief="flat", padx=16, pady=5,
            command=self._single_connect, cursor="hand2",
        )
        self.s_connect_btn.pack(side=LEFT, padx=(0, 8))
        self._log_action_btns(btn_row, LOG_SINGLE, bg=CARD)
        self._make_log(outer, "로그", 10)
        cmd_fr = Frame(outer, bg=BG)
        cmd_fr.pack(fill=X, pady=(8, 0))
        Label(cmd_fr, text="명령:", bg=BG, font=("Segoe UI", 10)).pack(side=LEFT)
        self.cmd_var = StringVar()
        Entry(cmd_fr, textvariable=self.cmd_var, font=("Consolas", 10), width=50).pack(side=LEFT, padx=6)
        Button(
            cmd_fr, text="전송", font=("Segoe UI", 10),
            bg=BTN_BG, relief="solid", borderwidth=1, padx=12,
            command=self._single_cmd, cursor="hand2",
        ).pack(side=LEFT)
        self._ssh_session = None

    def _single_connect(self):
        ip = self.s_ip.get().strip()
        if not ip:
            messagebox.showwarning("안내", "AP IP를 입력하세요.")
            return
        if self._worker and self._worker.is_alive():
            return
        self.s_connect_btn.config(state=DISABLED)
        self._log(f"\n===== 연결: {ip} =====\n")

        def work():
            old = sys.stdout
            sys.stdout = TextRedirector(self._log_queue)
            ssh = None
            try:
                ssh = RuckusSSH(timeout=10, debug=True, verbose=True)
                ok, msg = ssh.connect(
                    ip,
                    username=self.s_user.get().strip() or DEFAULT_USER,
                    password=self.s_pass.get().strip() or None,
                    new_password=STANDARD_PASSWORD,
                    standard_password=STANDARD_PASSWORD,
                )
                if ok:
                    self._log_queue.put("\n로그인 성공 — boarddata 조회\n")
                    serial, mac = ssh.get_boarddata()
                    self._log_queue.put(f"Serial: {serial}\nMAC: {mac}\n")
                    self._ssh_session = ssh
                else:
                    self._log_queue.put(f"\n실패: {msg}\n")
                    if ssh:
                        ssh.close()
            except Exception as e:
                self._log_queue.put(f"\n예외: {e}\n")
                if ssh:
                    try:
                        ssh.close()
                    except Exception:
                        pass
            finally:
                sys.stdout = old
                self.after(0, lambda: self.s_connect_btn.config(state=NORMAL))
                self._save_session_log(LOG_SINGLE, "single")

        self._worker = threading.Thread(target=work, daemon=True)
        self._worker.start()

    def _single_cmd(self):
        cmd = self.cmd_var.get().strip()
        if not cmd:
            return
        ssh = self._ssh_session
        if not ssh or not ssh.shell:
            messagebox.showwarning("안내", "먼저 연결 테스트를 성공하세요.")
            return

        def work():
            old = sys.stdout
            sys.stdout = TextRedirector(self._log_queue)
            try:
                ok, out = ssh.run(cmd, success_pattern=r"rkscli", timeout=12)
                self._log_queue.put(out + "\n")
            except Exception as e:
                self._log_queue.put(f"오류: {e}\n")
            finally:
                sys.stdout = old
                self._save_session_log(LOG_SINGLE, "single_cmd")

        threading.Thread(target=work, daemon=True).start()


    # ------------------------------------------------------------------
    # 메뉴 11: OUI
    # ------------------------------------------------------------------
    def _build_oui(self):
        self._clear_page()
        outer = self._scroll_page(padx=20, pady=16)
        self._back_btn(outer)
        Label(outer, text="11. OUI 조회", font=("Segoe UI", 14, "bold"), fg=ACCENT, bg=BG).pack(anchor="w")
        Label(
            outer,
            text="IEEE OUI 리스트를 다운로드하여 results/oui 에 저장합니다.",
            font=("Segoe UI", 9), fg="#666", bg=BG,
        ).pack(anchor="w", pady=(2, 8))
        btn_row = Frame(outer, bg=BG)
        btn_row.pack(anchor="w", pady=8)
        Button(
            btn_row, text="IEEE OUI 다운로드", font=("Segoe UI", 11),
            bg=LINK, fg="white", relief="flat", padx=16, pady=8,
            command=self._download_oui, cursor="hand2",
        ).pack(side=LEFT, padx=(0, 8))
        Button(
            btn_row, text="OUI 파일 열기", font=("Segoe UI", 10),
            bg=BTN_BG, relief="solid", borderwidth=1, padx=12, pady=8,
            command=self._open_oui_file, cursor="hand2",
        ).pack(side=LEFT)
        self._log_action_btns(btn_row, LOG_OUI, bg=BG)
        self._make_log(outer, "로그", 20)

    def _latest_result_btn(self, parent, folder: Path, suffixes=None):
        Button(
            parent, text="최근 결과 다운로드", font=("Segoe UI", 10),
            bg="#28a745", fg="white", relief="flat", padx=12, pady=5,
            command=lambda: self._download_latest_results(folder, suffixes),
            cursor="hand2",
        ).pack(side=LEFT, padx=(6, 0))

    def _log_action_btns(self, parent, folder: Path, bg=None):
        """결과 버튼 오른쪽에 붙는 작은 로그 폴더/최근 로그 버튼."""
        bg = bg if bg is not None else CARD
        Button(
            parent, text="로그 폴더", font=("Segoe UI", 10),
            bg="#e2f22e", fg="#333", relief="solid", borderwidth=1, padx=10, pady=5,
            command=lambda f=folder: self._open_path(f), cursor="hand2",
        ).pack(side=LEFT, padx=(6, 0))
        Button(
            parent, text="최근 로그", font=("Segoe UI", 10),
            bg="#ced4da", fg="#333", relief="solid", borderwidth=1, padx=10, pady=5,
            command=lambda f=folder: self._download_latest_log(f), cursor="hand2",
        ).pack(side=LEFT, padx=(6, 0))

    def _download_latest_log(self, folder: Path):
        folder = Path(folder)
        files = list(folder.glob("*.log")) if folder.is_dir() else []
        if not files:
            messagebox.showwarning("안내", "다운로드할 로그가 없습니다. 먼저 실행하세요.")
            return
        src = max(files, key=lambda p: p.stat().st_mtime)
        dest = filedialog.asksaveasfilename(
            title="최근 로그 저장",
            initialfile=src.name,
            defaultextension=".log",
            filetypes=[("Log", "*.log"), ("Text", "*.txt"), ("All", "*.*")],
        )
        if not dest:
            return
        try:
            shutil.copy2(src, dest)
            messagebox.showinfo("다운로드", f"최근 로그 저장:\n{dest}")
        except Exception as e:
            messagebox.showerror("다운로드 실패", str(e))

    def _download_latest_results(self, folder: Path, suffixes=None):
        """가장 마지막에 생성된 결과 파일을 사용자가 고른 폴더로 복사."""
        folder = Path(folder)
        if not folder.is_dir():
            messagebox.showwarning("안내", "결과 폴더가 없습니다. 먼저 조회하세요.")
            return
        cands = []
        for f in folder.iterdir():
            if not f.is_file():
                continue
            if suffixes and not any(f.name.endswith(s) for s in suffixes):
                continue
            cands.append(f)
        if not cands:
            messagebox.showwarning("안내", "다운로드할 결과 파일이 없습니다. 먼저 조회하세요.")
            return
        newest = max(cands, key=lambda p: p.stat().st_mtime)
        files = [newest]
        if suffixes:
            for s in suffixes:
                if newest.name.endswith(s):
                    prefix = newest.name[: -len(s)]
                    files = [folder / (prefix + x) for x in suffixes if (folder / (prefix + x)).is_file()]
                    break
        dest = filedialog.askdirectory(title="저장할 폴더 선택")
        if not dest:
            return
        dest = Path(dest)
        names = []
        for src in files:
            try:
                shutil.copy2(src, dest / src.name)
                names.append(src.name)
            except Exception as e:
                messagebox.showerror("다운로드 실패", f"{src.name}\n{e}")
                return
        messagebox.showinfo("다운로드", "최근 결과 저장:\n" + "\n".join(names) + f"\n\n{dest}")
        try:
            os.startfile(str(dest.resolve()))
        except Exception:
            pass

    def _hint_result(self, parent, bg=None):
        Label(parent, text=RESULT_HINT, font=("Segoe UI", 8),
              fg="#888", bg=bg or CARD).pack(side=LEFT, padx=(8, 0))

    def _open_path(self, path: Path):
        path = Path(path)
        if path.exists() and path.is_file():
            target = path
        else:
            folder = path if path.suffix == "" else path.parent
            folder.mkdir(parents=True, exist_ok=True)
            target = folder
        try:
            os.startfile(str(target.resolve()))
        except Exception:
            messagebox.showinfo("경로", str(target.resolve()))

    def _open_oui_file(self):
        path = RESULTS_OUI / "oui.txt"
        if not path.exists():
            messagebox.showwarning("안내", "먼저 IEEE OUI를 다운로드하세요.")
            return
        webbrowser.open(path.resolve().as_uri())

    def _download_oui(self):
        def work():
            import re
            import requests
            import urllib3
            # 회사 SSL 가로채기(자체서명/사설 CA) 환경 대응
            urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
            RESULTS_OUI.mkdir(parents=True, exist_ok=True)
            headers = {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/120.0.0.0 Safari/537.36"
                ),
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.9",
            }
            # IEEE 봇 차단·회사 프록시 SSL 가로채기 대비 (verify=False)
            self._log_queue.put(
                "참고: 회사 SSL 검사 환경이면 인증서 검증을 건너뜁니다 (verify=False).\n"
            )
            urls = [
                "https://standards-oui.ieee.org/oui/oui.txt",
                "https://standards-oui.ieee.org/oui.txt",
                "http://standards-oui.ieee.org/oui/oui.txt",
                "https://standards-oui.ieee.org/",
            ]
            content = None
            used = None
            last_err = None
            for url in urls:
                try:
                    self._log_queue.put(f"시도: {url}\n")
                    r = requests.get(url, headers=headers, timeout=60, verify=False)
                    self._log_queue.put(f"  HTTP {r.status_code}, size={len(r.content)}\n")
                    if r.status_code == 200 and r.content and len(r.content) > 1000:
                        content = r.text
                        used = url
                        break
                    last_err = f"HTTP {r.status_code}"
                except Exception as e:
                    last_err = str(e)
                    self._log_queue.put(f"  오류: {e}\n")

            if not content:
                self._log_queue.put(f"실패: IEEE 다운로드 불가 ({last_err})\n")
                return

            self._log_queue.put(f"성공 소스: {used}\n파싱 중...\n")
            lines = []
            # 1) 공식 oui.txt 형식: "XX-XX-XX   (hex)\tVendor"
            for line in content.splitlines():
                if "(hex)" in line.lower():
                    # 예: AA-BB-CC   (hex)		Vendor Name
                    m = re.match(
                        r"^\s*([0-9A-Fa-f]{2}[-:][0-9A-Fa-f]{2}[-:][0-9A-Fa-f]{2})\s*\(hex\)\s*(.*)$",
                        line,
                    )
                    if m:
                        oui = m.group(1).replace("-", ":").upper()
                        vendor = m.group(2).strip()
                        lines.append(f"{oui}\t{vendor}")
                        continue
                    parts = re.split(r"\(hex\)", line, maxsplit=1, flags=re.IGNORECASE)
                    if len(parts) >= 2:
                        oui = parts[0].strip().replace("-", ":").upper()
                        vendor = parts[1].strip()
                        if oui and vendor:
                            lines.append(f"{oui}\t{vendor}")

            # 중복 제거 (순서 유지)
            seen = set()
            uniq = []
            for ln in lines:
                key = ln.split("\t", 1)[0]
                if key not in seen:
                    seen.add(key)
                    uniq.append(ln)
            lines = uniq

            # 최신본 하나만 유지 (oui.txt 덮어쓰기)
            txt_path = RESULTS_OUI / "oui.txt"
            header = (
                f"* {datetime.now().strftime('%Y/%m/%d %H:%M')} 업데이트\n"
                f"* source: {used}\n\n"
                "OUI\t\tVendor\n"
                + ("-" * 80) + "\n"
            )
            body = "\n".join(lines) + "\n" + ("-" * 80) + "\nEND\n"
            with open(txt_path, "w", encoding="utf-8") as f:
                f.write(header + body)
            self._log_queue.put(f"완료: {len(lines)} entries\n  {txt_path}\n")
            self._save_session_log(LOG_OUI, "oui")
            if len(lines) == 0:
                self._log_queue.put("경고: 파싱된 OUI가 0건입니다. 응답 형식이 변경되었을 수 있습니다.\n")

        threading.Thread(target=work, daemon=True).start()

    # ------------------------------------------------------------------
    # 로그 헬퍼
    # ------------------------------------------------------------------
    def _save_session_log(self, folder: Path, prefix: str):
        """실행 로그를 log/하위폴더에 저장하고 5개만 유지."""
        def _do():
            try:
                while True:
                    s = self._log_queue.get_nowait()
                    self._log(s)
            except queue.Empty:
                pass
            try:
                folder.mkdir(parents=True, exist_ok=True)
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                path = folder / f"{prefix}_{ts}.log"
                text = ""
                if hasattr(self, "log_text"):
                    text = self.log_text.get("1.0", "end-1c")
                path.write_text(text or "", encoding="utf-8")
                keep_latest_results(folder, keep=RESULT_KEEP)
                self._log(f"로그 저장: {path}\n")
            except Exception as e:
                try:
                    self._log(f"로그 저장 실패: {e}\n")
                except Exception:
                    pass
        self.after(0, _do)

    def _log(self, msg: str):
        if hasattr(self, "log_text"):
            self.log_text.insert(END, msg)
            self.log_text.see(END)

    def _drain_log(self):
        try:
            while True:
                s = self._log_queue.get_nowait()
                self._log(s)
        except queue.Empty:
            pass
        self.after(100, self._drain_log)


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
